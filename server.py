"""
Loan DPR & CMA Preparation Software - Backend API
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Cookie
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import jwt
import bcrypt
import io
import requests
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Literal, Any, Dict
from datetime import datetime, timezone, timedelta

try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore
    EMERGENT_LLM_AVAILABLE = True
except ImportError:
    LlmChat = None  # type: ignore
    UserMessage = None  # type: ignore
    EMERGENT_LLM_AVAILABLE = False

# Excel & PDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me')
JWT_ALGO = "HS256"
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

app = FastAPI(title="Loan DPR & CMA API")
api_router = APIRouter(prefix="/api")

# ============================== MODELS ==============================

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: EmailStr
    name: str
    picture: Optional[str] = None
    auth_provider: Literal["email", "google", "guest"] = "email"
    referral_code: str = ""
    referred_by: str = ""
    referral_credits: int = 0  # number of ₹500-discount credits available
    free_dpr_credits: int = 0  # number of free WATERMARKED DPR downloads (earned after first paid subscription via referral)
    wallet_balance: float = 0.0  # INR balance — usable to pay for DPRs
    is_guest: bool = False  # true for no-password "Quick Buy" accounts; pricing falls back to base_price
    is_admin: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    referral_code: Optional[str] = ""


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class AuthResponse(BaseModel):
    user: User
    token: Optional[str] = None


# Project Models
class PromoterInfo(BaseModel):
    name: str = ""
    age: Optional[int] = None
    qualification: str = ""
    experience: str = ""
    address: str = ""
    contact: str = ""
    pan: str = ""
    aadhaar: str = ""


class ApplicantInfo(BaseModel):
    """Detailed applicant info for the loan/DPR cover."""
    full_name: str = ""
    father_name: str = ""
    dob: str = ""
    gender: str = ""
    category: str = ""  # General / OBC / SC / ST
    aadhaar: str = ""
    pan: str = ""
    mobile: str = ""
    email: str = ""
    address_line1: str = ""
    address_line2: str = ""
    city: str = ""
    state: str = ""
    pincode: str = ""


class PreparedBy(BaseModel):
    name: str = ""
    designation: str = ""
    firm: str = ""
    contact: str = ""
    email: str = ""


class HistoricalYear(BaseModel):
    """Actuals for an existing business."""
    year_label: str = ""  # e.g. "FY 2023-24"
    revenue: float = 0
    raw_material: float = 0
    salaries: float = 0
    power_fuel: float = 0
    rent: float = 0
    other_expenses: float = 0
    depreciation: float = 0
    interest: float = 0


class CostItem(BaseModel):
    name: str = ""
    amount: float = 0


class YearProjection(BaseModel):
    year: int = 1
    revenue: float = 0
    raw_material: float = 0
    salaries: float = 0
    power_fuel: float = 0
    rent: float = 0
    other_expenses: float = 0
    depreciation: float = 0
    interest: float = 0
    tax_rate: float = 25


class FinanceMeans(BaseModel):
    promoter_contribution: float = 0
    term_loan: float = 0
    working_capital_loan: float = 0
    subsidy: float = 0
    other_sources: float = 0
    interest_rate: float = 11.0
    loan_tenure_years: int = 7
    moratorium_months: int = 6


class WorkingCapital(BaseModel):
    """Nayak / standard method inputs for WC assessment."""
    raw_material_days: int = 30
    finished_goods_days: int = 15
    receivables_days: int = 45
    payables_days: int = 30
    cash_required: float = 50000
    method: str = "Nayak"  # Nayak / MPBF / Turnover


class Collateral(BaseModel):
    type: str = ""  # "Land", "Building", "Plant & Machinery", "FD", "LIC", "Vehicle", "Gold", "Other"
    description: str = ""
    location: str = ""
    owner: str = ""
    market_value: float = 0
    realisable_value: float = 0


class UserIndustryTemplate(BaseModel):
    """User-defined editable industry template."""
    model_config = ConfigDict(extra="ignore")
    template_id: str = Field(default_factory=lambda: f"utpl_{uuid.uuid4().hex[:10]}")
    user_id: str
    name: str
    business_type: str = "Manufacturing"
    default_revenue: float = 5000000
    raw_material_pct: float = 0.50
    salaries_pct: float = 0.12
    power_fuel_pct: float = 0.06
    rent_pct: float = 0.04
    other_pct: float = 0.06
    depreciation_pct: float = 0.05
    tax_rate: float = 25
    rm_days: int = 30
    fg_days: int = 15
    recv_days: int = 45
    pay_days: int = 30
    cost_heads: List[CostItem] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class AINarrative(BaseModel):
    executive_summary: str = ""
    project_description: str = ""
    marketing_strategy: str = ""
    swot_strengths: List[str] = []
    swot_weaknesses: List[str] = []
    swot_opportunities: List[str] = []
    swot_threats: List[str] = []


class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    project_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    # Basic
    business_name: str = ""
    business_type: str = "Manufacturing"  # Manufacturing / Service / Trading
    industry: str = ""
    industry_template: str = ""  # e.g. textile_unit, food_processing OR user template id
    constitution: str = "Proprietorship"
    location: str = ""
    state: str = ""
    loan_scheme: str = "PMEGP"
    loan_scheme_custom_name: str = ""  # if scheme is custom or state-specific
    subsidy_pct: float = 0  # auto-applied subsidy %
    loan_amount: float = 0
    projection_years: int = 5
    # Business stage
    business_stage: str = "new"  # "new" or "existing"
    historical_years_count: int = 0  # 0 / 1 / 2 / 3
    base_year_index: int = -1  # index of historical year used as base for growth; -1 = latest
    growth_rate: float = 15.0  # annual growth %
    historical_actuals: List[HistoricalYear] = []
    # Applicant & prepared by
    applicant: ApplicantInfo = Field(default_factory=ApplicantInfo)
    prepared_by: PreparedBy = Field(default_factory=PreparedBy)
    # Details
    promoters: List[PromoterInfo] = []
    project_cost: List[CostItem] = []
    means_of_finance: FinanceMeans = Field(default_factory=FinanceMeans)
    projections: List[YearProjection] = []
    working_capital: WorkingCapital = Field(default_factory=WorkingCapital)
    collateral: List[Collateral] = []
    narrative: AINarrative = Field(default_factory=AINarrative)
    # Meta
    status: str = "draft"
    # Payment (manual UPI flow — admin can verify)
    payment_status: str = "unpaid"  # unpaid / submitted / paid
    payment_txn_id: str = ""
    payment_amount: float = 0
    payment_method: str = ""  # GPay / PhonePe / Paytm / BHIM / Bank
    paid_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ProjectCreate(BaseModel):
    business_name: str
    business_type: str = "Manufacturing"
    loan_scheme: str = "PMEGP"
    loan_scheme_custom_name: str = ""
    projection_years: int = 5
    business_stage: str = "new"
    industry_template: str = ""
    loan_amount: float = 0
    interest_rate: float = 11.0
    loan_tenure_years: int = 7
    moratorium_months: int = 6


class ProjectUpdate(BaseModel):
    business_name: Optional[str] = None
    business_type: Optional[str] = None
    industry: Optional[str] = None
    industry_template: Optional[str] = None
    constitution: Optional[str] = None
    location: Optional[str] = None
    state: Optional[str] = None
    loan_scheme: Optional[str] = None
    loan_scheme_custom_name: Optional[str] = None
    subsidy_pct: Optional[float] = None
    loan_amount: Optional[float] = None
    projection_years: Optional[int] = None
    business_stage: Optional[str] = None
    historical_years_count: Optional[int] = None
    base_year_index: Optional[int] = None
    growth_rate: Optional[float] = None
    historical_actuals: Optional[List[HistoricalYear]] = None
    applicant: Optional[ApplicantInfo] = None
    prepared_by: Optional[PreparedBy] = None
    promoters: Optional[List[PromoterInfo]] = None
    project_cost: Optional[List[CostItem]] = None
    means_of_finance: Optional[FinanceMeans] = None
    projections: Optional[List[YearProjection]] = None
    working_capital: Optional[WorkingCapital] = None
    collateral: Optional[List[Collateral]] = None
    narrative: Optional[AINarrative] = None
    status: Optional[str] = None


class PaymentSubmit(BaseModel):
    txn_id: str
    amount: float = 199
    method: str = "GPay"


class AutoProjectRequest(BaseModel):
    growth_rate: Optional[float] = None
    base_year_index: Optional[int] = None


# ============================== AUTH HELPERS ==============================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode(), hashed.encode())
    except Exception:
        return False


def create_jwt(user_id: str) -> str:
    payload = {
        "user_id": user_id,
        "jti": uuid.uuid4().hex,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def is_token_revoked(jti: Optional[str], user_id: str, iat: Optional[int]) -> bool:
    """Check if a JWT has been revoked. Two mechanisms: jti blocklist OR user-wide invalidation cutoff."""
    if jti:
        rev = await db.revoked_jtis.find_one({"jti": jti}, {"_id": 0})
        if rev:
            return True
    # User-wide invalidation: any token issued before this timestamp is invalid
    user_rev = await db.users.find_one({"user_id": user_id}, {"_id": 0, "tokens_invalidated_before": 1})
    if user_rev and user_rev.get("tokens_invalidated_before"):
        cutoff = user_rev["tokens_invalidated_before"]
        if isinstance(cutoff, str):
            try:
                dt = datetime.fromisoformat(cutoff)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                cutoff = dt.timestamp()
            except ValueError:
                cutoff = 0
        elif isinstance(cutoff, datetime):
            if cutoff.tzinfo is None:
                cutoff = cutoff.replace(tzinfo=timezone.utc)
            cutoff = cutoff.timestamp()
        if iat and iat < cutoff:
            return True
    return False


async def get_current_user(
    request: Request,
    session_token: Optional[str] = Cookie(None),
) -> User:
    # Try session_token cookie (Google OAuth)
    if session_token:
        session_doc = await db.user_sessions.find_one(
            {"session_token": session_token}, {"_id": 0}
        )
        if session_doc:
            expires_at = session_doc.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at and expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at and expires_at >= datetime.now(timezone.utc):
                user_doc = await db.users.find_one(
                    {"user_id": session_doc["user_id"]}, {"_id": 0}
                )
                if user_doc:
                    if isinstance(user_doc.get("created_at"), str):
                        user_doc["created_at"] = datetime.fromisoformat(user_doc["created_at"])
                    return User(**user_doc)

    # Try Authorization header (JWT)
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth.split(" ", 1)[1]
        # Try JWT first
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
            if await is_token_revoked(payload.get("jti"), payload["user_id"], payload.get("iat")):
                raise HTTPException(status_code=401, detail="Token revoked")
            user_doc = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
            if user_doc:
                if isinstance(user_doc.get("created_at"), str):
                    user_doc["created_at"] = datetime.fromisoformat(user_doc["created_at"])
                return User(**user_doc)
        except jwt.PyJWTError:
            pass
        # Try as session_token
        session_doc = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if session_doc:
            user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
            if user_doc:
                if isinstance(user_doc.get("created_at"), str):
                    user_doc["created_at"] = datetime.fromisoformat(user_doc["created_at"])
                return User(**user_doc)

    raise HTTPException(status_code=401, detail="Not authenticated")


# ============================== AUTH ROUTES ==============================

@api_router.post("/auth/register", response_model=AuthResponse)
async def register(req: RegisterRequest):
    existing = await db.users.find_one({"email": req.email.lower()}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    user_id = f"user_{uuid.uuid4().hex[:12]}"
    referral_code = f"MBDS-{uuid.uuid4().hex[:6].upper()}"
    now = datetime.now(timezone.utc)
    # Validate referral code (if provided) — does referrer exist?
    referred_by = ""
    if req.referral_code:
        rc = req.referral_code.strip().upper()
        ref_user = await db.users.find_one({"referral_code": rc}, {"_id": 0, "user_id": 1})
        if ref_user:
            referred_by = rc
            # Note: free_dpr_credits is awarded AFTER the user subscribes (makes first payment),
            # NOT immediately on signup. See submit_payment / pay_from_wallet.

    is_admin = req.email.lower() in ADMIN_EMAILS

    user_doc = {
        "user_id": user_id,
        "email": req.email.lower(),
        "name": req.name,
        "picture": None,
        "auth_provider": "email",
        "password_hash": hash_password(req.password),
        "referral_code": referral_code,
        "referred_by": referred_by,
        "referral_credits": 0,
        "free_dpr_credits": 0,
        "wallet_balance": 0.0,
        "is_guest": False,
        "is_admin": is_admin,
        "created_at": now.isoformat(),
    }
    await db.users.insert_one(user_doc)
    token = create_jwt(user_id)
    return AuthResponse(
        user=User(user_id=user_id, email=req.email.lower(), name=req.name,
                  auth_provider="email", referral_code=referral_code,
                  referred_by=referred_by, free_dpr_credits=0,
                  is_admin=is_admin, created_at=now),
        token=token,
    )


@api_router.post("/auth/login", response_model=AuthResponse)
async def login(req: LoginRequest):
    user_doc = await db.users.find_one({"email": req.email.lower()}, {"_id": 0})
    if not user_doc or not user_doc.get("password_hash"):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not verify_password(req.password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Auto-promote: if email is in ADMIN_EMAILS but user's is_admin is False, fix it.
    email_lower = user_doc.get("email", "").lower()
    if email_lower in ADMIN_EMAILS and not user_doc.get("is_admin"):
        await db.users.update_one({"user_id": user_doc["user_id"]}, {"$set": {"is_admin": True}})
        user_doc["is_admin"] = True

    if isinstance(user_doc.get("created_at"), str):
        user_doc["created_at"] = datetime.fromisoformat(user_doc["created_at"])
    token = create_jwt(user_doc["user_id"])
    return AuthResponse(user=User(**user_doc), token=token)


@api_router.post("/auth/google-session")
async def google_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")

    try:
        r = requests.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id},
            timeout=10,
        )
        if r.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session_id")
        data = r.json()
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Auth service error: {e}")

    email = data["email"].lower()
    user_doc = await db.users.find_one({"email": email}, {"_id": 0})
    now = datetime.now(timezone.utc)
    if not user_doc:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "email": email,
            "name": data.get("name", email),
            "picture": data.get("picture"),
            "auth_provider": "google",
            "created_at": now.isoformat(),
        }
        await db.users.insert_one(user_doc)
    else:
        user_id = user_doc["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": data.get("name", user_doc.get("name")),
                      "picture": data.get("picture")}}
        )

    session_token = data["session_token"]
    expires_at = now + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": now.isoformat(),
    })

    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60,
    )

    user_obj = {**user_doc, "user_id": user_id}
    if isinstance(user_obj.get("created_at"), str):
        user_obj["created_at"] = datetime.fromisoformat(user_obj["created_at"])
    user_obj.pop("password_hash", None)
    return {"user": User(**user_obj).model_dump(mode="json"), "session_token": session_token}


@api_router.get("/auth/me", response_model=User)
async def me(user: User = Depends(get_current_user)):
    # Auto-promote: if user's email is in ADMIN_EMAILS but DB has is_admin=False, fix it.
    if user.email.lower() in ADMIN_EMAILS and not user.is_admin:
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"is_admin": True}})
        user.is_admin = True
    return user


@api_router.post("/auth/logout")
async def logout(
    request: Request,
    response: Response,
    session_token: Optional[str] = Cookie(None),
):
    # Clear Google OAuth session
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})

    # Revoke Bearer JWT if present
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth.split(" ", 1)[1]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO], options={"verify_exp": False})
            jti = payload.get("jti")
            user_id = payload.get("user_id")
            exp = payload.get("exp")
            if jti:
                await db.revoked_jtis.insert_one({
                    "jti": jti,
                    "user_id": user_id,
                    "revoked_at": datetime.now(timezone.utc).isoformat(),
                    "exp": exp,
                })
            else:
                # legacy token w/o jti: invalidate all tokens for this user
                if user_id:
                    await db.users.update_one(
                        {"user_id": user_id},
                        {"$set": {"tokens_invalidated_before": datetime.now(timezone.utc).isoformat()}},
                    )
            # Also try as session_token (Google bearer)
            await db.user_sessions.delete_one({"session_token": token})
        except jwt.PyJWTError:
            await db.user_sessions.delete_one({"session_token": token})

    response.delete_cookie("session_token", path="/")
    return {"ok": True}


@api_router.post("/auth/logout-all")
async def logout_all(user: User = Depends(get_current_user)):
    """Revoke ALL tokens for the current user (every device)."""
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"tokens_invalidated_before": now_iso}},
    )
    await db.user_sessions.delete_many({"user_id": user.user_id})
    return {"ok": True}


# ============================== PROJECT ROUTES ==============================

LOAN_SCHEMES = [
    # Central / Government
    {"id": "PMEGP", "name": "PMEGP", "max_loan": 5000000, "category": "Government", "subsidy_pct": 25, "subsidy_note": "25% (urban) / 35% (rural / SC-ST / Women / NER)",
     "description": "Prime Minister's Employment Generation Programme. Up to ₹50 Lakh."},
    {"id": "MUDRA_SHISHU", "name": "Mudra - Shishu", "max_loan": 50000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Loan up to ₹50,000 for micro enterprises starting up."},
    {"id": "MUDRA_KISHORE", "name": "Mudra - Kishore", "max_loan": 500000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Loan ₹50,001 to ₹5 Lakh for growing micro enterprises."},
    {"id": "MUDRA_TARUN", "name": "Mudra - Tarun", "max_loan": 1000000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Loan ₹5 Lakh to ₹10 Lakh for established micro enterprises."},
    {"id": "STAND_UP_INDIA", "name": "Stand-Up India", "max_loan": 10000000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "Interest concession via DICGC; no upfront subsidy",
     "description": "Loan ₹10 Lakh to ₹1 Crore for SC/ST/Women entrepreneurs."},
    {"id": "CGTMSE", "name": "CGTMSE", "max_loan": 50000000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "Credit guarantee scheme (no upfront subsidy)",
     "description": "Credit Guarantee scheme for MSEs up to ₹5 Crore without collateral."},
    {"id": "PM_VISHWAKARMA", "name": "PM Vishwakarma", "max_loan": 300000, "category": "Government", "subsidy_pct": 0, "subsidy_note": "Toolkit incentive ₹15,000 + interest subvention",
     "description": "Loan + skill incentive for traditional craftsmen / artisans."},
    {"id": "PMFME", "name": "PMFME (Food Processing)", "max_loan": 1000000, "category": "Government", "subsidy_pct": 35, "subsidy_note": "35% credit-linked subsidy up to ₹10 Lakh",
     "description": "PM Formalisation of Micro Food Processing Enterprises."},
    # Bank-led
    {"id": "BANK_TERM_LOAN", "name": "Bank Term Loan", "max_loan": 1000000000, "category": "Bank", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Standard term loan from commercial banks."},
    {"id": "WORKING_CAPITAL", "name": "Working Capital", "max_loan": 1000000000, "category": "Bank", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Loan for day-to-day working capital needs."},
    {"id": "OD_CC", "name": "OD / Cash Credit", "max_loan": 1000000000, "category": "Bank", "subsidy_pct": 0, "subsidy_note": "",
     "description": "Overdraft / Cash Credit facility against stock & receivables."},
    # State schemes (selected major ones)
    {"id": "MMYSY_MP", "name": "Mukhyamantri Yuva Swarojgar (MP)", "max_loan": 5000000, "category": "State (Madhya Pradesh)", "subsidy_pct": 30, "subsidy_note": "30% general / 50% SC-ST-OBC-Women",
     "description": "MP State self-employment scheme for youth."},
    {"id": "MYSY_GUJ", "name": "Mukhyamantri Yuva Swavalamban (Gujarat)", "max_loan": 1000000, "category": "State (Gujarat)", "subsidy_pct": 100, "subsidy_note": "Up to 100% tuition/interest subsidy in select categories",
     "description": "Gujarat State scheme — supports tuition + business start-up."},
    {"id": "CMEGP_WB", "name": "CMEGP (West Bengal)", "max_loan": 2500000, "category": "State (West Bengal)", "subsidy_pct": 35, "subsidy_note": "15-35% margin money subsidy",
     "description": "Chief Minister's Employment Generation Programme — WB."},
    {"id": "MMRSK_BIHAR", "name": "Mukhyamantri Udyami Yojana (Bihar)", "max_loan": 1000000, "category": "State (Bihar)", "subsidy_pct": 50, "subsidy_note": "50% subsidy up to ₹5 Lakh + 50% interest-free loan",
     "description": "Bihar State entrepreneurship scheme."},
    {"id": "MSEFC_TN", "name": "NEEDS (Tamil Nadu)", "max_loan": 10000000, "category": "State (Tamil Nadu)", "subsidy_pct": 25, "subsidy_note": "25% capital subsidy",
     "description": "New Entrepreneur-cum-Enterprise Development Scheme — TN."},
    # Manual / custom catch-all
    {"id": "CUSTOM", "name": "Custom / Other Scheme", "max_loan": 1000000000, "category": "Custom", "subsidy_pct": 0, "subsidy_note": "Define manually",
     "description": "Enter your own scheme name and subsidy %."},
]


def _find_scheme(scheme_id: str) -> Optional[dict]:
    for s in LOAN_SCHEMES:
        if s["id"] == scheme_id:
            return s
    return None


@api_router.get("/loan-schemes")
async def get_loan_schemes():
    return LOAN_SCHEMES


# ============================== COMPANY / PAYMENT ==============================

COMPANY_INFO = {
    "name": "Mother Bless Digital Solutions",
    "address_line1": "Shop No. 32, Above State Bank ATM",
    "address_line2": "Sagwara Road",
    "city": "Bagidora",
    "state": "Rajasthan",
    "pincode": "314035",
    "country": "India",
    "phones": ["+91 7300213623", "+91 7877387012", "+91 7976578741"],
    "primary_phone": "7300213623",
    "whatsapp": "917300213623",
    "upi_id": "7300213623@okbizaxis",
    "upi_name": "Mother Bless Digital Solutions",
    "qr_image_url": "/payment-qr.jpg",
    "price_inr": 799,
    "referral_price_inr": 599,
    "referral_reward_text": "Refer a friend — when they sign up using your code, they get 1 free DPR (with DPRForge watermark). Their first paid DPR earns you a ₹200 wallet credit.",
    "tagline": "Loan DPR & CMA Reports made simple.",
}

ADMIN_EMAILS = [e.strip().lower() for e in os.environ.get("ADMIN_EMAILS", "motherblessopc@gmail.com,admin@motherbless.in,7300213623@motherbless.in").split(",") if e.strip()]


# ---- Persistent settings overlay (admin-editable) ----
_SETTINGS_ID = "company_settings"

# Fields the admin is allowed to change at runtime
_EDITABLE_KEYS = {
    "upi_id", "upi_name", "qr_image_url", "price_inr", "referral_price_inr",
    "payment_methods", "primary_phone", "whatsapp",
    "razorpay_key_id", "razorpay_key_secret", "razorpay_enabled",
}

# Default list of accepted payment methods (admin can edit)
COMPANY_INFO["payment_methods"] = ["GPay", "PhonePe", "Paytm", "BHIM UPI", "Bank Transfer"]


async def get_company_settings() -> dict:
    """Return COMPANY_INFO merged with any admin overrides stored in MongoDB."""
    merged = dict(COMPANY_INFO)
    doc = await db.settings.find_one({"_id": _SETTINGS_ID}, {"_id": 0})
    if doc:
        for k, v in doc.items():
            if k in _EDITABLE_KEYS and v is not None:
                merged[k] = v
    return merged


@api_router.get("/company")
async def get_company():
    return await get_company_settings()


@api_router.get("/projects/{project_id}/pricing")
async def get_pricing(project_id: str, user: User = Depends(get_current_user)):
    """Return the effective price for THIS user on THIS project.
    - Logged-in (non-guest) users always get the bulk/premium price (referral_price_inr).
    - Guest accounts (auth_provider='guest', is_guest=True) pay the base_price (799).
    Free watermarked credits are reported separately so the frontend can show 'free preview' option.
    """
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    credits = (user_doc or {}).get("referral_credits", 0)
    free_dpr = (user_doc or {}).get("free_dpr_credits", 0)
    wallet = float((user_doc or {}).get("wallet_balance", 0))
    is_guest = bool((user_doc or {}).get("is_guest", False))
    settings = await get_company_settings()
    base = int(settings["price_inr"])           # 799 — guest one-time
    bulk = int(settings["referral_price_inr"])  # 599 — logged-in bulk
    your_price = base if is_guest else bulk
    return {
        "base_price": base,
        "discounted_price": bulk,
        "your_price": your_price,
        "is_guest": is_guest,
        "has_referral_credit": credits > 0,
        "referral_credits": credits,
        "free_dpr_credits": free_dpr,
        "wallet_balance": wallet,
        "can_pay_from_wallet": wallet >= your_price,
    }


@api_router.post("/projects/{project_id}/submit-payment", response_model=Project)
async def submit_payment(project_id: str, payload: PaymentSubmit, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    if not payload.txn_id or len(payload.txn_id.strip()) < 4:
        raise HTTPException(status_code=400, detail="Valid transaction ID required")

    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    is_guest = bool((user_doc or {}).get("is_guest", False))
    has_credit = (user_doc or {}).get("referral_credits", 0) > 0
    settings = await get_company_settings()
    # Guests pay base (799); logged-in users pay bulk (599)
    expected_price = int(settings["price_inr"]) if is_guest else int(settings["referral_price_inr"])
    used_credit = False

    now = datetime.now(timezone.utc)
    update = {
        "payment_status": "paid",
        "payment_txn_id": payload.txn_id.strip(),
        "payment_amount": float(payload.amount),
        "payment_method": payload.method,
        "paid_at": now.isoformat(),
        "updated_at": now.isoformat(),
    }
    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id}, {"$set": update}
    )

    # First-time payment by this user → reward the referrer + give referee 1 free watermarked DPR.
    await _award_referral_if_first_paid(user_doc)

    # Audit log
    await db.payment_logs.insert_one({
        "log_id": str(uuid.uuid4()),
        "user_id": user.user_id,
        "user_email": (user_doc or {}).get("email"),
        "project_id": project_id,
        "txn_id": payload.txn_id.strip(),
        "amount": float(payload.amount),
        "expected_price": expected_price,
        "method": payload.method,
        "submitted_at": now.isoformat(),
        "verification_status": "pending",  # pending / verified / rejected
        "verified_by": None,
        "verified_at": None,
        "used_referral_credit": used_credit,
    })

    new_doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    return Project(**_serialize_project(new_doc))


# ============================== REFERRAL ==============================

@api_router.get("/referral/me")
async def my_referral(user: User = Depends(get_current_user)):
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    code = (user_doc or {}).get("referral_code", "")
    credits = (user_doc or {}).get("referral_credits", 0)
    # Count successful referrals
    referees = await db.users.find({"referred_by": code}, {"_id": 0, "email": 1, "name": 1, "created_at": 1, "referral_rewarded": 1}).to_list(200)
    successful = sum(1 for r in referees if r.get("referral_rewarded"))
    settings = await get_company_settings()
    return {
        "referral_code": code,
        "referral_credits": credits,
        "total_referrals": len(referees),
        "successful_referrals": successful,
        "discounted_price": int(settings["referral_price_inr"]),
        "base_price": int(settings["price_inr"]),
        "share_text": (
            f"Use DPRForge by Mother Bless Digital Solutions for your bank-ready Loan DPR & CMA. "
            f"Sign up with my code *{code}* and your first DPR is only ₹{int(settings['referral_price_inr'])}. "
            f"https://dprforge.com/register?ref={code}"
        ),
    }


# ============================== LOAN INQUIRY (Public) ==============================

class InquiryCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str = Field(..., min_length=2, max_length=120)
    mobile: str = Field(..., min_length=7, max_length=15)
    email: EmailStr
    loan_scheme: Optional[str] = ""
    loan_amount: Optional[float] = 0
    city: Optional[str] = ""
    state: Optional[str] = ""
    business_name: Optional[str] = ""
    message: Optional[str] = ""


@api_router.post("/inquiry")
async def submit_inquiry(payload: InquiryCreate, request: Request):
    """Public endpoint — accepts loan inquiries from landing page."""
    # Light spam guard
    mobile = "".join(c for c in payload.mobile if c.isdigit())
    if len(mobile) < 7:
        raise HTTPException(status_code=400, detail="Valid mobile number required")
    inquiry = {
        "inquiry_id": str(uuid.uuid4()),
        "name": payload.name.strip(),
        "mobile": mobile,
        "email": str(payload.email).lower(),
        "loan_scheme": (payload.loan_scheme or "").strip(),
        "loan_amount": float(payload.loan_amount or 0),
        "city": (payload.city or "").strip(),
        "state": (payload.state or "").strip(),
        "business_name": (payload.business_name or "").strip(),
        "message": (payload.message or "").strip(),
        "status": "new",  # new / contacted / converted / closed
        "ip": (request.client.host if request.client else ""),
        "user_agent": request.headers.get("user-agent", "")[:300],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.inquiries.insert_one(dict(inquiry))
    return {"ok": True, "inquiry_id": inquiry["inquiry_id"]}


# Admin-only inquiry endpoints are declared below, after `require_admin`.



# ============================== ADMIN ==============================

async def require_admin(user: User = Depends(get_current_user)) -> User:
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "is_admin": 1, "email": 1})
    if not user_doc:
        raise HTTPException(status_code=403, detail="Not admin")
    if user_doc.get("is_admin") or user_doc.get("email", "").lower() in ADMIN_EMAILS:
        return user
    raise HTTPException(status_code=403, detail="Admin access required")


@api_router.get("/admin/payments")
async def admin_payments(_: User = Depends(require_admin), status: Optional[str] = None):
    q = {}
    if status:
        q["verification_status"] = status
    logs = await db.payment_logs.find(q, {"_id": 0}).sort("submitted_at", -1).to_list(500)
    # Enrich each log with project + user info
    for log in logs:
        prj = await db.projects.find_one({"project_id": log.get("project_id")}, {"_id": 0, "business_name": 1, "loan_scheme": 1, "loan_amount": 1, "payment_status": 1})
        if prj:
            log["business_name"] = prj.get("business_name")
            log["loan_scheme"] = prj.get("loan_scheme")
            log["loan_amount"] = prj.get("loan_amount")
            log["current_payment_status"] = prj.get("payment_status")
    return logs


@api_router.post("/admin/payments/{log_id}/verify")
async def admin_verify(log_id: str, admin: User = Depends(require_admin)):
    log = await db.payment_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Payment log not found")
    now = datetime.now(timezone.utc).isoformat()
    await db.payment_logs.update_one(
        {"log_id": log_id},
        {"$set": {"verification_status": "verified", "verified_by": admin.email, "verified_at": now}},
    )
    return {"ok": True}


@api_router.post("/admin/payments/{log_id}/reject")
async def admin_reject(log_id: str, admin: User = Depends(require_admin)):
    log = await db.payment_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Payment log not found")
    now = datetime.now(timezone.utc).isoformat()
    await db.payment_logs.update_one(
        {"log_id": log_id},
        {"$set": {"verification_status": "rejected", "verified_by": admin.email, "verified_at": now}},
    )
    # Re-lock the project
    await db.projects.update_one(
        {"project_id": log["project_id"]},
        {"$set": {"payment_status": "unpaid", "updated_at": now}},
    )
    return {"ok": True}


@api_router.post("/admin/users/{user_id}/make-admin")
async def make_admin(user_id: str, _: User = Depends(require_admin)):
    r = await db.users.update_one({"user_id": user_id}, {"$set": {"is_admin": True}})
    return {"ok": r.modified_count > 0}


@api_router.get("/admin/stats")
async def admin_stats(_: User = Depends(require_admin)):
    total_users = await db.users.count_documents({})
    total_projects = await db.projects.count_documents({})
    total_paid = await db.projects.count_documents({"payment_status": "paid"})
    pending = await db.payment_logs.count_documents({"verification_status": "pending"})
    verified = await db.payment_logs.count_documents({"verification_status": "verified"})
    rejected = await db.payment_logs.count_documents({"verification_status": "rejected"})
    # Revenue
    logs = await db.payment_logs.find({"verification_status": {"$ne": "rejected"}}, {"_id": 0, "amount": 1}).to_list(10000)
    revenue = sum(float(x.get("amount") or 0) for x in logs)
    return {
        "users": total_users,
        "projects": total_projects,
        "paid_projects": total_paid,
        "payments_pending": pending,
        "payments_verified": verified,
        "payments_rejected": rejected,
        "revenue_inr": revenue,
    }


# ---------- Admin: Editable Settings (UPI, Price, Methods) ----------

class SettingsUpdate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    upi_id: Optional[str] = None
    upi_name: Optional[str] = None
    qr_image_url: Optional[str] = None
    price_inr: Optional[int] = None
    referral_price_inr: Optional[int] = None
    payment_methods: Optional[List[str]] = None
    primary_phone: Optional[str] = None
    whatsapp: Optional[str] = None
    razorpay_key_id: Optional[str] = None
    razorpay_key_secret: Optional[str] = None
    razorpay_enabled: Optional[bool] = None


@api_router.get("/admin/settings")
async def admin_get_settings(_: User = Depends(require_admin)):
    return await get_company_settings()


@api_router.post("/admin/settings")
async def admin_update_settings(payload: SettingsUpdate, admin: User = Depends(require_admin)):
    data = {k: v for k, v in payload.model_dump().items() if v is not None and k in _EDITABLE_KEYS}
    if not data:
        raise HTTPException(status_code=400, detail="Nothing to update")
    if "price_inr" in data and data["price_inr"] <= 0:
        raise HTTPException(status_code=400, detail="Price must be greater than 0")
    if "referral_price_inr" in data and data["referral_price_inr"] <= 0:
        raise HTTPException(status_code=400, detail="Referral price must be greater than 0")
    if "payment_methods" in data:
        # de-dupe and trim
        data["payment_methods"] = [m.strip() for m in data["payment_methods"] if m and m.strip()]
        if not data["payment_methods"]:
            raise HTTPException(status_code=400, detail="At least one payment method required")
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = admin.email
    await db.settings.update_one({"_id": _SETTINGS_ID}, {"$set": data}, upsert=True)
    return await get_company_settings()


# ---------- Admin: Loan Inquiries ----------

@api_router.get("/admin/inquiries")
async def admin_list_inquiries(_: User = Depends(require_admin), status: Optional[str] = None):
    q = {}
    if status:
        q["status"] = status
    rows = await db.inquiries.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return rows


@api_router.post("/admin/inquiries/{inquiry_id}/status")
async def admin_update_inquiry_status(inquiry_id: str, payload: Dict[str, Any], _: User = Depends(require_admin)):
    status = (payload.get("status") or "").strip().lower()
    if status not in {"new", "contacted", "converted", "closed"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    r = await db.inquiries.update_one(
        {"inquiry_id": inquiry_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inquiry not found")
    return {"ok": True}





def _ensure_paid(project: Project):
    if project.payment_status != "paid":
        raise HTTPException(
            status_code=402,
            detail="Payment required to download the report. Use Pay & Unlock on the preview page.",
        )


# ============================== INDUSTRY TEMPLATES ==============================
# Each template gives typical cost-structure ratios + scale-based defaults.
# Used to AUTO-GENERATE projections for a NEW business that has no historicals.
INDUSTRY_TEMPLATES = [
    {
        "id": "manufacturing_general",
        "name": "Manufacturing (General)",
        "icon": "factory",
        "default_revenue": 6000000,
        "capacity_utilization": [0.55, 0.70, 0.80, 0.85, 0.90, 0.92, 0.92],
        "raw_material_pct": 0.55, "salaries_pct": 0.10, "power_fuel_pct": 0.06,
        "rent_pct": 0.03, "other_pct": 0.06, "depreciation_pct": 0.05,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 45, "finished_goods_days": 20, "receivables_days": 45, "payables_days": 30},
        "default_cost_heads": [
            ("Land", 0), ("Building / Civil Works", 800000),
            ("Plant & Machinery", 2500000), ("Furniture & Fixtures", 200000),
            ("Pre-operative Expenses", 100000), ("Working Capital Margin", 200000),
            ("Contingency", 200000),
        ],
    },
    {
        "id": "textile_garment",
        "name": "Textile / Garment Unit",
        "icon": "shirt",
        "default_revenue": 8000000,
        "capacity_utilization": [0.50, 0.65, 0.75, 0.85, 0.90, 0.92, 0.92],
        "raw_material_pct": 0.60, "salaries_pct": 0.12, "power_fuel_pct": 0.05,
        "rent_pct": 0.04, "other_pct": 0.05, "depreciation_pct": 0.04,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 60, "finished_goods_days": 30, "receivables_days": 60, "payables_days": 30},
        "default_cost_heads": [
            ("Building / Civil Works", 600000), ("Stitching Machines", 1500000),
            ("Cutting / Pressing Equipment", 400000), ("Fabric Stock (Initial)", 600000),
            ("Furniture & Fixtures", 150000), ("Pre-operative Expenses", 100000),
            ("Working Capital Margin", 300000), ("Contingency", 150000),
        ],
    },
    {
        "id": "food_processing",
        "name": "Food Processing",
        "icon": "utensils",
        "default_revenue": 7500000,
        "capacity_utilization": [0.55, 0.70, 0.80, 0.85, 0.90, 0.92, 0.92],
        "raw_material_pct": 0.58, "salaries_pct": 0.10, "power_fuel_pct": 0.07,
        "rent_pct": 0.04, "other_pct": 0.06, "depreciation_pct": 0.05,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 30, "finished_goods_days": 15, "receivables_days": 30, "payables_days": 20},
        "default_cost_heads": [
            ("Building / Civil Works", 700000), ("Processing Machinery", 2000000),
            ("Cold Storage / Refrigeration", 500000), ("Packaging Equipment", 300000),
            ("Furniture & Fixtures", 100000), ("Pre-operative Expenses", 100000),
            ("Working Capital Margin", 250000), ("Contingency", 150000),
        ],
    },
    {
        "id": "restaurant_cafe",
        "name": "Restaurant / Cafe",
        "icon": "coffee",
        "default_revenue": 4800000,
        "capacity_utilization": [0.45, 0.65, 0.75, 0.82, 0.88, 0.90, 0.90],
        "raw_material_pct": 0.35, "salaries_pct": 0.22, "power_fuel_pct": 0.07,
        "rent_pct": 0.10, "other_pct": 0.08, "depreciation_pct": 0.05,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 7, "finished_goods_days": 1, "receivables_days": 5, "payables_days": 15},
        "default_cost_heads": [
            ("Interior / Civil Works", 800000), ("Kitchen Equipment", 600000),
            ("Furniture & Fixtures", 400000), ("POS / IT Systems", 80000),
            ("Initial Inventory", 100000), ("Pre-operative Expenses", 80000),
            ("Working Capital Margin", 150000), ("Contingency", 100000),
        ],
    },
    {
        "id": "retail_store",
        "name": "Retail / Trading Store",
        "icon": "shopping-cart",
        "default_revenue": 7200000,
        "capacity_utilization": [0.60, 0.75, 0.85, 0.90, 0.92, 0.92, 0.92],
        "raw_material_pct": 0.72, "salaries_pct": 0.07, "power_fuel_pct": 0.03,
        "rent_pct": 0.05, "other_pct": 0.04, "depreciation_pct": 0.02,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 40, "finished_goods_days": 0, "receivables_days": 15, "payables_days": 30},
        "default_cost_heads": [
            ("Shop Interior / Fitout", 500000), ("Furniture & Fixtures", 200000),
            ("POS / IT Systems", 80000), ("Initial Stock / Inventory", 1200000),
            ("Pre-operative Expenses", 50000), ("Working Capital Margin", 300000),
            ("Contingency", 100000),
        ],
    },
    {
        "id": "it_service",
        "name": "IT / Software Service",
        "icon": "monitor",
        "default_revenue": 9000000,
        "capacity_utilization": [0.50, 0.70, 0.85, 0.92, 0.95, 0.95, 0.95],
        "raw_material_pct": 0.05, "salaries_pct": 0.55, "power_fuel_pct": 0.03,
        "rent_pct": 0.08, "other_pct": 0.08, "depreciation_pct": 0.04,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 0, "finished_goods_days": 0, "receivables_days": 60, "payables_days": 30},
        "default_cost_heads": [
            ("Office Setup / Interior", 400000), ("Computers / Laptops", 800000),
            ("Servers / Networking", 300000), ("Furniture & Fixtures", 200000),
            ("Software Licenses", 150000), ("Pre-operative Expenses", 100000),
            ("Working Capital Margin", 400000), ("Contingency", 150000),
        ],
    },
    {
        "id": "dairy_farm",
        "name": "Dairy / Farm",
        "icon": "sprout",
        "default_revenue": 4500000,
        "capacity_utilization": [0.55, 0.75, 0.85, 0.90, 0.92, 0.92, 0.92],
        "raw_material_pct": 0.45, "salaries_pct": 0.12, "power_fuel_pct": 0.08,
        "rent_pct": 0.02, "other_pct": 0.10, "depreciation_pct": 0.06,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 15, "finished_goods_days": 3, "receivables_days": 10, "payables_days": 15},
        "default_cost_heads": [
            ("Shed / Civil Works", 500000), ("Livestock Purchase", 800000),
            ("Milking / Feed Equipment", 300000), ("Storage / Cooling", 200000),
            ("Feed Stock (Initial)", 150000), ("Pre-operative Expenses", 50000),
            ("Working Capital Margin", 150000), ("Contingency", 100000),
        ],
    },
    {
        "id": "service_general",
        "name": "Service Business (General)",
        "icon": "briefcase",
        "default_revenue": 5000000,
        "capacity_utilization": [0.50, 0.70, 0.82, 0.88, 0.92, 0.92, 0.92],
        "raw_material_pct": 0.15, "salaries_pct": 0.40, "power_fuel_pct": 0.05,
        "rent_pct": 0.08, "other_pct": 0.10, "depreciation_pct": 0.04,
        "tax_rate": 25,
        "wc_days": {"raw_material_days": 15, "finished_goods_days": 0, "receivables_days": 45, "payables_days": 30},
        "default_cost_heads": [
            ("Office Setup", 400000), ("Equipment / Tools", 600000),
            ("Furniture & Fixtures", 200000), ("IT Systems", 150000),
            ("Pre-operative Expenses", 80000), ("Working Capital Margin", 200000),
            ("Contingency", 100000),
        ],
    },
]


@api_router.get("/industry-templates")
async def get_industry_templates():
    """Return industry templates for the UI picker (excluding heavy detail)."""
    return [
        {"id": t["id"], "name": t["name"], "icon": t["icon"],
         "default_revenue": t["default_revenue"],
         "raw_material_pct": t["raw_material_pct"],
         "salaries_pct": t["salaries_pct"]}
        for t in INDUSTRY_TEMPLATES
    ]


def _find_template(template_id: str) -> Optional[dict]:
    for t in INDUSTRY_TEMPLATES:
        if t["id"] == template_id:
            return t
    return None


def _serialize_project(doc: dict) -> dict:
    """Convert ISO strings back to datetime for Pydantic."""
    for key in ("created_at", "updated_at"):
        if isinstance(doc.get(key), str):
            doc[key] = datetime.fromisoformat(doc[key])
    return doc


@api_router.get("/projects", response_model=List[Project])
async def list_projects(user: User = Depends(get_current_user)):
    docs = await db.projects.find({"user_id": user.user_id}, {"_id": 0}).sort("updated_at", -1).to_list(500)
    return [Project(**_serialize_project(d)) for d in docs]


@api_router.post("/projects", response_model=Project)
async def create_project(payload: ProjectCreate, user: User = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    years = payload.projection_years
    project = Project(
        user_id=user.user_id,
        business_name=payload.business_name,
        business_type=payload.business_type,
        loan_scheme=payload.loan_scheme,
        loan_scheme_custom_name=payload.loan_scheme_custom_name or "",
        loan_amount=payload.loan_amount or 0,
        projection_years=years,
        business_stage=payload.business_stage or "new",
        industry_template=payload.industry_template or "",
        projections=[YearProjection(year=i + 1) for i in range(years)],
    )
    # Apply finance defaults from payload (interest rate / tenure / moratorium)
    project.means_of_finance.interest_rate = float(payload.interest_rate or 11.0)
    project.means_of_finance.loan_tenure_years = int(payload.loan_tenure_years or 7)
    project.means_of_finance.moratorium_months = int(payload.moratorium_months or 6)
    if project.loan_amount and project.loan_amount > 0:
        project.means_of_finance.term_loan = float(project.loan_amount)

    # Auto-set subsidy % from scheme master
    scheme = _find_scheme(payload.loan_scheme)
    if scheme:
        project.subsidy_pct = float(scheme.get("subsidy_pct") or 0)

    # If user picked an industry template for a NEW business, auto-seed costs + projections + WC.
    if project.business_stage == "new" and payload.industry_template:
        tpl = _find_template(payload.industry_template) or await _find_user_template(user.user_id, payload.industry_template)
        if tpl:
            _apply_industry_template(project, tpl)

    # Apply auto-subsidy from %
    _autocompute_subsidy(project)
    # Auto-fill yearly interest from amortization schedule
    _apply_interest_to_projections(project)

    doc = project.model_dump()
    doc["created_at"] = now.isoformat()
    doc["updated_at"] = now.isoformat()
    await db.projects.insert_one(doc)
    return project


async def _find_user_template(user_id: str, template_id: str) -> Optional[dict]:
    """Look up a user-defined template by id, return dict in same shape as INDUSTRY_TEMPLATES."""
    doc = await db.user_templates.find_one(
        {"user_id": user_id, "template_id": template_id}, {"_id": 0}
    )
    if not doc:
        return None
    cost_heads = doc.get("cost_heads") or []
    return {
        "id": doc["template_id"],
        "name": doc["name"],
        "default_revenue": doc.get("default_revenue") or 5000000,
        "capacity_utilization": [0.55, 0.70, 0.80, 0.85, 0.90, 0.92, 0.92],
        "raw_material_pct": doc.get("raw_material_pct") or 0.5,
        "salaries_pct": doc.get("salaries_pct") or 0.12,
        "power_fuel_pct": doc.get("power_fuel_pct") or 0.06,
        "rent_pct": doc.get("rent_pct") or 0.04,
        "other_pct": doc.get("other_pct") or 0.06,
        "depreciation_pct": doc.get("depreciation_pct") or 0.05,
        "tax_rate": doc.get("tax_rate") or 25,
        "wc_days": {
            "raw_material_days": doc.get("rm_days") or 30,
            "finished_goods_days": doc.get("fg_days") or 15,
            "receivables_days": doc.get("recv_days") or 45,
            "payables_days": doc.get("pay_days") or 30,
        },
        "default_cost_heads": [(c["name"], float(c["amount"])) for c in cost_heads],
    }


def _apply_industry_template(project: Project, tpl: dict):
    """Seed project_cost, working_capital, industry name & projections based on template.
    If project.loan_amount > 0, scale the default cost-heads so that the TOTAL roughly equals
    (loan_amount + reasonable promoter contribution) — i.e. the project cost fits the user's loan ask.
    """
    project.industry = project.industry or tpl["name"]
    base_cost_heads = [(n, float(a)) for (n, a) in tpl["default_cost_heads"]]
    base_total = sum(a for _, a in base_cost_heads) or 1.0

    # Determine target project cost.
    # Convention: project_cost ≈ loan_amount / 0.75 (i.e. bank funds ~75%, promoter ~25%).
    if project.loan_amount and project.loan_amount > 0:
        target_total = max(project.loan_amount / 0.75, project.loan_amount)
        scale = target_total / base_total
    else:
        scale = 1.0
    project.project_cost = [CostItem(name=n, amount=round(a * scale, 0)) for (n, a) in base_cost_heads]

    project.working_capital = WorkingCapital(**{**tpl["wc_days"], "cash_required": 50000, "method": "Nayak"})

    rev_base = float(tpl["default_revenue"]) * scale
    projections = []
    util = tpl["capacity_utilization"]
    for i in range(project.projection_years):
        u = util[i] if i < len(util) else util[-1]
        revenue = rev_base * u
        projections.append(YearProjection(
            year=i + 1,
            revenue=round(revenue, 2),
            raw_material=round(revenue * tpl["raw_material_pct"], 2),
            salaries=round(revenue * tpl["salaries_pct"], 2),
            power_fuel=round(revenue * tpl["power_fuel_pct"], 2),
            rent=round(revenue * tpl["rent_pct"], 2),
            other_expenses=round(revenue * tpl["other_pct"], 2),
            depreciation=round(revenue * tpl["depreciation_pct"], 2),
            interest=0,
            tax_rate=tpl["tax_rate"],
        ))
    project.projections = projections


def _autocompute_subsidy(project: Project):
    """Auto-set subsidy in means_of_finance from scheme % * loan_amount."""
    if project.subsidy_pct and project.loan_amount:
        subsidy_amt = round(project.loan_amount * project.subsidy_pct / 100, 0)
        project.means_of_finance.subsidy = subsidy_amt


async def _award_referral_if_first_paid(user_doc: dict) -> None:
    """If this is the user's first paid project AND they were referred,
    credit the referrer ₹200 wallet, give the referee 1 free watermarked DPR,
    and mark `referral_rewarded=True` to make it one-time. Idempotent.
    """
    if not user_doc:
        return
    if user_doc.get("referral_rewarded"):
        return
    referred_by = (user_doc or {}).get("referred_by") or ""
    if not referred_by:
        return
    paid_count = await db.projects.count_documents({
        "user_id": user_doc.get("user_id"), "payment_status": "paid"
    })
    if paid_count != 1:
        return
    # Credit the referrer
    await db.users.update_one(
        {"referral_code": referred_by},
        {"$inc": {"wallet_balance": 200.0}},
    )
    # Reward referee with 1 free watermarked DPR + mark rewarded
    await db.users.update_one(
        {"user_id": user_doc.get("user_id")},
        {"$set": {"referral_rewarded": True}, "$inc": {"free_dpr_credits": 1}},
    )


def _yearly_interest_schedule(loan_amount: float, annual_rate: float,
                              tenure_years: int, moratorium_months: int,
                              projection_years: int) -> List[float]:
    """Standard EMI-based amortization. Returns the yearly TOTAL interest paid
    for each projection year (length = projection_years).
    During moratorium, interest is accrued and PAID but no principal repayment.
    After moratorium, regular EMIs cover both interest + principal.
    """
    out = [0.0] * max(1, projection_years)
    if loan_amount <= 0 or annual_rate <= 0 or tenure_years <= 0:
        return out
    r = (annual_rate / 100.0) / 12.0  # monthly rate
    n = max(1, tenure_years * 12)  # tenure in months
    morat = max(0, int(moratorium_months or 0))
    # During moratorium → no principal repayment, interest = loan * r each month
    # After moratorium, n_remaining months of EMI on (still full) principal
    n_emi = max(1, n - morat)
    emi = loan_amount * r * ((1 + r) ** n_emi) / (((1 + r) ** n_emi) - 1) if r > 0 else (loan_amount / n_emi)

    outstanding = loan_amount
    total_months = projection_years * 12
    for m in range(1, total_months + 1):
        year_idx = (m - 1) // 12
        if year_idx >= projection_years:
            break
        if m <= morat:
            interest_this_month = outstanding * r
            # No principal repayment during moratorium
        else:
            if outstanding <= 0:
                continue
            interest_this_month = outstanding * r
            principal_this_month = max(0.0, emi - interest_this_month)
            outstanding = max(0.0, outstanding - principal_this_month)
        out[year_idx] += round(interest_this_month, 2)
    return [round(x, 2) for x in out]


def _apply_interest_to_projections(project: Project):
    """Recalculate the per-year `interest` value for each projection from the
    loan amount/rate/tenure in means_of_finance. Skips if no loan_amount.
    """
    rate = float(project.means_of_finance.interest_rate or 0)
    tenure = int(project.means_of_finance.loan_tenure_years or 0)
    morat = int(project.means_of_finance.moratorium_months or 0)
    years = project.projection_years or len(project.projections) or 5
    schedule = _yearly_interest_schedule(
        project.loan_amount or 0.0, rate, tenure, morat, years
    )
    for i, p in enumerate(project.projections):
        if i < len(schedule):
            p.interest = schedule[i]


@api_router.post("/projects/{project_id}/auto-project", response_model=Project)
async def auto_project(project_id: str, payload: AutoProjectRequest, user: User = Depends(get_current_user)):
    """Auto-generate projections.
    - NEW business: uses selected industry_template (with capacity utilisation ramp)
    - EXISTING business: uses historical_actuals + growth_rate from the chosen base year
    """
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    project = Project(**_serialize_project(doc))

    if payload.growth_rate is not None:
        project.growth_rate = float(payload.growth_rate)
    if payload.base_year_index is not None:
        project.base_year_index = int(payload.base_year_index)

    if project.business_stage == "existing":
        actuals = project.historical_actuals or []
        if not actuals:
            raise HTTPException(status_code=400, detail="No historical actuals provided for existing business")
        idx = project.base_year_index
        if idx < 0 or idx >= len(actuals):
            idx = len(actuals) - 1  # latest year
        base = actuals[idx]
        g = (project.growth_rate or 15) / 100.0
        projections = []
        for i in range(project.projection_years):
            mul = (1 + g) ** (i + 1)
            projections.append(YearProjection(
                year=i + 1,
                revenue=round(base.revenue * mul, 2),
                raw_material=round(base.raw_material * mul, 2),
                salaries=round(base.salaries * mul, 2),
                power_fuel=round(base.power_fuel * mul, 2),
                rent=round(base.rent * mul, 2),
                other_expenses=round(base.other_expenses * mul, 2),
                depreciation=round(base.depreciation * mul, 2),
                interest=round(base.interest * mul, 2),
                tax_rate=25,
            ))
        project.projections = projections
    else:
        tpl_id = project.industry_template
        tpl = _find_template(tpl_id) if tpl_id else None
        if not tpl:
            raise HTTPException(status_code=400, detail="Select an industry template first for a new business")
        _apply_industry_template(project, tpl)

    # Apply yearly interest auto-calculation from loan amount × rate × tenure
    _apply_interest_to_projections(project)

    update_doc = project.model_dump()
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_doc.pop("created_at", None)
    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id},
        {"$set": update_doc},
    )
    new_doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    return Project(**_serialize_project(new_doc))


@api_router.post("/projects/{project_id}/duplicate", response_model=Project)
async def duplicate_project(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    now = datetime.now(timezone.utc)
    new_id = str(uuid.uuid4())
    doc["project_id"] = new_id
    doc["business_name"] = (doc.get("business_name") or "Untitled") + " (Copy)"
    doc["status"] = "draft"
    doc["created_at"] = now.isoformat()
    doc["updated_at"] = now.isoformat()
    await db.projects.insert_one(doc)
    return Project(**_serialize_project(doc))


@api_router.post("/projects/{project_id}/scale-to-loan", response_model=Project)
async def scale_to_loan(project_id: str, user: User = Depends(get_current_user)):
    """Scale all project_cost amounts so that the total ≈ loan_amount / 0.75 (75/25 bank/promoter split)."""
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    project = Project(**_serialize_project(doc))
    if not project.loan_amount or project.loan_amount <= 0:
        raise HTTPException(status_code=400, detail="Set loan_amount first")
    current_total = sum(c.amount for c in project.project_cost) or 1
    target = max(project.loan_amount / 0.75, project.loan_amount)
    factor = target / current_total
    project.project_cost = [CostItem(name=c.name, amount=round(c.amount * factor, 0)) for c in project.project_cost]
    update_doc = project.model_dump()
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_doc.pop("created_at", None)
    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id},
        {"$set": update_doc},
    )
    return project


# ============================== USER INDUSTRY TEMPLATES ==============================

@api_router.get("/user-templates")
async def list_user_templates(user: User = Depends(get_current_user)):
    docs = await db.user_templates.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for d in docs:
        if isinstance(d.get("created_at"), str):
            d["created_at"] = datetime.fromisoformat(d["created_at"])
    return docs


class UserTemplateCreate(BaseModel):
    name: str
    business_type: str = "Manufacturing"
    default_revenue: float = 5000000
    raw_material_pct: float = 0.50
    salaries_pct: float = 0.12
    power_fuel_pct: float = 0.06
    rent_pct: float = 0.04
    other_pct: float = 0.06
    depreciation_pct: float = 0.05
    tax_rate: float = 25
    rm_days: int = 30
    fg_days: int = 15
    recv_days: int = 45
    pay_days: int = 30
    cost_heads: List[CostItem] = []


@api_router.post("/user-templates", response_model=UserIndustryTemplate)
async def create_user_template(payload: UserTemplateCreate, user: User = Depends(get_current_user)):
    tpl = UserIndustryTemplate(user_id=user.user_id, **payload.model_dump())
    doc = tpl.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.user_templates.insert_one(doc)
    return tpl


@api_router.put("/user-templates/{template_id}", response_model=UserIndustryTemplate)
async def update_user_template(template_id: str, payload: UserTemplateCreate, user: User = Depends(get_current_user)):
    doc = await db.user_templates.find_one({"template_id": template_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Template not found")
    update_data = payload.model_dump()
    await db.user_templates.update_one(
        {"template_id": template_id, "user_id": user.user_id},
        {"$set": update_data},
    )
    new_doc = await db.user_templates.find_one({"template_id": template_id, "user_id": user.user_id}, {"_id": 0})
    if isinstance(new_doc.get("created_at"), str):
        new_doc["created_at"] = datetime.fromisoformat(new_doc["created_at"])
    return UserIndustryTemplate(**new_doc)


@api_router.delete("/user-templates/{template_id}")
async def delete_user_template(template_id: str, user: User = Depends(get_current_user)):
    result = await db.user_templates.delete_one({"template_id": template_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"ok": True}


@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    return Project(**_serialize_project(doc))


@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, payload: ProjectUpdate, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    update_data = payload.model_dump(exclude_unset=True)
    # If projection_years changed, resize projections list
    if "projection_years" in update_data and "projections" not in update_data:
        years = update_data["projection_years"]
        existing = doc.get("projections", [])
        if len(existing) < years:
            for i in range(len(existing), years):
                existing.append(YearProjection(year=i + 1).model_dump())
        else:
            existing = existing[:years]
        update_data["projections"] = existing

    # If scheme changed, auto-pick subsidy_pct from master (unless explicitly overridden)
    if "loan_scheme" in update_data and "subsidy_pct" not in update_data:
        scheme = _find_scheme(update_data["loan_scheme"])
        if scheme:
            update_data["subsidy_pct"] = float(scheme.get("subsidy_pct") or 0)

    # Auto-compute subsidy amount if subsidy_pct OR loan_amount changed
    merged = {**doc, **update_data}
    pct = merged.get("subsidy_pct") or 0
    loan_amt = merged.get("loan_amount") or 0
    if pct and loan_amt and "means_of_finance" not in update_data:
        mof = merged.get("means_of_finance") or {}
        if isinstance(mof, dict):
            mof = {**mof, "subsidy": round(loan_amt * pct / 100, 0)}
            update_data["means_of_finance"] = mof

    # Auto-populate interest in each projection year using proper amortization schedule.
    # Triggered when loan_amount, means_of_finance, or projection_years changes.
    triggers = {"means_of_finance", "loan_amount", "projection_years", "projections"}
    if triggers & set(update_data.keys()) and "projections" in {*update_data.keys(), *merged.keys()}:
        mof = merged.get("means_of_finance") or {}
        if isinstance(mof, dict):
            tl = float(mof.get("term_loan") or merged.get("loan_amount") or 0)
            rate = float(mof.get("interest_rate") or 11.0)
            tenure = int(mof.get("loan_tenure_years") or 7)
            morat = int(mof.get("moratorium_months") or 0)
            years = int(merged.get("projection_years") or len(merged.get("projections") or []) or 5)
            projections = update_data.get("projections") or merged.get("projections") or []
            if tl > 0 and rate > 0:
                schedule = _yearly_interest_schedule(tl, rate, tenure, morat, years)
                updated_projections = []
                for i, pr in enumerate(projections):
                    pr_copy = dict(pr) if isinstance(pr, dict) else pr
                    if i < len(schedule):
                        pr_copy["interest"] = schedule[i]
                    updated_projections.append(pr_copy)
                update_data["projections"] = updated_projections

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id},
        {"$set": update_data},
    )
    new_doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    return Project(**_serialize_project(new_doc))


@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: User = Depends(get_current_user)):
    result = await db.projects.delete_one({"project_id": project_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"ok": True}


# ============================== AI NARRATIVE ==============================

@api_router.post("/projects/{project_id}/generate-narrative", response_model=AINarrative)
async def generate_narrative(project_id: str, user: User = Depends(get_current_user)):
    if not EMERGENT_LLM_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="AI narrative generation is not configured on this deployment. Install 'emergentintegrations' to enable.",
        )
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    project = Project(**_serialize_project(doc))

    total_cost = sum(c.amount for c in project.project_cost) or project.loan_amount
    year1 = project.projections[0] if project.projections else YearProjection()
    promoter_names = ", ".join([p.name for p in project.promoters if p.name]) or "the promoter"

    prompt = f"""You are an expert financial consultant preparing a Detailed Project Report (DPR) for a loan application in India.

Generate professional, bank-ready narrative content for this project. Return ONLY valid JSON, no markdown, no commentary.

Project Details:
- Business Name: {project.business_name}
- Industry: {project.industry or project.business_type}
- Business Type: {project.business_type}
- Constitution: {project.constitution}
- Location: {project.location}, {project.state}
- Loan Scheme: {project.loan_scheme}
- Total Project Cost: INR {total_cost:,.0f}
- Loan Required: INR {project.loan_amount:,.0f}
- Promoter(s): {promoter_names}
- Year 1 Revenue Projection: INR {year1.revenue:,.0f}

Return JSON with this exact structure:
{{
  "executive_summary": "3-4 paragraph professional executive summary suitable for bank submission",
  "project_description": "2-3 paragraph detailed description of the business, products/services, and operations",
  "marketing_strategy": "2 paragraph marketing & sales strategy including target market, pricing, distribution",
  "swot_strengths": ["4-5 specific strengths"],
  "swot_weaknesses": ["3-4 specific weaknesses"],
  "swot_opportunities": ["4-5 specific opportunities"],
  "swot_threats": ["3-4 specific threats"]
}}
"""

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"narrative-{project_id}",
        system_message="You are an expert Chartered Accountant and DPR writer for Indian MSME loans. Output strictly valid JSON only.",
    ).with_model("anthropic", "claude-sonnet-4-5-20250929")

    response_text = await chat.send_message(UserMessage(text=prompt))

    # Parse JSON from response
    import json
    import re
    cleaned = response_text.strip()
    # Remove ```json fences if present
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    # Find first { and last }
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start != -1 and end != -1:
        cleaned = cleaned[start:end + 1]

    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="AI returned invalid JSON")

    narrative = AINarrative(
        executive_summary=data.get("executive_summary", ""),
        project_description=data.get("project_description", ""),
        marketing_strategy=data.get("marketing_strategy", ""),
        swot_strengths=data.get("swot_strengths", []),
        swot_weaknesses=data.get("swot_weaknesses", []),
        swot_opportunities=data.get("swot_opportunities", []),
        swot_threats=data.get("swot_threats", []),
    )

    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id},
        {"$set": {"narrative": narrative.model_dump(),
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    return narrative


# ============================== COMPUTATIONS ==============================

def compute_emi(principal: float, rate_annual: float, tenure_years: int) -> float:
    if principal <= 0 or tenure_years <= 0:
        return 0
    r = rate_annual / 100 / 12
    n = tenure_years * 12
    if r == 0:
        return principal / n
    return principal * r * (1 + r) ** n / ((1 + r) ** n - 1)


def compute_year_summary(p: Project) -> List[Dict[str, Any]]:
    """Compute P&L summary per year."""
    results = []
    annual_interest = p.means_of_finance.term_loan * (p.means_of_finance.interest_rate / 100)
    for proj in p.projections:
        total_expenses = (proj.raw_material + proj.salaries + proj.power_fuel
                          + proj.rent + proj.other_expenses)
        ebitda = proj.revenue - total_expenses
        # Use stored interest if non-zero, else estimated annual interest
        interest = proj.interest if proj.interest > 0 else annual_interest
        ebt = ebitda - proj.depreciation - interest
        tax = max(0, ebt) * (proj.tax_rate / 100)
        net_profit = ebt - tax
        # DSCR = (Net Profit + Depreciation + Interest) / (Interest + Principal Repayment)
        principal_repay = p.means_of_finance.term_loan / max(p.means_of_finance.loan_tenure_years, 1)
        denom = interest + principal_repay
        dscr = ((net_profit + proj.depreciation + interest) / denom) if denom > 0 else 0
        results.append({
            "year": proj.year,
            "revenue": proj.revenue,
            "total_expenses": total_expenses,
            "ebitda": ebitda,
            "depreciation": proj.depreciation,
            "interest": interest,
            "ebt": ebt,
            "tax": tax,
            "net_profit": net_profit,
            "dscr": dscr,
        })
    return results


@api_router.get("/projects/{project_id}/summary")
async def project_summary(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    project = Project(**_serialize_project(doc))
    total_cost = sum(c.amount for c in project.project_cost)
    total_finance = (project.means_of_finance.promoter_contribution
                     + project.means_of_finance.term_loan
                     + project.means_of_finance.working_capital_loan
                     + project.means_of_finance.subsidy
                     + project.means_of_finance.other_sources)
    emi = compute_emi(project.means_of_finance.term_loan,
                      project.means_of_finance.interest_rate,
                      project.means_of_finance.loan_tenure_years)
    years_data = compute_year_summary(project)
    avg_dscr = (sum(y["dscr"] for y in years_data) / len(years_data)) if years_data else 0
    return {
        "total_project_cost": total_cost,
        "total_means_of_finance": total_finance,
        "gap": total_cost - total_finance,
        "monthly_emi": emi,
        "annual_emi": emi * 12,
        "yearly_pl": years_data,
        "average_dscr": avg_dscr,
    }


def compute_working_capital(p: Project) -> Dict[str, Any]:
    """Compute year-wise working capital requirement based on days held.
    WC = (RM*RM_days + FG*FG_days + Receivables*Recv_days)/365 - Payables*Payables_days/365 + cash
    Using Nayak method: bank can fund 75% of net WC gap; 25% to be promoter's margin.
    """
    wc = p.working_capital
    rows = []
    for proj in p.projections:
        rm_held = (proj.raw_material * wc.raw_material_days) / 365
        fg_held = ((proj.raw_material + proj.salaries + proj.power_fuel) * wc.finished_goods_days) / 365
        recv = (proj.revenue * wc.receivables_days) / 365
        payables = (proj.raw_material * wc.payables_days) / 365
        gross_wc = rm_held + fg_held + recv + wc.cash_required
        net_wc = gross_wc - payables
        nayak_bank = net_wc * 0.75
        nayak_margin = net_wc * 0.25
        rows.append({
            "year": proj.year,
            "raw_material_stock": round(rm_held, 0),
            "finished_goods_stock": round(fg_held, 0),
            "receivables": round(recv, 0),
            "cash": wc.cash_required,
            "gross_wc": round(gross_wc, 0),
            "payables": round(payables, 0),
            "net_wc": round(net_wc, 0),
            "bank_finance_75": round(nayak_bank, 0),
            "promoter_margin_25": round(nayak_margin, 0),
        })
    return {"method": wc.method, "rows": rows}


def compute_balance_sheet(p: Project) -> Dict[str, Any]:
    """Compute simple projected Balance Sheet year-wise.
    Liabilities: Capital + Reserves (Retained Earnings) + Term Loan O/S + WC Loan + Sundry Creditors
    Assets:      Fixed Assets (net of depreciation) + Current Assets (Stocks + Receivables + Cash)
    """
    rows = []
    fixed_assets_gross = sum(c.amount for c in p.project_cost if c.name and "working capital" not in c.name.lower())
    accumulated_dep = 0
    retained_earnings = 0
    # Initial principal outstanding
    tl_principal = p.means_of_finance.term_loan
    annual_principal = tl_principal / max(p.means_of_finance.loan_tenure_years, 1) if tl_principal else 0
    wc_loan = p.means_of_finance.working_capital_loan
    capital = p.means_of_finance.promoter_contribution
    subsidy = p.means_of_finance.subsidy

    summary = compute_year_summary(p)
    wc_data = compute_working_capital(p)["rows"]

    cash_balance = capital + subsidy + p.means_of_finance.other_sources - sum(c.amount for c in p.project_cost)
    cash_balance = max(0, cash_balance)

    for i, proj in enumerate(p.projections):
        accumulated_dep += proj.depreciation
        net_fixed_assets = max(0, fixed_assets_gross - accumulated_dep)
        net_profit = summary[i]["net_profit"]
        retained_earnings += net_profit
        # Reduce term loan principal year-on-year (post moratorium)
        moratorium_year = (p.means_of_finance.moratorium_months or 0) / 12
        if (i + 1) > moratorium_year:
            tl_principal = max(0, tl_principal - annual_principal)
        wc_row = wc_data[i] if i < len(wc_data) else {"raw_material_stock": 0, "finished_goods_stock": 0, "receivables": 0, "cash": 0, "payables": 0}
        current_assets_inventory = wc_row["raw_material_stock"] + wc_row["finished_goods_stock"]
        receivables = wc_row["receivables"]
        cash_balance = cash_balance + net_profit + proj.depreciation - annual_principal
        cash_balance = max(0, cash_balance)
        current_assets = current_assets_inventory + receivables + cash_balance
        current_liabilities = wc_row["payables"] + wc_loan

        total_liabilities = capital + subsidy + retained_earnings + tl_principal + current_liabilities
        total_assets = net_fixed_assets + current_assets
        rows.append({
            "year": proj.year,
            "capital": round(capital, 0),
            "subsidy": round(subsidy, 0),
            "retained_earnings": round(retained_earnings, 0),
            "term_loan_outstanding": round(tl_principal, 0),
            "wc_loan": round(wc_loan, 0),
            "sundry_creditors": round(wc_row["payables"], 0),
            "total_liabilities": round(total_liabilities, 0),
            "net_fixed_assets": round(net_fixed_assets, 0),
            "inventory": round(current_assets_inventory, 0),
            "receivables": round(receivables, 0),
            "cash": round(cash_balance, 0),
            "total_assets": round(total_assets, 0),
        })
    return {"rows": rows}


@api_router.get("/projects/{project_id}/working-capital")
async def working_capital_endpoint(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    return compute_working_capital(Project(**_serialize_project(doc)))


@api_router.get("/projects/{project_id}/balance-sheet")
async def balance_sheet_endpoint(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    return compute_balance_sheet(Project(**_serialize_project(doc)))


def compute_ratios(p: Project) -> Dict[str, Any]:
    """Compute key financial ratios per projected year. Returns list of dicts + averages."""
    summary = compute_year_summary(p)
    wc_data = compute_working_capital(p)["rows"]
    bs_rows = compute_balance_sheet(p)["rows"]
    rows = []
    for i, yr in enumerate(p.projections):
        s = summary[i] if i < len(summary) else {}
        wc = wc_data[i] if i < len(wc_data) else {}
        bs = bs_rows[i] if i < len(bs_rows) else {}
        revenue = s.get("revenue", 0)
        net_profit = s.get("net_profit", 0)
        ebitda = s.get("ebitda", 0)
        interest = s.get("interest", 0) or 0.01
        # Balance sheet figures
        equity = (bs.get("capital", 0) + bs.get("subsidy", 0) + bs.get("retained_earnings", 0))
        debt = bs.get("term_loan_outstanding", 0) + bs.get("wc_loan", 0)
        current_assets = bs.get("inventory", 0) + bs.get("receivables", 0) + bs.get("cash", 0)
        current_liabs = bs.get("sundry_creditors", 0) + bs.get("wc_loan", 0) or 1
        quick_assets = bs.get("receivables", 0) + bs.get("cash", 0)
        total_capital_employed = equity + bs.get("term_loan_outstanding", 0)
        # Ratios
        current_ratio = current_assets / current_liabs if current_liabs else 0
        quick_ratio = quick_assets / current_liabs if current_liabs else 0
        debt_equity = debt / equity if equity > 0 else 0
        interest_coverage = ebitda / interest if interest > 0 else 0
        net_margin = (net_profit / revenue * 100) if revenue > 0 else 0
        ebitda_margin = (ebitda / revenue * 100) if revenue > 0 else 0
        roce = ((net_profit + interest) / total_capital_employed * 100) if total_capital_employed > 0 else 0
        roe = (net_profit / equity * 100) if equity > 0 else 0
        dscr = s.get("dscr", 0)
        rows.append({
            "year": yr.year,
            "current_ratio": round(current_ratio, 2),
            "quick_ratio": round(quick_ratio, 2),
            "debt_equity": round(debt_equity, 2),
            "interest_coverage": round(interest_coverage, 2),
            "ebitda_margin_pct": round(ebitda_margin, 2),
            "net_margin_pct": round(net_margin, 2),
            "roce_pct": round(roce, 2),
            "roe_pct": round(roe, 2),
            "dscr": round(dscr, 2),
        })
    def avg(key):
        vals = [r[key] for r in rows if r[key] not in (None, 0)]
        return round(sum(vals) / len(vals), 2) if vals else 0
    averages = {k: avg(k) for k in ("current_ratio", "quick_ratio", "debt_equity",
                                     "interest_coverage", "ebitda_margin_pct",
                                     "net_margin_pct", "roce_pct", "roe_pct", "dscr")}
    return {"rows": rows, "averages": averages}


@api_router.get("/projects/{project_id}/ratios")
async def ratios_endpoint(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    return compute_ratios(Project(**_serialize_project(doc)))


# ============================== EXPORTS ==============================

def _fmt_inr(amt: float) -> str:
    try:
        return f"₹ {amt:,.0f}"
    except Exception:
        return str(amt)


@api_router.get("/projects/{project_id}/download/excel")
async def download_excel(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    p = Project(**_serialize_project(doc))
    _ensure_paid(p)
    summary = compute_year_summary(p)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    # Sheet 1: Project Overview
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "DETAILED PROJECT REPORT"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")
    rows = [
        ("Business Name", p.business_name),
        ("Business Type", p.business_type),
        ("Industry", p.industry),
        ("Constitution", p.constitution),
        ("Location", f"{p.location}, {p.state}"),
        ("Loan Scheme", p.loan_scheme),
        ("Loan Amount", _fmt_inr(p.loan_amount)),
        ("Projection Years", p.projection_years),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # Sheet 2: Project Cost
    ws2 = wb.create_sheet("Project Cost")
    ws2.append(["Cost Head", "Amount (INR)"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    total = 0
    for item in p.project_cost:
        ws2.append([item.name, item.amount])
        total += item.amount
    ws2.append(["TOTAL", total])
    ws2[ws2.max_row][0].font = Font(bold=True)
    ws2[ws2.max_row][1].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    # Sheet 3: Means of Finance
    ws3 = wb.create_sheet("Means of Finance")
    ws3.append(["Source", "Amount (INR)"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
    mof = p.means_of_finance
    for label, val in [("Promoter Contribution", mof.promoter_contribution),
                       ("Term Loan", mof.term_loan),
                       ("Working Capital Loan", mof.working_capital_loan),
                       ("Subsidy", mof.subsidy),
                       ("Other Sources", mof.other_sources)]:
        ws3.append([label, val])
    ws3.append(["TOTAL", mof.promoter_contribution + mof.term_loan + mof.working_capital_loan + mof.subsidy + mof.other_sources])
    ws3.append([])
    ws3.append(["Interest Rate (%)", mof.interest_rate])
    ws3.append(["Tenure (Years)", mof.loan_tenure_years])
    ws3.append(["Moratorium (Months)", mof.moratorium_months])
    emi = compute_emi(mof.term_loan, mof.interest_rate, mof.loan_tenure_years)
    ws3.append(["Monthly EMI", round(emi, 2)])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    # Sheet 4: Projected P&L (CMA-style)
    ws4 = wb.create_sheet("P&L Projections")
    header = ["Particulars"] + [f"Year {y['year']}" for y in summary]
    ws4.append(header)
    for c in ws4[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    rows_def = [
        ("Revenue", "revenue"),
        ("Total Expenses", "total_expenses"),
        ("EBITDA", "ebitda"),
        ("Depreciation", "depreciation"),
        ("Interest", "interest"),
        ("Profit Before Tax", "ebt"),
        ("Tax", "tax"),
        ("Net Profit", "net_profit"),
        ("DSCR", "dscr"),
    ]
    for label, key in rows_def:
        row = [label] + [round(y[key], 2) for y in summary]
        ws4.append(row)
    ws4.column_dimensions["A"].width = 25
    for col_idx in range(2, len(summary) + 2):
        ws4.column_dimensions[chr(64 + col_idx)].width = 16

    # Sheet 5: Promoters
    ws5 = wb.create_sheet("Promoters")
    ws5.append(["Name", "Age", "Qualification", "Experience", "Contact", "PAN"])
    for c in ws5[1]:
        c.font = header_font
        c.fill = header_fill
    for pr in p.promoters:
        ws5.append([pr.name, pr.age, pr.qualification, pr.experience, pr.contact, pr.pan])
    for col in "ABCDEF":
        ws5.column_dimensions[col].width = 20

    # Sheet 6: Narrative
    ws6 = wb.create_sheet("Narrative")
    ws6["A1"] = "EXECUTIVE SUMMARY"
    ws6["A1"].font = Font(bold=True, size=14)
    ws6["A2"] = p.narrative.executive_summary
    ws6["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws6["A4"] = "PROJECT DESCRIPTION"
    ws6["A4"].font = Font(bold=True, size=14)
    ws6["A5"] = p.narrative.project_description
    ws6["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws6["A7"] = "MARKETING STRATEGY"
    ws6["A7"].font = Font(bold=True, size=14)
    ws6["A8"] = p.narrative.marketing_strategy
    ws6["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws6.column_dimensions["A"].width = 100

    # Sheet 7: Working Capital Assessment
    wc_data = compute_working_capital(p)
    ws7 = wb.create_sheet("Working Capital")
    ws7.append(["Working Capital Assessment", "Method: " + (wc_data.get("method") or "Nayak")])
    ws7["A1"].font = Font(bold=True, size=14)
    ws7.append([])
    header_row = ["Particulars"] + [f"Year {r['year']}" for r in wc_data["rows"]]
    ws7.append(header_row)
    for c in ws7[ws7.max_row]:
        c.font = header_font; c.fill = header_fill
    rows_def = [
        ("Raw Material Stock", "raw_material_stock"),
        ("Finished Goods Stock", "finished_goods_stock"),
        ("Receivables", "receivables"),
        ("Cash Required", "cash"),
        ("Gross Working Capital", "gross_wc"),
        ("Less: Payables", "payables"),
        ("Net Working Capital", "net_wc"),
        ("Bank Finance (75%)", "bank_finance_75"),
        ("Promoter Margin (25%)", "promoter_margin_25"),
    ]
    for label, key in rows_def:
        ws7.append([label] + [r[key] for r in wc_data["rows"]])
    ws7.column_dimensions["A"].width = 28

    # Sheet 8: Projected Balance Sheet
    bs_data = compute_balance_sheet(p)
    ws8 = wb.create_sheet("Balance Sheet")
    ws8.append(["Projected Balance Sheet"])
    ws8["A1"].font = Font(bold=True, size=14)
    ws8.append([])
    ws8.append(["Particulars"] + [f"Year {r['year']}" for r in bs_data["rows"]])
    for c in ws8[ws8.max_row]:
        c.font = header_font; c.fill = header_fill
    ws8.append(["LIABILITIES"])
    ws8[ws8.max_row][0].font = Font(bold=True)
    for label, key in [
        ("Capital", "capital"),
        ("Subsidy", "subsidy"),
        ("Reserves & Surplus", "retained_earnings"),
        ("Term Loan Outstanding", "term_loan_outstanding"),
        ("WC Loan", "wc_loan"),
        ("Sundry Creditors", "sundry_creditors"),
        ("Total Liabilities", "total_liabilities"),
    ]:
        ws8.append([label] + [r[key] for r in bs_data["rows"]])
    ws8.append(["ASSETS"])
    ws8[ws8.max_row][0].font = Font(bold=True)
    for label, key in [
        ("Net Fixed Assets", "net_fixed_assets"),
        ("Inventory", "inventory"),
        ("Receivables", "receivables"),
        ("Cash & Bank", "cash"),
        ("Total Assets", "total_assets"),
    ]:
        ws8.append([label] + [r[key] for r in bs_data["rows"]])
    ws8.column_dimensions["A"].width = 28

    # Sheet 9: Ratio Analysis
    ratio_data = compute_ratios(p)
    ws9 = wb.create_sheet("Ratio Analysis")
    ws9.append(["Key Financial Ratios"])
    ws9["A1"].font = Font(bold=True, size=14)
    ws9.append([])
    ws9.append(["Ratio"] + [f"Year {r['year']}" for r in ratio_data["rows"]] + ["Average"])
    for c in ws9[ws9.max_row]:
        c.font = header_font; c.fill = header_fill
    rdef = [
        ("Current Ratio", "current_ratio"),
        ("Quick Ratio", "quick_ratio"),
        ("Debt-Equity", "debt_equity"),
        ("Interest Coverage", "interest_coverage"),
        ("EBITDA Margin (%)", "ebitda_margin_pct"),
        ("Net Profit Margin (%)", "net_margin_pct"),
        ("ROCE (%)", "roce_pct"),
        ("ROE (%)", "roe_pct"),
        ("DSCR", "dscr"),
    ]
    for label, key in rdef:
        ws9.append([label] + [r[key] for r in ratio_data["rows"]] + [ratio_data["averages"].get(key, 0)])
    ws9.column_dimensions["A"].width = 28

    # Sheet 10: Collateral / Security
    if p.collateral:
        ws10 = wb.create_sheet("Collateral")
        ws10.append(["Collateral / Security Offered"])
        ws10["A1"].font = Font(bold=True, size=14)
        ws10.append([])
        ws10.append(["Type", "Description", "Location", "Owner", "Market Value", "Realisable Value"])
        for c in ws10[ws10.max_row]:
            c.font = header_font; c.fill = header_fill
        total_mv = 0
        total_rv = 0
        for col in p.collateral:
            ws10.append([col.type, col.description, col.location, col.owner, col.market_value, col.realisable_value])
            total_mv += col.market_value
            total_rv += col.realisable_value
        ws10.append(["TOTAL", "", "", "", total_mv, total_rv])
        ws10[ws10.max_row][0].font = Font(bold=True)
        for col_letter in "ABCDEF":
            ws10.column_dimensions[col_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"DPR_{p.business_name or 'project'}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/projects/{project_id}/download/pdf")
async def download_pdf(project_id: str, user: User = Depends(get_current_user)):
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    p = Project(**_serialize_project(doc))
    _ensure_paid(p)
    summary = compute_year_summary(p)

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=20 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                 fontSize=22, textColor=colors.HexColor("#0F172A"),
                                 spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                               fontSize=11, textColor=colors.HexColor("#475569"),
                               alignment=1, spaceAfter=18)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                        fontSize=14, textColor=colors.HexColor("#1D4ED8"),
                        spaceBefore=14, spaceAfter=8)
    body = ParagraphStyle("Body", parent=styles["Normal"],
                          fontSize=10, leading=14, textColor=colors.HexColor("#0F172A"))

    story = []
    story.append(Paragraph("DETAILED PROJECT REPORT", title_style))
    story.append(Paragraph(p.business_name or "Untitled Project", sub_style))

    # Applicant block (if filled)
    if p.applicant.full_name or p.applicant.aadhaar or p.applicant.pan:
        ap = p.applicant
        addr = ", ".join([x for x in [ap.address_line1, ap.address_line2, ap.city, ap.state, ap.pincode] if x])
        story.append(Paragraph("APPLICANT", h2))
        ap_data = [
            ["Name", ap.full_name or "-", "Father's Name", ap.father_name or "-"],
            ["DOB", ap.dob or "-", "Gender", ap.gender or "-"],
            ["Category", ap.category or "-", "Mobile", ap.mobile or "-"],
            ["Email", ap.email or "-", "Aadhaar", ap.aadhaar or "-"],
            ["PAN", ap.pan or "-", "Pincode", ap.pincode or "-"],
        ]
        at = Table(ap_data, colWidths=[28 * mm, 60 * mm, 28 * mm, 60 * mm])
        at.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F1F5F9")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(at)
        if addr:
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Address:</b> {addr}", body))

    # Cover info table
    cover_data = [
        ["Business Type", p.business_type, "Loan Scheme", p.loan_scheme],
        ["Constitution", p.constitution, "Industry", p.industry or "-"],
        ["Stage", "Existing Business" if p.business_stage == "existing" else "New Business",
         "Loan Amount", _fmt_inr(p.loan_amount)],
        ["Location", f"{p.location}, {p.state}", "Projection", f"{p.projection_years} years"],
    ]
    t = Table(cover_data, colWidths=[35 * mm, 55 * mm, 35 * mm, 55 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    # Executive Summary
    if p.narrative.executive_summary:
        story.append(Paragraph("1. EXECUTIVE SUMMARY", h2))
        story.append(Paragraph(p.narrative.executive_summary.replace("\n", "<br/>"), body))

    if p.narrative.project_description:
        story.append(Paragraph("2. PROJECT DESCRIPTION", h2))
        story.append(Paragraph(p.narrative.project_description.replace("\n", "<br/>"), body))

    # Promoters
    if p.promoters:
        story.append(Paragraph("3. PROMOTER DETAILS", h2))
        prom_data = [["Name", "Qual.", "Experience", "Contact"]]
        for pr in p.promoters:
            prom_data.append([pr.name, pr.qualification, pr.experience, pr.contact])
        pt = Table(prom_data, colWidths=[40 * mm, 40 * mm, 50 * mm, 40 * mm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(pt)

    # Project Cost
    story.append(Paragraph("4. PROJECT COST", h2))
    cost_data = [["Cost Head", "Amount (₹)"]]
    total = 0
    for item in p.project_cost:
        cost_data.append([item.name, f"{item.amount:,.0f}"])
        total += item.amount
    cost_data.append(["TOTAL", f"{total:,.0f}"])
    ct = Table(cost_data, colWidths=[110 * mm, 60 * mm])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(ct)

    # Means of Finance
    story.append(Paragraph("5. MEANS OF FINANCE", h2))
    mof = p.means_of_finance
    mof_data = [
        ["Source", "Amount (₹)"],
        ["Promoter Contribution", f"{mof.promoter_contribution:,.0f}"],
        ["Term Loan", f"{mof.term_loan:,.0f}"],
        ["Working Capital Loan", f"{mof.working_capital_loan:,.0f}"],
        ["Subsidy", f"{mof.subsidy:,.0f}"],
        ["Other Sources", f"{mof.other_sources:,.0f}"],
        ["TOTAL", f"{mof.promoter_contribution + mof.term_loan + mof.working_capital_loan + mof.subsidy + mof.other_sources:,.0f}"],
    ]
    mt = Table(mof_data, colWidths=[110 * mm, 60 * mm])
    mt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(mt)

    emi = compute_emi(mof.term_loan, mof.interest_rate, mof.loan_tenure_years)
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Interest Rate: <b>{mof.interest_rate}%</b> &nbsp; Tenure: <b>{mof.loan_tenure_years} years</b> &nbsp; Monthly EMI: <b>₹ {emi:,.0f}</b>",
        body))

    # P&L Projections
    story.append(PageBreak())
    story.append(Paragraph("6. PROJECTED PROFITABILITY (CMA)", h2))
    pl_header = ["Particulars"] + [f"Year {y['year']}" for y in summary]
    pl_data = [pl_header]
    rows_def = [
        ("Revenue", "revenue"),
        ("Total Expenses", "total_expenses"),
        ("EBITDA", "ebitda"),
        ("Depreciation", "depreciation"),
        ("Interest", "interest"),
        ("Profit Before Tax", "ebt"),
        ("Tax", "tax"),
        ("Net Profit", "net_profit"),
        ("DSCR", "dscr"),
    ]
    for label, key in rows_def:
        row = [label] + [(f"{y[key]:,.2f}" if key == "dscr" else f"{y[key]:,.0f}") for y in summary]
        pl_data.append(row)
    col_widths = [40 * mm] + [(130 / max(len(summary), 1)) * mm for _ in summary]
    pl = Table(pl_data, colWidths=col_widths)
    pl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(pl)

    # Marketing
    if p.narrative.marketing_strategy:
        story.append(Paragraph("7. MARKETING STRATEGY", h2))
        story.append(Paragraph(p.narrative.marketing_strategy.replace("\n", "<br/>"), body))

    # SWOT
    has_swot = any([p.narrative.swot_strengths, p.narrative.swot_weaknesses,
                    p.narrative.swot_opportunities, p.narrative.swot_threats])
    if has_swot:
        story.append(Paragraph("8. SWOT ANALYSIS", h2))
        swot_data = [
            [Paragraph("<b>Strengths</b>", body), Paragraph("<b>Weaknesses</b>", body)],
            [Paragraph("<br/>".join(f"• {s}" for s in p.narrative.swot_strengths), body),
             Paragraph("<br/>".join(f"• {s}" for s in p.narrative.swot_weaknesses), body)],
            [Paragraph("<b>Opportunities</b>", body), Paragraph("<b>Threats</b>", body)],
            [Paragraph("<br/>".join(f"• {s}" for s in p.narrative.swot_opportunities), body),
             Paragraph("<br/>".join(f"• {s}" for s in p.narrative.swot_threats), body)],
        ]
        sw = Table(swot_data, colWidths=[85 * mm, 85 * mm])
        sw.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#F1F5F9")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(sw)

    # Working Capital Assessment
    story.append(PageBreak())
    story.append(Paragraph("9. WORKING CAPITAL ASSESSMENT", h2))
    wc_data = compute_working_capital(p)
    story.append(Paragraph(f"<b>Method:</b> {wc_data.get('method') or 'Nayak'}", body))
    wc_rows = wc_data["rows"]
    if wc_rows:
        wc_header = ["Particulars"] + [f"Year {r['year']}" for r in wc_rows]
        wc_pdf_data = [wc_header]
        for label, key in [
            ("Raw Material Stock", "raw_material_stock"),
            ("Finished Goods Stock", "finished_goods_stock"),
            ("Receivables", "receivables"),
            ("Cash Required", "cash"),
            ("Gross Working Capital", "gross_wc"),
            ("Less: Payables", "payables"),
            ("Net Working Capital", "net_wc"),
            ("Bank Finance (75%)", "bank_finance_75"),
            ("Promoter Margin (25%)", "promoter_margin_25"),
        ]:
            wc_pdf_data.append([label] + [f"{r[key]:,.0f}" for r in wc_rows])
        col_widths_wc = [50 * mm] + [(120 / max(len(wc_rows), 1)) * mm for _ in wc_rows]
        wt = Table(wc_pdf_data, colWidths=col_widths_wc)
        wt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(wt)

    # Projected Balance Sheet
    story.append(Paragraph("10. PROJECTED BALANCE SHEET", h2))
    bs_data = compute_balance_sheet(p)
    bs_rows = bs_data["rows"]
    if bs_rows:
        bs_pdf_data = [["Particulars"] + [f"Year {r['year']}" for r in bs_rows]]
        bs_pdf_data.append(["LIABILITIES"] + [""] * len(bs_rows))
        for label, key in [
            ("Capital", "capital"),
            ("Subsidy", "subsidy"),
            ("Reserves & Surplus", "retained_earnings"),
            ("Term Loan Outstanding", "term_loan_outstanding"),
            ("WC Loan", "wc_loan"),
            ("Sundry Creditors", "sundry_creditors"),
            ("Total Liabilities", "total_liabilities"),
        ]:
            bs_pdf_data.append([label] + [f"{r[key]:,.0f}" for r in bs_rows])
        bs_pdf_data.append(["ASSETS"] + [""] * len(bs_rows))
        for label, key in [
            ("Net Fixed Assets", "net_fixed_assets"),
            ("Inventory", "inventory"),
            ("Receivables", "receivables"),
            ("Cash & Bank", "cash"),
            ("Total Assets", "total_assets"),
        ]:
            bs_pdf_data.append([label] + [f"{r[key]:,.0f}" for r in bs_rows])
        col_widths_bs = [50 * mm] + [(120 / max(len(bs_rows), 1)) * mm for _ in bs_rows]
        bt = Table(bs_pdf_data, colWidths=col_widths_bs)
        bt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F1F5F9")),
            ("BACKGROUND", (0, 9), (-1, 9), colors.HexColor("#F1F5F9")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(bt)

    # Ratio Analysis
    ratio_data = compute_ratios(p)
    if ratio_data["rows"]:
        story.append(Paragraph("11. RATIO ANALYSIS", h2))
        rt_data = [["Ratio"] + [f"Year {r['year']}" for r in ratio_data["rows"]] + ["Avg"]]
        rdef = [
            ("Current Ratio", "current_ratio"),
            ("Quick Ratio", "quick_ratio"),
            ("Debt-Equity", "debt_equity"),
            ("Interest Coverage", "interest_coverage"),
            ("EBITDA Margin (%)", "ebitda_margin_pct"),
            ("Net Profit Margin (%)", "net_margin_pct"),
            ("ROCE (%)", "roce_pct"),
            ("ROE (%)", "roe_pct"),
            ("DSCR", "dscr"),
        ]
        for label, key in rdef:
            rt_data.append([label] + [f"{r[key]:,.2f}" for r in ratio_data["rows"]] + [f"{ratio_data['averages'].get(key, 0):,.2f}"])
        col_widths_rt = [42 * mm] + [(110 / max(len(ratio_data["rows"]) + 1, 1)) * mm for _ in range(len(ratio_data["rows"]) + 1)]
        rt = Table(rt_data, colWidths=col_widths_rt)
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(rt)

    # Collateral / Security
    if p.collateral:
        story.append(Paragraph("12. COLLATERAL / SECURITY", h2))
        col_data = [["Type", "Description", "Owner", "Market Value (₹)", "Realisable (₹)"]]
        tot_mv = 0
        tot_rv = 0
        for c in p.collateral:
            col_data.append([c.type, c.description[:50], c.owner, f"{c.market_value:,.0f}", f"{c.realisable_value:,.0f}"])
            tot_mv += c.market_value
            tot_rv += c.realisable_value
        col_data.append(["TOTAL", "", "", f"{tot_mv:,.0f}", f"{tot_rv:,.0f}"])
        ct = Table(col_data, colWidths=[28 * mm, 60 * mm, 30 * mm, 30 * mm, 25 * mm])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(ct)

    # Prepared By signature block
    pb = p.prepared_by
    if pb.name or pb.firm:
        story.append(Spacer(1, 14))
        story.append(Paragraph("PREPARED BY", h2))
        prep_lines = []
        if pb.name: prep_lines.append(f"<b>{pb.name}</b>")
        if pb.designation: prep_lines.append(pb.designation)
        if pb.firm: prep_lines.append(pb.firm)
        if pb.contact: prep_lines.append(f"Contact: {pb.contact}")
        if pb.email: prep_lines.append(f"Email: {pb.email}")
        story.append(Paragraph("<br/>".join(prep_lines), body))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<i>Generated on {datetime.now(timezone.utc).strftime('%d %B %Y')} via DPRForge — Loan DPR & CMA Software</i>",
        ParagraphStyle("Foot", parent=body, alignment=1, textColor=colors.HexColor("#94A3B8"), fontSize=8)))

    pdf.build(story)
    buf.seek(0)
    filename = f"DPR_{p.business_name or 'project'}.pdf".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================== ADMIN LOGIN (SEPARATE) + QUICK BUY (GUEST) ==============================

class QuickBuyRequest(BaseModel):
    name: str = Field(..., min_length=2, max_length=120)
    email: EmailStr
    mobile: str = Field(..., min_length=7, max_length=15)


@api_router.post("/auth/quick-buy", response_model=AuthResponse)
async def quick_buy_signup(req: QuickBuyRequest):
    """Create a passwordless GUEST account so a visitor can buy a DPR at ₹799 without registering.
    If the email already belongs to a real (non-guest) account, asks them to log in.
    """
    mobile = "".join(c for c in req.mobile if c.isdigit())
    if len(mobile) < 7:
        raise HTTPException(status_code=400, detail="Valid mobile number required")
    email = str(req.email).lower()
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    now = datetime.now(timezone.utc)

    if existing and not existing.get("is_guest"):
        raise HTTPException(
            status_code=400,
            detail="This email already has an account. Please sign in instead.",
        )

    if existing and existing.get("is_guest"):
        # Re-use the guest account so they don't accumulate orphan rows
        user_id = existing["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": req.name, "mobile": mobile}},
        )
        refreshed = await db.users.find_one({"user_id": user_id}, {"_id": 0}) or existing
        if isinstance(refreshed.get("created_at"), str):
            refreshed["created_at"] = datetime.fromisoformat(refreshed["created_at"])
        token = create_jwt(user_id)
        return AuthResponse(user=User(**refreshed), token=token)

    user_id = f"guest_{uuid.uuid4().hex[:12]}"
    referral_code = f"MBDS-{uuid.uuid4().hex[:6].upper()}"
    user_doc = {
        "user_id": user_id,
        "email": email,
        "name": req.name,
        "mobile": mobile,
        "picture": None,
        "auth_provider": "guest",
        "password_hash": "",  # passwordless
        "referral_code": referral_code,
        "referred_by": "",
        "referral_credits": 0,
        "free_dpr_credits": 0,
        "wallet_balance": 0.0,
        "is_guest": True,
        "is_admin": False,
        "created_at": now.isoformat(),
    }
    await db.users.insert_one(user_doc)
    token = create_jwt(user_id)
    return AuthResponse(
        user=User(user_id=user_id, email=email, name=req.name,
                  auth_provider="guest", referral_code=referral_code,
                  referred_by="", free_dpr_credits=0, wallet_balance=0.0,
                  is_guest=True, is_admin=False, created_at=now),
        token=token,
    )


@api_router.post("/auth/admin-login", response_model=AuthResponse)
async def admin_login(req: LoginRequest):
    """Admin-only login endpoint. Rejects non-admin credentials so admins can't accidentally
    log in via the public user page and vice versa."""
    user_doc = await db.users.find_one({"email": req.email.lower()}, {"_id": 0})
    if not user_doc or not user_doc.get("password_hash"):
        raise HTTPException(status_code=401, detail="Invalid admin credentials")
    if not verify_password(req.password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid admin credentials")

    email_lower = user_doc.get("email", "").lower()
    # Auto-promote configured admin emails
    if email_lower in ADMIN_EMAILS and not user_doc.get("is_admin"):
        await db.users.update_one({"user_id": user_doc["user_id"]}, {"$set": {"is_admin": True}})
        user_doc["is_admin"] = True

    if not (user_doc.get("is_admin") or email_lower in ADMIN_EMAILS):
        raise HTTPException(status_code=403, detail="Not an admin account")

    if isinstance(user_doc.get("created_at"), str):
        user_doc["created_at"] = datetime.fromisoformat(user_doc["created_at"])
    token = create_jwt(user_doc["user_id"])
    return AuthResponse(user=User(**user_doc), token=token)


# ============================== WALLET ==============================

class WalletTopupRequest(BaseModel):
    txn_id: str
    amount: float
    method: str = "GPay"


@api_router.get("/wallet/me")
async def wallet_me(user: User = Depends(get_current_user)):
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    txns = await db.wallet_txns.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    settings = await get_company_settings()
    return {
        "balance": float((user_doc or {}).get("wallet_balance", 0)),
        "free_dpr_credits": int((user_doc or {}).get("free_dpr_credits", 0)),
        "referral_credits": int((user_doc or {}).get("referral_credits", 0)),
        "bulk_price": int(settings["referral_price_inr"]),
        "transactions": txns,
    }


@api_router.post("/wallet/topup")
async def wallet_topup(payload: WalletTopupRequest, user: User = Depends(get_current_user)):
    """User submits a UPI top-up transaction. Pending verification by admin.
    Balance is credited immediately but flagged 'pending' so admin can reverse it if needed.
    """
    if not payload.txn_id or len(payload.txn_id.strip()) < 4:
        raise HTTPException(status_code=400, detail="Valid transaction ID required")
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")

    now = datetime.now(timezone.utc)
    txn = {
        "txn_uid": str(uuid.uuid4()),
        "user_id": user.user_id,
        "user_email": user.email,
        "type": "topup",
        "amount": float(payload.amount),
        "txn_id": payload.txn_id.strip(),
        "method": payload.method,
        "status": "pending",
        "created_at": now.isoformat(),
    }
    await db.wallet_txns.insert_one(dict(txn))
    # Credit balance now (optimistic). Admin can reverse if not verified.
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$inc": {"wallet_balance": float(payload.amount)}},
    )
    new_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "wallet_balance": 1})
    return {"ok": True, "balance": float(new_doc.get("wallet_balance", 0))}


@api_router.post("/projects/{project_id}/pay-from-wallet", response_model=Project)
async def pay_from_wallet(project_id: str, user: User = Depends(get_current_user)):
    """Pay for a project using wallet balance — no UPI/transaction needed."""
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    if doc.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Already paid")

    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    is_guest = bool((user_doc or {}).get("is_guest", False))
    settings = await get_company_settings()
    # Guests would pay 799, logged-in users 599
    price = int(settings["price_inr"]) if is_guest else int(settings["referral_price_inr"])

    balance = float((user_doc or {}).get("wallet_balance", 0))
    if balance < price:
        raise HTTPException(
            status_code=402,
            detail=f"Insufficient wallet balance. Need ₹{price}, you have ₹{balance:.0f}",
        )

    now = datetime.now(timezone.utc)
    # Deduct from wallet
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$inc": {"wallet_balance": -price}},
    )
    # Mark project paid
    await db.projects.update_one(
        {"project_id": project_id, "user_id": user.user_id},
        {"$set": {
            "payment_status": "paid",
            "payment_txn_id": f"WALLET-{uuid.uuid4().hex[:8].upper()}",
            "payment_amount": float(price),
            "payment_method": "Wallet",
            "paid_at": now.isoformat(),
            "updated_at": now.isoformat(),
        }},
    )
    # Wallet txn record
    await db.wallet_txns.insert_one({
        "txn_uid": str(uuid.uuid4()),
        "user_id": user.user_id,
        "user_email": user.email,
        "type": "debit",
        "amount": float(price),
        "project_id": project_id,
        "status": "verified",
        "created_at": now.isoformat(),
    })
    # Payment log (so admin can see it)
    await db.payment_logs.insert_one({
        "log_id": str(uuid.uuid4()),
        "user_id": user.user_id,
        "user_email": user.email,
        "project_id": project_id,
        "txn_id": f"WALLET-{uuid.uuid4().hex[:8].upper()}",
        "amount": float(price),
        "expected_price": price,
        "method": "Wallet",
        "submitted_at": now.isoformat(),
        "verification_status": "verified",
        "verified_by": "wallet-auto",
        "verified_at": now.isoformat(),
        "used_referral_credit": False,
    })
    # Reward referrer + give referee free watermarked DPR if first paid project
    await _award_referral_if_first_paid(user_doc)

    new_doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    return Project(**_serialize_project(new_doc))


# ============================== ADMIN: WALLET MANAGEMENT ==============================

class AdminWalletCredit(BaseModel):
    user_email: EmailStr
    amount: float
    note: Optional[str] = ""


@api_router.post("/admin/wallet/credit")
async def admin_wallet_credit(payload: AdminWalletCredit, admin: User = Depends(require_admin)):
    """Admin can credit/debit any user's wallet (use negative amount to debit)."""
    target = await db.users.find_one({"email": payload.user_email.lower()}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    await db.users.update_one(
        {"user_id": target["user_id"]},
        {"$inc": {"wallet_balance": float(payload.amount)}},
    )
    now = datetime.now(timezone.utc).isoformat()
    await db.wallet_txns.insert_one({
        "txn_uid": str(uuid.uuid4()),
        "user_id": target["user_id"],
        "user_email": target["email"],
        "type": "admin_credit" if payload.amount > 0 else "admin_debit",
        "amount": float(payload.amount),
        "status": "verified",
        "note": payload.note or "",
        "admin_email": admin.email,
        "created_at": now,
    })
    new_doc = await db.users.find_one({"user_id": target["user_id"]}, {"_id": 0, "wallet_balance": 1, "email": 1})
    return {"ok": True, "email": new_doc["email"], "balance": float(new_doc["wallet_balance"])}


@api_router.get("/admin/users")
async def admin_list_users(_: User = Depends(require_admin)):
    rows = await db.users.find(
        {}, {"_id": 0, "password_hash": 0, "tokens_invalidated_before": 0}
    ).sort("created_at", -1).to_list(500)
    for r in rows:
        if isinstance(r.get("created_at"), str):
            r["created_at"] = r["created_at"]
    return rows


@api_router.get("/admin/wallet/txns")
async def admin_wallet_txns(_: User = Depends(require_admin)):
    rows = await db.wallet_txns.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return rows


# ============================== FREE WATERMARKED DPR ==============================

def _watermark_canvas(canvas, doc):
    """Draw a diagonal watermark on every page of the PDF."""
    canvas.saveState()
    canvas.setFont("Helvetica-Bold", 60)
    canvas.setFillColorRGB(0.85, 0.85, 0.85)
    canvas.translate(105 * mm, 148 * mm)
    canvas.rotate(45)
    canvas.drawCentredString(0, 0, "DPRForge SAMPLE")
    canvas.setFont("Helvetica", 14)
    canvas.drawCentredString(0, -22, "www.dprforge.com — Upgrade to remove watermark")
    canvas.restoreState()


@api_router.get("/projects/{project_id}/download/free-watermarked-pdf")
async def download_free_watermarked_pdf(project_id: str, user: User = Depends(get_current_user)):
    """Download a watermarked sample PDF — uses 1 free_dpr_credits from referral signup."""
    doc = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Project not found")
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    free_credits = int((user_doc or {}).get("free_dpr_credits", 0))
    if free_credits <= 0 and doc.get("payment_status") != "paid":
        raise HTTPException(
            status_code=402,
            detail="No free DPR credits available. Sign up with a referral code to get 1 free watermarked DPR, or pay to unlock.",
        )

    p = Project(**_serialize_project(doc))
    summary = compute_year_summary(p)

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=20 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                 fontSize=22, textColor=colors.HexColor("#0F172A"),
                                 spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                               fontSize=11, textColor=colors.HexColor("#475569"),
                               alignment=1, spaceAfter=18)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                        fontSize=14, textColor=colors.HexColor("#1D4ED8"),
                        spaceBefore=14, spaceAfter=8)
    body = ParagraphStyle("Body", parent=styles["Normal"],
                          fontSize=10, leading=14, textColor=colors.HexColor("#0F172A"))

    story = [
        Paragraph("DETAILED PROJECT REPORT (SAMPLE)", title_style),
        Paragraph(p.business_name or "Untitled Project", sub_style),
        Paragraph("FREE PREVIEW — generated by DPRForge", body),
        Spacer(1, 8),
        Paragraph("Business Snapshot", h2),
    ]
    snap_data = [
        ["Business Name", p.business_name or "-", "Scheme", p.loan_scheme or "-"],
        ["Type", p.business_type or "-", "Loan Amount", f"₹{p.loan_amount:,.0f}"],
        ["Location", p.location or "-", "Years", str(p.projection_years)],
    ]
    snap = Table(snap_data, colWidths=[28 * mm, 60 * mm, 28 * mm, 60 * mm])
    snap.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(snap)

    # Year summary
    if summary:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Projected P&L Summary", h2))
        ys_header = ["Year", "Revenue", "EBITDA", "PAT"]
        ys_rows = [ys_header]
        for row in summary[:5]:
            ys_rows.append([
                f"Y{row['year']}",
                f"{row['revenue']:,.0f}",
                f"{row['ebitda']:,.0f}",
                f"{row['net_profit']:,.0f}",
            ])
        ys = Table(ys_rows, repeatRows=1)
        ys.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ]))
        story.append(ys)

    story.append(Spacer(1, 16))
    story.append(Paragraph(
        "<b>This is a FREE watermarked sample.</b> To download a clean, bank-ready DPR + CMA Excel "
        "report, please pay ₹599 (logged-in price) from your wallet or via UPI on the preview page.",
        body))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"<i>Generated on {datetime.now(timezone.utc).strftime('%d %B %Y')} via DPRForge — www.dprforge.com</i>",
        ParagraphStyle("Foot", parent=body, alignment=1, textColor=colors.HexColor("#94A3B8"), fontSize=8)))

    pdf.build(story, onFirstPage=_watermark_canvas, onLaterPages=_watermark_canvas)

    # Consume one free credit (only if not already paid)
    if doc.get("payment_status") != "paid":
        await db.users.update_one(
            {"user_id": user.user_id, "free_dpr_credits": {"$gt": 0}},
            {"$inc": {"free_dpr_credits": -1}},
        )

    buf.seek(0)
    filename = f"DPR_Sample_{p.business_name or 'project'}.pdf".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================== STARTUP: SEED ADMIN ==============================

@app.on_event("startup")
async def seed_admin_on_startup():
    """Ensure the configured admin account exists with a working password. Idempotent."""
    seed_email = (os.environ.get("ADMIN_SEED_EMAIL") or "").strip().lower()
    seed_password = os.environ.get("ADMIN_SEED_PASSWORD") or ""
    seed_name = os.environ.get("ADMIN_SEED_NAME") or "Admin"
    if not seed_email or not seed_password:
        return
    try:
        existing = await db.users.find_one({"email": seed_email}, {"_id": 0})
        now = datetime.now(timezone.utc).isoformat()
        if not existing:
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            await db.users.insert_one({
                "user_id": user_id,
                "email": seed_email,
                "name": seed_name,
                "picture": None,
                "auth_provider": "email",
                "password_hash": hash_password(seed_password),
                "referral_code": f"MBDS-{uuid.uuid4().hex[:6].upper()}",
                "referred_by": "",
                "referral_credits": 0,
                "free_dpr_credits": 0,
                "wallet_balance": 0.0,
                "is_admin": True,
                "created_at": now,
            })
            logger.info(f"[seed] Created admin account: {seed_email}")
        else:
            # Always reset password to seed value on startup, ensure is_admin=True
            await db.users.update_one(
                {"user_id": existing["user_id"]},
                {"$set": {
                    "password_hash": hash_password(seed_password),
                    "is_admin": True,
                    "name": existing.get("name") or seed_name,
                }},
            )
            logger.info(f"[seed] Refreshed admin credentials: {seed_email}")
    except Exception as e:
        logger.warning(f"[seed] Failed to seed admin: {e}")


# ============================== APP SETUP ==============================

app.include_router(api_router)

# CORS — supports both:
#   1. Explicit comma-separated list: CORS_ORIGINS="https://dprforge.com,https://www.dprforge.com"
#   2. Wildcard "*" — falls back to allow_origin_regex so credentials still work
_cors_env = os.environ.get('CORS_ORIGINS', '*').strip()
if _cors_env == '*' or _cors_env == '':
    # Browsers reject wildcard origin + credentials. Use regex match-all instead.
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=True,
        allow_origin_regex=".*",
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    _origins = [o.strip().rstrip("/") for o in _cors_env.split(',') if o.strip()]
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=True,
        allow_origins=_origins,
        allow_methods=["*"],
        allow_headers=["*"],
    )

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
