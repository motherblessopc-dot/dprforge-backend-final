"""
CMA Data Preparation Module — Standalone module appended to the DPR app.

Provides:
  - CMA Pydantic models
  - CRUD endpoints for CMA statements
  - Computations (Ratios, MPBF, Fund Flow)
  - Excel export (7 bank-standard sheets)
  - PDF export
  - Excel template (blank) + upload parser

All endpoints are mounted under /api/cma/* via `register_cma_routes(api_router, db, get_current_user)`.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)


# =============================== MODELS ===============================

class CMAYearLabel(BaseModel):
    """Column definition for a single financial year in CMA."""
    label: str  # e.g. "FY 2022-23"
    type: str   # "audited" | "provisional" | "projected"


class CMAOpStatement(BaseModel):
    """Form II — Operating Statement (one year of P&L)."""
    # Revenue
    domestic_sales: float = 0
    export_sales: float = 0
    # Less: Excise / GST on output
    less_excise: float = 0
    other_income: float = 0
    # Cost of Sales
    opening_stock_rm: float = 0
    raw_materials_purchase: float = 0
    closing_stock_rm: float = 0
    power_fuel: float = 0
    direct_labour: float = 0
    other_mfg_expenses: float = 0
    depreciation: float = 0
    opening_stock_wip: float = 0
    closing_stock_wip: float = 0
    opening_stock_fg: float = 0
    closing_stock_fg: float = 0
    # Operating Expenses
    selling_expenses: float = 0
    admin_expenses: float = 0
    interest_on_bank_borrowing: float = 0
    interest_on_others: float = 0
    # Tax
    tax_rate: float = 25.0


class CMABalanceSheet(BaseModel):
    """Form III — Analysis of Balance Sheet (one year)."""
    # === LIABILITIES ===
    # Current Liabilities
    short_term_borrowing_banks: float = 0  # CC / OD / WC loans
    sundry_creditors: float = 0
    advance_from_customers: float = 0
    provision_for_tax: float = 0
    other_current_liabilities: float = 0
    # Term Liabilities
    debentures: float = 0
    term_loan_banks: float = 0
    term_loan_others: float = 0
    deferred_payment_credits: float = 0
    other_term_liabilities: float = 0
    # Net Worth
    paid_up_capital: float = 0
    reserves_surplus: float = 0
    revaluation_reserve: float = 0
    other_reserves: float = 0
    # === ASSETS ===
    # Current Assets
    cash_bank_balances: float = 0
    investments_short_term: float = 0
    receivables_domestic: float = 0
    receivables_export: float = 0
    inventory_rm: float = 0
    inventory_wip: float = 0
    inventory_fg: float = 0
    advance_to_suppliers: float = 0
    other_current_assets: float = 0
    # Fixed Assets
    gross_block: float = 0
    accumulated_depreciation: float = 0
    capital_wip: float = 0
    # Other Non-Current Assets
    investments_long_term: float = 0
    intangible_assets: float = 0
    deferred_revenue_expenditure: float = 0
    other_non_current_assets: float = 0


class CMAExistingLimits(BaseModel):
    """Form I — Existing and Proposed Limits row."""
    facility: str = ""           # e.g. "Cash Credit", "Term Loan", "OD"
    bank: str = ""
    nature: str = ""             # Fund based / Non-fund based
    existing_limit: float = 0
    outstanding: float = 0
    proposed_limit: float = 0
    security: str = ""
    rate_of_interest: float = 0
    margin_pct: float = 0


class CMASensitivityScenario(BaseModel):
    """A what-if scenario for Form VIII."""
    name: str = "Adverse Scenario"
    description: str = ""
    sales_delta_pct: float = -10.0       # % change to net sales
    rm_cost_delta_pct: float = 5.0       # % change to raw material consumed
    interest_delta_pct: float = 10.0     # % change to interest expense
    other_expenses_delta_pct: float = 0.0  # % change to operating expenses


class CMASensitivity(BaseModel):
    """Form VIII — Sensitivity Analysis configuration."""
    enabled: bool = False
    target_year_index: int = -1  # -1 = last projected year
    scenarios: List[CMASensitivityScenario] = Field(
        default_factory=lambda: [
            CMASensitivityScenario(name="Base Case", description="As per projections",
                                   sales_delta_pct=0, rm_cost_delta_pct=0,
                                   interest_delta_pct=0, other_expenses_delta_pct=0),
            CMASensitivityScenario(name="Adverse (Sales -10%, RM +5%, Int +10%)",
                                   sales_delta_pct=-10, rm_cost_delta_pct=5,
                                   interest_delta_pct=10, other_expenses_delta_pct=0),
            CMASensitivityScenario(name="Stress (Sales -20%, RM +10%, Int +20%)",
                                   sales_delta_pct=-20, rm_cost_delta_pct=10,
                                   interest_delta_pct=20, other_expenses_delta_pct=5),
        ]
    )


class CMAYearData(BaseModel):
    """All financial data for one year."""
    year_label: str
    year_type: str = "audited"  # audited | provisional | projected
    op_statement: CMAOpStatement = Field(default_factory=CMAOpStatement)
    balance_sheet: CMABalanceSheet = Field(default_factory=CMABalanceSheet)


class CMAStatement(BaseModel):
    """Top-level CMA document."""
    model_config = ConfigDict(extra="ignore")
    cma_id: str = Field(default_factory=lambda: f"cma_{uuid.uuid4().hex[:12]}")
    user_id: str
    company_name: str = ""
    constitution: str = "Private Limited"  # Proprietary / Partnership / LLP / Pvt Ltd / Public Ltd
    industry: str = ""
    business_activity: str = ""
    registered_address: str = ""
    pan: str = ""
    gstin: str = ""
    date_of_incorporation: str = ""
    promoters: str = ""
    banker_name: str = ""
    facility_type: str = "CC"  # CC / OD / TL / Composite
    purpose: str = ""
    years: List[CMAYearData] = Field(default_factory=list)
    existing_limits: List[CMAExistingLimits] = Field(default_factory=list)
    sensitivity: CMASensitivity = Field(default_factory=CMASensitivity)
    notes: str = ""
    # Payment / locking (per-CMA)
    payment_status: str = "unpaid"   # unpaid | paid (admin or wallet flow)
    paid_at: Optional[datetime] = None
    source_dpr_project_id: str = ""  # set if generated from a DPR
    ai_summary: str = ""  # latest AI-generated bank narrative
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class CMACreate(BaseModel):
    company_name: str
    constitution: str = "Private Limited"
    industry: str = ""
    facility_type: str = "CC"
    banker_name: str = ""
    audited_count: int = 2
    provisional_count: int = 1
    projected_count: int = 3
    first_audited_fy: str = "FY 2022-23"  # used to label years sequentially


class CMAPaymentSubmit(BaseModel):
    txn_id: str
    amount: float = 0
    method: str = "GPay"


class CMARazorpayVerify(BaseModel):
    razorpay_payment_id: str
    razorpay_order_id: str
    razorpay_signature: str


class CMAUpdate(BaseModel):
    """Loose update payload — accepts any subset of CMAStatement fields."""
    model_config = ConfigDict(extra="allow")
    company_name: Optional[str] = None
    constitution: Optional[str] = None
    industry: Optional[str] = None
    business_activity: Optional[str] = None
    registered_address: Optional[str] = None
    pan: Optional[str] = None
    gstin: Optional[str] = None
    date_of_incorporation: Optional[str] = None
    promoters: Optional[str] = None
    banker_name: Optional[str] = None
    facility_type: Optional[str] = None
    purpose: Optional[str] = None
    years: Optional[List[CMAYearData]] = None
    existing_limits: Optional[List[CMAExistingLimits]] = None
    sensitivity: Optional[CMASensitivity] = None
    notes: Optional[str] = None


# =============================== HELPERS ===============================

def _next_fy_label(label: str) -> str:
    """Given 'FY 2022-23' produce 'FY 2023-24'. Best effort."""
    try:
        # tolerate variations like "FY 22-23"
        clean = label.replace("FY", "").strip()
        start, end = clean.split("-")
        start = start.strip()
        end = end.strip()
        if len(start) == 4 and len(end) == 2:
            s = int(start) + 1
            e = int(end) + 1
            return f"FY {s}-{str(e).zfill(2)}"
        if len(start) == 4 and len(end) == 4:
            s = int(start) + 1
            e = int(end) + 1
            return f"FY {s}-{e}"
        # numeric YY-YY
        s = int(start) + 1
        e = int(end) + 1
        return f"FY {str(s).zfill(2)}-{str(e).zfill(2)}"
    except Exception:
        return label + " +1"


def _scaffold_years(payload: CMACreate) -> List[CMAYearData]:
    years: List[CMAYearData] = []
    current = payload.first_audited_fy.strip() or "FY 2022-23"
    for _ in range(max(0, payload.audited_count)):
        years.append(CMAYearData(year_label=current, year_type="audited"))
        current = _next_fy_label(current)
    for _ in range(max(0, payload.provisional_count)):
        years.append(CMAYearData(year_label=current, year_type="provisional"))
        current = _next_fy_label(current)
    for _ in range(max(0, payload.projected_count)):
        years.append(CMAYearData(year_label=current, year_type="projected"))
        current = _next_fy_label(current)
    return years


# =============================== COMPUTATIONS ===============================

def _safe_div(a: float, b: float) -> float:
    return float(a) / float(b) if b not in (0, 0.0, None) else 0.0


def compute_year_metrics(yd: CMAYearData) -> Dict[str, float]:
    """Compute all derived metrics for a single year (Op Statement + Balance Sheet)."""
    op = yd.op_statement
    bs = yd.balance_sheet

    # ---- Op Statement ----
    gross_sales = op.domestic_sales + op.export_sales
    net_sales = gross_sales - op.less_excise
    total_revenue = net_sales + op.other_income

    # Raw material consumed
    rm_consumed = op.opening_stock_rm + op.raw_materials_purchase - op.closing_stock_rm

    # Cost of production
    cost_of_production = (
        rm_consumed + op.power_fuel + op.direct_labour
        + op.other_mfg_expenses + op.depreciation
        + (op.opening_stock_wip - op.closing_stock_wip)
    )
    cost_of_sales = cost_of_production + (op.opening_stock_fg - op.closing_stock_fg)
    gross_profit = net_sales - cost_of_sales

    interest_total = op.interest_on_bank_borrowing + op.interest_on_others
    operating_expenses = op.selling_expenses + op.admin_expenses
    operating_profit = gross_profit - operating_expenses + op.other_income
    pbdit = operating_profit + op.depreciation
    pbt = operating_profit - interest_total
    tax = max(0.0, pbt * (op.tax_rate / 100.0))
    pat = pbt - tax
    cash_accrual = pat + op.depreciation

    # ---- Balance Sheet groupings ----
    current_liab = (
        bs.short_term_borrowing_banks + bs.sundry_creditors
        + bs.advance_from_customers + bs.provision_for_tax
        + bs.other_current_liabilities
    )
    other_current_liab = current_liab - bs.short_term_borrowing_banks  # OCL (used in MPBF)

    term_liab = (
        bs.debentures + bs.term_loan_banks + bs.term_loan_others
        + bs.deferred_payment_credits + bs.other_term_liabilities
    )
    net_worth = (
        bs.paid_up_capital + bs.reserves_surplus
        + bs.revaluation_reserve + bs.other_reserves
    )
    tangible_net_worth = net_worth - bs.revaluation_reserve - bs.intangible_assets - bs.deferred_revenue_expenditure
    total_liab = current_liab + term_liab + net_worth

    current_assets = (
        bs.cash_bank_balances + bs.investments_short_term
        + bs.receivables_domestic + bs.receivables_export
        + bs.inventory_rm + bs.inventory_wip + bs.inventory_fg
        + bs.advance_to_suppliers + bs.other_current_assets
    )
    inventory_total = bs.inventory_rm + bs.inventory_wip + bs.inventory_fg
    receivables_total = bs.receivables_domestic + bs.receivables_export
    quick_assets = current_assets - inventory_total

    net_block = bs.gross_block - bs.accumulated_depreciation
    fixed_assets_total = net_block + bs.capital_wip
    non_current_assets = (
        fixed_assets_total + bs.investments_long_term
        + bs.intangible_assets + bs.deferred_revenue_expenditure
        + bs.other_non_current_assets
    )
    total_assets = current_assets + non_current_assets

    # Working Capital Gap (WCG) = CA – OCL
    wcg = current_assets - other_current_liab
    nwc = current_assets - current_liab  # Net Working Capital

    # MPBF (Tandon Committee)
    mpbf_method_1 = 0.75 * wcg
    mpbf_method_2 = 0.75 * current_assets - other_current_liab

    # Long-term Solvency Ratios
    current_ratio = _safe_div(current_assets, current_liab)
    quick_ratio = _safe_div(quick_assets, current_liab)
    absolute_liquid_ratio = _safe_div(bs.cash_bank_balances + bs.investments_short_term, current_liab)
    debt_equity = _safe_div(term_liab, tangible_net_worth)
    tol_tnw = _safe_div(term_liab + current_liab, tangible_net_worth)
    debt_to_net_worth = _safe_div(term_liab, net_worth)
    net_worth_to_total_assets = _safe_div(net_worth, total_assets) * 100.0
    capital_gearing = _safe_div(term_liab, net_worth + term_liab)
    fixed_assets_to_long_term = _safe_div(fixed_assets_total, net_worth + term_liab)
    proprietary_ratio = _safe_div(net_worth, total_assets) * 100.0
    # Profitability
    np_margin = _safe_div(pat, net_sales) * 100.0
    op_margin = _safe_div(operating_profit, net_sales) * 100.0
    op_margin_after_int = _safe_div(operating_profit - interest_total, net_sales) * 100.0
    gp_margin = _safe_div(gross_profit, net_sales) * 100.0
    cash_profit_ratio = _safe_div(cash_accrual, net_sales) * 100.0
    return_on_net_worth = _safe_div(pat, net_worth) * 100.0
    capital_employed = net_worth + term_liab
    roce = _safe_div(pbt + interest_total, capital_employed) * 100.0
    # Activity
    stock_turnover = _safe_div(net_sales, inventory_total) if inventory_total else 0.0
    debtors_velocity_days = _safe_div(receivables_total * 365.0, net_sales) if net_sales else 0.0
    creditors_velocity_days = _safe_div(bs.sundry_creditors * 365.0, op.raw_materials_purchase or net_sales) if (op.raw_materials_purchase or net_sales) else 0.0
    debtors_turnover = _safe_div(net_sales, receivables_total) if receivables_total else 0.0
    creditors_turnover = _safe_div(op.raw_materials_purchase or net_sales, bs.sundry_creditors) if bs.sundry_creditors else 0.0
    fixed_assets_turnover = _safe_div(net_sales, fixed_assets_total) if fixed_assets_total else 0.0
    assets_turnover = _safe_div(net_sales, total_assets) if total_assets else 0.0
    working_capital_turnover = _safe_div(net_sales, nwc) if nwc else 0.0
    sales_to_capital_employed = _safe_div(net_sales, capital_employed) if capital_employed else 0.0
    interest_coverage = _safe_div(pbdit, interest_total) if interest_total else 0.0
    dscr = _safe_div(pat + op.depreciation + interest_total, interest_total) if interest_total else 0.0
    # Operating ratios (% of net sales)
    domestic_sales_pct = _safe_div(op.domestic_sales, gross_sales) * 100.0
    export_sales_pct = _safe_div(op.export_sales, gross_sales) * 100.0
    material_cost_ratio = _safe_div(rm_consumed, net_sales) * 100.0
    direct_labour_ratio = _safe_div(op.direct_labour, net_sales) * 100.0
    other_overheads_ratio = _safe_div(op.power_fuel + op.other_mfg_expenses, net_sales) * 100.0
    indirect_cost_ratio = _safe_div(operating_expenses, net_sales) * 100.0
    interest_cost_ratio = _safe_div(interest_total, net_sales) * 100.0
    operating_cost_ratio = _safe_div(cost_of_sales + operating_expenses, net_sales) * 100.0
    pbit = operating_profit  # before finance charges

    return {
        # P&L
        "gross_sales": gross_sales,
        "net_sales": net_sales,
        "total_revenue": total_revenue,
        "rm_consumed": rm_consumed,
        "cost_of_production": cost_of_production,
        "cost_of_sales": cost_of_sales,
        "gross_profit": gross_profit,
        "operating_expenses": operating_expenses,
        "operating_profit": operating_profit,
        "pbdit": pbdit,
        "interest_total": interest_total,
        "pbt": pbt,
        "tax": tax,
        "pat": pat,
        "cash_accrual": cash_accrual,
        # BS groupings
        "current_liab": current_liab,
        "other_current_liab": other_current_liab,
        "term_liab": term_liab,
        "net_worth": net_worth,
        "tangible_net_worth": tangible_net_worth,
        "total_liab": total_liab,
        "current_assets": current_assets,
        "inventory_total": inventory_total,
        "receivables_total": receivables_total,
        "quick_assets": quick_assets,
        "net_block": net_block,
        "fixed_assets_total": fixed_assets_total,
        "non_current_assets": non_current_assets,
        "total_assets": total_assets,
        "wcg": wcg,
        "nwc": nwc,
        "mpbf_method_1": mpbf_method_1,
        "mpbf_method_2": mpbf_method_2,
        "capital_employed": capital_employed,
        "pbit": pbit,
        # Ratios
        "current_ratio": current_ratio,
        "quick_ratio": quick_ratio,
        "absolute_liquid_ratio": absolute_liquid_ratio,
        "debt_equity": debt_equity,
        "tol_tnw": tol_tnw,
        "debt_to_net_worth": debt_to_net_worth,
        "net_worth_to_total_assets": net_worth_to_total_assets,
        "capital_gearing": capital_gearing,
        "fixed_assets_to_long_term": fixed_assets_to_long_term,
        "proprietary_ratio": proprietary_ratio,
        "np_margin": np_margin,
        "op_margin": op_margin,
        "op_margin_after_int": op_margin_after_int,
        "gp_margin": gp_margin,
        "cash_profit_ratio": cash_profit_ratio,
        "return_on_net_worth": return_on_net_worth,
        "roce": roce,
        "stock_turnover": stock_turnover,
        "debtors_turnover": debtors_turnover,
        "creditors_turnover": creditors_turnover,
        "fixed_assets_turnover": fixed_assets_turnover,
        "assets_turnover": assets_turnover,
        "working_capital_turnover": working_capital_turnover,
        "sales_to_capital_employed": sales_to_capital_employed,
        "debtors_velocity_days": debtors_velocity_days,
        "creditors_velocity_days": creditors_velocity_days,
        "interest_coverage": interest_coverage,
        "dscr": dscr,
        # Operating ratios
        "domestic_sales_pct": domestic_sales_pct,
        "export_sales_pct": export_sales_pct,
        "material_cost_ratio": material_cost_ratio,
        "direct_labour_ratio": direct_labour_ratio,
        "other_overheads_ratio": other_overheads_ratio,
        "indirect_cost_ratio": indirect_cost_ratio,
        "interest_cost_ratio": interest_cost_ratio,
        "operating_cost_ratio": operating_cost_ratio,
    }


def compute_fund_flow(prev_m: Dict[str, float], curr_m: Dict[str, float],
                      prev_y: CMAYearData, curr_y: CMAYearData) -> Dict[str, float]:
    """Sources & Uses between two consecutive years."""
    sources = {
        "PAT (Cash Profit)": curr_m["pat"],
        "Depreciation": curr_y.op_statement.depreciation,
        "Increase in Net Worth (Capital + Reserves)": max(0.0, curr_m["net_worth"] - prev_m["net_worth"]) - curr_m["pat"],
        "Increase in Term Liabilities": max(0.0, curr_m["term_liab"] - prev_m["term_liab"]),
        "Decrease in Fixed Assets (net)": max(0.0, prev_m["fixed_assets_total"] - curr_m["fixed_assets_total"]),
        "Increase in Bank Borrowing (CC/OD)": max(0.0, curr_y.balance_sheet.short_term_borrowing_banks - prev_y.balance_sheet.short_term_borrowing_banks),
    }
    uses = {
        "Increase in Fixed Assets (net)": max(0.0, curr_m["fixed_assets_total"] - prev_m["fixed_assets_total"]),
        "Decrease in Term Liabilities": max(0.0, prev_m["term_liab"] - curr_m["term_liab"]),
        "Dividends / Drawings": max(0.0, curr_m["pat"] - max(0.0, curr_m["net_worth"] - prev_m["net_worth"])) if curr_m["pat"] > 0 else 0.0,
        "Increase in Current Assets (net)": max(0.0, curr_m["current_assets"] - prev_m["current_assets"]),
        "Decrease in Current Liabilities": max(0.0, prev_m["other_current_liab"] - curr_m["other_current_liab"]),
    }
    total_sources = sum(sources.values())
    total_uses = sum(uses.values())
    return {
        "sources": sources,
        "uses": uses,
        "total_sources": total_sources,
        "total_uses": total_uses,
        "surplus_deficit": total_sources - total_uses,
    }


def compute_sensitivity(stmt: CMAStatement) -> Dict[str, Any]:
    """Form VIII — apply each scenario's deltas to the target year and recompute key metrics."""
    if not stmt.years:
        return {"enabled": stmt.sensitivity.enabled, "target_year_label": "—", "scenarios": []}
    idx = stmt.sensitivity.target_year_index
    if idx < 0 or idx >= len(stmt.years):
        idx = len(stmt.years) - 1
    base_year = stmt.years[idx]
    results = []
    for sc in stmt.sensitivity.scenarios:
        # Clone and apply deltas
        op = base_year.op_statement.model_copy()
        sf = 1.0 + (sc.sales_delta_pct or 0) / 100.0
        rf = 1.0 + (sc.rm_cost_delta_pct or 0) / 100.0
        intf = 1.0 + (sc.interest_delta_pct or 0) / 100.0
        oef = 1.0 + (sc.other_expenses_delta_pct or 0) / 100.0
        op.domestic_sales *= sf
        op.export_sales *= sf
        op.raw_materials_purchase *= rf
        op.opening_stock_rm *= rf
        op.closing_stock_rm *= rf
        op.power_fuel *= oef
        op.direct_labour *= oef
        op.other_mfg_expenses *= oef
        op.selling_expenses *= oef
        op.admin_expenses *= oef
        op.interest_on_bank_borrowing *= intf
        op.interest_on_others *= intf
        modified = base_year.model_copy()
        modified.op_statement = op
        m = compute_year_metrics(modified)
        results.append({
            "name": sc.name,
            "description": sc.description,
            "deltas": {
                "sales_delta_pct": sc.sales_delta_pct,
                "rm_cost_delta_pct": sc.rm_cost_delta_pct,
                "interest_delta_pct": sc.interest_delta_pct,
                "other_expenses_delta_pct": sc.other_expenses_delta_pct,
            },
            "metrics": {
                "net_sales": m["net_sales"],
                "gross_profit": m["gross_profit"],
                "operating_profit": m["operating_profit"],
                "interest_total": m["interest_total"],
                "pbt": m["pbt"],
                "tax": m["tax"],
                "pat": m["pat"],
                "cash_accrual": m["cash_accrual"],
                "np_margin": m["np_margin"],
                "dscr": m["dscr"],
                "interest_coverage": m["interest_coverage"],
            },
        })
    return {
        "enabled": stmt.sensitivity.enabled,
        "target_year_label": base_year.year_label,
        "target_year_index": idx,
        "scenarios": results,
    }


def build_analysis(stmt: CMAStatement) -> Dict[str, Any]:
    """Compute everything for the whole CMA."""
    per_year = []
    for yd in stmt.years:
        per_year.append({
            "year_label": yd.year_label,
            "year_type": yd.year_type,
            "metrics": compute_year_metrics(yd),
        })

    fund_flows = []
    for i in range(1, len(stmt.years)):
        prev_m = per_year[i - 1]["metrics"]
        curr_m = per_year[i]["metrics"]
        ff = compute_fund_flow(prev_m, curr_m, stmt.years[i - 1], stmt.years[i])
        fund_flows.append({
            "from_year": stmt.years[i - 1].year_label,
            "to_year": stmt.years[i].year_label,
            **ff,
        })

    return {
        "cma_id": stmt.cma_id,
        "company_name": stmt.company_name,
        "facility_type": stmt.facility_type,
        "years": per_year,
        "fund_flows": fund_flows,
        "sensitivity": compute_sensitivity(stmt),
    }


# =============================== EXCEL EXPORT ===============================

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_SUBHEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
_TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")


def _h(cell, white=True):
    cell.font = Font(bold=True, color="FFFFFF" if white else "0F172A", size=11)
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _sh(cell):
    cell.font = Font(bold=True, color="0F172A", size=10)
    cell.fill = _SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _b(cell):
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00;[Red]-#,##0.00;"-"'


def _label(cell, bold=False):
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if bold:
        cell.font = Font(bold=True, size=10)


def _total(cell, value=True):
    cell.border = _BORDER
    cell.font = Font(bold=True, size=10)
    cell.fill = _TOTAL_FILL
    if value:
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0.00;[Red]-#,##0.00;"-"'
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _set_widths(ws, first=42, rest=18):
    ws.column_dimensions["A"].width = first
    for i in range(2, 20):
        ws.column_dimensions[get_column_letter(i)].width = rest


def _write_year_header(ws, row: int, years: List[CMAYearData]) -> int:
    ws.cell(row=row, column=1, value="Particulars (₹ in Lakhs)")
    _h(ws.cell(row=row, column=1))
    for i, y in enumerate(years):
        c = ws.cell(row=row, column=2 + i, value=f"{y.year_label}\n({y.year_type.title()})")
        _h(c)
    ws.row_dimensions[row].height = 32
    return row + 1


def _row(ws, row: int, label: str, values: List[float], bold: bool = False, total: bool = False):
    c = ws.cell(row=row, column=1, value=label)
    if total:
        _total(c, value=False)
    else:
        _label(c, bold=bold)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=2 + i, value=round(float(v or 0), 2))
        if total:
            _total(cell)
        else:
            _b(cell)
            if bold:
                cell.font = Font(bold=True, size=10)


def _section(ws, row: int, title: str, span: int):
    c = ws.cell(row=row, column=1, value=title)
    _sh(c)
    for i in range(1, span):
        cc = ws.cell(row=row, column=1 + i, value="")
        _sh(cc)
    return row + 1


def generate_cma_excel(stmt: CMAStatement) -> bytes:
    wb = Workbook()
    years = stmt.years
    metrics = [compute_year_metrics(y) for y in years]
    n_years = len(years)
    span = 1 + n_years

    # ============ SHEET 0: Cover ============
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "CMA DATA"
    cover["A1"].font = Font(bold=True, size=22, color="0F172A")
    cover["A2"] = "Credit Monitoring Arrangement — Bank Submission Format"
    cover["A2"].font = Font(italic=True, size=11, color="475569")

    cover_rows = [
        ("Company / Borrower", stmt.company_name),
        ("Constitution", stmt.constitution),
        ("Industry / Business", stmt.industry),
        ("Business Activity", stmt.business_activity),
        ("Registered Address", stmt.registered_address),
        ("PAN", stmt.pan),
        ("GSTIN", stmt.gstin),
        ("Date of Incorporation", stmt.date_of_incorporation),
        ("Promoters / Directors", stmt.promoters),
        ("Banker", stmt.banker_name),
        ("Facility Type", stmt.facility_type),
        ("Purpose", stmt.purpose),
        ("Prepared On", datetime.now(timezone.utc).strftime("%d %B %Y")),
    ]
    r = 4
    for k, v in cover_rows:
        cover.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        cover.cell(row=r, column=2, value=v or "—").font = Font(size=10)
        r += 1
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 60
    cover["A18"] = "Sheets in this workbook:"
    cover["A18"].font = Font(bold=True, size=11)
    sheet_list = [
        "Form I — Existing & Proposed Limits",
        "Form II — Operating Statement (P&L)",
        "Form III — Analysis of Balance Sheet",
        "Form IV — Comparative Statement of CA & CL",
        "Form V — Computation of MPBF (Method I & II)",
        "Form VI — Funds Flow Statement",
        "Form VII — Summary of Financial Statements for Ratio Analysis",
        "Form VIII — Ratio Analysis (Solvency / Profitability / Activity / Operating)",
        "Form IX — Statement of Changes in Working Capital",
    ]
    for i, s in enumerate(sheet_list):
        cover.cell(row=19 + i, column=1, value=f"  • {s}").font = Font(size=10)

    # ============ SHEET 1: Form I — Existing & Proposed Limits ============
    ws = wb.create_sheet("Form I - Limits")
    headers = ["S.No.", "Facility", "Bank", "Nature", "Existing Limit", "Outstanding", "Proposed Limit", "Security", "ROI (%)", "Margin (%)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i + 1, value=h)
        _h(c)
    ws.row_dimensions[1].height = 30
    if not stmt.existing_limits:
        ws.cell(row=2, column=1, value="—").alignment = Alignment(horizontal="center")
    else:
        for idx, lim in enumerate(stmt.existing_limits):
            row = 2 + idx
            vals = [idx + 1, lim.facility, lim.bank, lim.nature, lim.existing_limit, lim.outstanding,
                    lim.proposed_limit, lim.security, lim.rate_of_interest, lim.margin_pct]
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=i + 1, value=v)
                cell.border = _BORDER
                if i in (4, 5, 6, 8, 9):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00;[Red]-#,##0.00;"-"'
        # Totals row
        total_row = 2 + len(stmt.existing_limits)
        ws.cell(row=total_row, column=1, value="TOTAL")
        _total(ws.cell(row=total_row, column=1), value=False)
        for i in (1, 2, 3, 6, 7, 8, 9):
            _total(ws.cell(row=total_row, column=i + 1), value=(i in (4, 5, 6, 8, 9)))
        for i, col in enumerate((4, 5, 6)):  # Existing, Outstanding, Proposed (1-based: 5,6,7)
            total = sum(getattr(l, ["existing_limit", "outstanding", "proposed_limit"][i]) for l in stmt.existing_limits)
            ws.cell(row=total_row, column=col + 1, value=round(total, 2))
            _total(ws.cell(row=total_row, column=col + 1))
    widths = [6, 22, 18, 16, 16, 16, 16, 24, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ============ SHEET 2: Form II — Operating Statement ============
    ws = wb.create_sheet("Form II - Op Statement")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)

    r = _section(ws, r, "GROSS SALES", span)
    _row(ws, r, "Domestic Sales", [y.op_statement.domestic_sales for y in years]); r += 1
    _row(ws, r, "Export Sales", [y.op_statement.export_sales for y in years]); r += 1
    _row(ws, r, "Gross Sales", [m["gross_sales"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Less: Excise / GST", [y.op_statement.less_excise for y in years]); r += 1
    _row(ws, r, "Net Sales", [m["net_sales"] for m in metrics], total=True); r += 1
    _row(ws, r, "Add: Other Income", [y.op_statement.other_income for y in years]); r += 1
    _row(ws, r, "Total Revenue", [m["total_revenue"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "COST OF SALES", span)
    _row(ws, r, "Opening Stock of Raw Material", [y.op_statement.opening_stock_rm for y in years]); r += 1
    _row(ws, r, "Add: Purchases of Raw Material", [y.op_statement.raw_materials_purchase for y in years]); r += 1
    _row(ws, r, "Less: Closing Stock of Raw Material", [y.op_statement.closing_stock_rm for y in years]); r += 1
    _row(ws, r, "Raw Material Consumed", [m["rm_consumed"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Power & Fuel", [y.op_statement.power_fuel for y in years]); r += 1
    _row(ws, r, "Direct Labour / Wages", [y.op_statement.direct_labour for y in years]); r += 1
    _row(ws, r, "Other Manufacturing Expenses", [y.op_statement.other_mfg_expenses for y in years]); r += 1
    _row(ws, r, "Depreciation", [y.op_statement.depreciation for y in years]); r += 1
    _row(ws, r, "Add: Opening WIP", [y.op_statement.opening_stock_wip for y in years]); r += 1
    _row(ws, r, "Less: Closing WIP", [y.op_statement.closing_stock_wip for y in years]); r += 1
    _row(ws, r, "Cost of Production", [m["cost_of_production"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Add: Opening Finished Goods", [y.op_statement.opening_stock_fg for y in years]); r += 1
    _row(ws, r, "Less: Closing Finished Goods", [y.op_statement.closing_stock_fg for y in years]); r += 1
    _row(ws, r, "Cost of Sales", [m["cost_of_sales"] for m in metrics], total=True); r += 1
    _row(ws, r, "GROSS PROFIT", [m["gross_profit"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "OPERATING EXPENSES", span)
    _row(ws, r, "Selling Expenses", [y.op_statement.selling_expenses for y in years]); r += 1
    _row(ws, r, "Administrative Expenses", [y.op_statement.admin_expenses for y in years]); r += 1
    _row(ws, r, "Operating Profit (before interest)", [m["operating_profit"] for m in metrics], total=True); r += 1
    _row(ws, r, "Interest on Bank Borrowings", [y.op_statement.interest_on_bank_borrowing for y in years]); r += 1
    _row(ws, r, "Interest on Others", [y.op_statement.interest_on_others for y in years]); r += 1
    _row(ws, r, "Profit Before Tax (PBT)", [m["pbt"] for m in metrics], total=True); r += 1
    _row(ws, r, "Provision for Tax", [m["tax"] for m in metrics]); r += 1
    _row(ws, r, "PROFIT AFTER TAX (PAT)", [m["pat"] for m in metrics], total=True); r += 1
    _row(ws, r, "Cash Accrual (PAT + Depreciation)", [m["cash_accrual"] for m in metrics], bold=True); r += 1

    # ============ SHEET 3: Form III — Balance Sheet ============
    ws = wb.create_sheet("Form III - Balance Sheet")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "LIABILITIES", span)
    r = _section(ws, r, "  Current Liabilities", span)
    _row(ws, r, "Short-term Borrowing from Banks (CC/OD/WC)", [y.balance_sheet.short_term_borrowing_banks for y in years]); r += 1
    _row(ws, r, "Sundry Creditors", [y.balance_sheet.sundry_creditors for y in years]); r += 1
    _row(ws, r, "Advance from Customers", [y.balance_sheet.advance_from_customers for y in years]); r += 1
    _row(ws, r, "Provision for Tax", [y.balance_sheet.provision_for_tax for y in years]); r += 1
    _row(ws, r, "Other Current Liabilities", [y.balance_sheet.other_current_liabilities for y in years]); r += 1
    _row(ws, r, "Total Current Liabilities", [m["current_liab"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "  Term Liabilities", span)
    _row(ws, r, "Debentures", [y.balance_sheet.debentures for y in years]); r += 1
    _row(ws, r, "Term Loan from Banks", [y.balance_sheet.term_loan_banks for y in years]); r += 1
    _row(ws, r, "Term Loan from Others", [y.balance_sheet.term_loan_others for y in years]); r += 1
    _row(ws, r, "Deferred Payment Credits", [y.balance_sheet.deferred_payment_credits for y in years]); r += 1
    _row(ws, r, "Other Term Liabilities", [y.balance_sheet.other_term_liabilities for y in years]); r += 1
    _row(ws, r, "Total Term Liabilities", [m["term_liab"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "  Net Worth", span)
    _row(ws, r, "Paid-up Capital", [y.balance_sheet.paid_up_capital for y in years]); r += 1
    _row(ws, r, "Reserves & Surplus", [y.balance_sheet.reserves_surplus for y in years]); r += 1
    _row(ws, r, "Revaluation Reserves", [y.balance_sheet.revaluation_reserve for y in years]); r += 1
    _row(ws, r, "Other Reserves", [y.balance_sheet.other_reserves for y in years]); r += 1
    _row(ws, r, "Total Net Worth", [m["net_worth"] for m in metrics], total=True); r += 1
    _row(ws, r, "TOTAL LIABILITIES", [m["total_liab"] for m in metrics], total=True); r += 1

    r = _section(ws, r, "ASSETS", span)
    r = _section(ws, r, "  Current Assets", span)
    _row(ws, r, "Cash & Bank Balances", [y.balance_sheet.cash_bank_balances for y in years]); r += 1
    _row(ws, r, "Short-term Investments", [y.balance_sheet.investments_short_term for y in years]); r += 1
    _row(ws, r, "Receivables — Domestic", [y.balance_sheet.receivables_domestic for y in years]); r += 1
    _row(ws, r, "Receivables — Export", [y.balance_sheet.receivables_export for y in years]); r += 1
    _row(ws, r, "Inventory — Raw Material", [y.balance_sheet.inventory_rm for y in years]); r += 1
    _row(ws, r, "Inventory — WIP", [y.balance_sheet.inventory_wip for y in years]); r += 1
    _row(ws, r, "Inventory — Finished Goods", [y.balance_sheet.inventory_fg for y in years]); r += 1
    _row(ws, r, "Advance to Suppliers", [y.balance_sheet.advance_to_suppliers for y in years]); r += 1
    _row(ws, r, "Other Current Assets", [y.balance_sheet.other_current_assets for y in years]); r += 1
    _row(ws, r, "Total Current Assets", [m["current_assets"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "  Fixed Assets", span)
    _row(ws, r, "Gross Block", [y.balance_sheet.gross_block for y in years]); r += 1
    _row(ws, r, "Less: Accumulated Depreciation", [y.balance_sheet.accumulated_depreciation for y in years]); r += 1
    _row(ws, r, "Net Block", [m["net_block"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Capital WIP", [y.balance_sheet.capital_wip for y in years]); r += 1
    _row(ws, r, "Total Fixed Assets", [m["fixed_assets_total"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "  Other Non-Current Assets", span)
    _row(ws, r, "Long-term Investments", [y.balance_sheet.investments_long_term for y in years]); r += 1
    _row(ws, r, "Intangible Assets", [y.balance_sheet.intangible_assets for y in years]); r += 1
    _row(ws, r, "Deferred Revenue Expenditure", [y.balance_sheet.deferred_revenue_expenditure for y in years]); r += 1
    _row(ws, r, "Other Non-Current Assets", [y.balance_sheet.other_non_current_assets for y in years]); r += 1
    _row(ws, r, "TOTAL ASSETS", [m["total_assets"] for m in metrics], total=True); r += 1

    # ============ SHEET 4: Form IV — Current Assets & Liabilities ============
    ws = wb.create_sheet("Form IV - CA & CL")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "CURRENT ASSETS", span)
    _row(ws, r, "Inventory (RM + WIP + FG)", [m["inventory_total"] for m in metrics]); r += 1
    _row(ws, r, "Receivables (Domestic + Export)", [m["receivables_total"] for m in metrics]); r += 1
    _row(ws, r, "Cash & Bank", [y.balance_sheet.cash_bank_balances for y in years]); r += 1
    _row(ws, r, "Short-term Investments", [y.balance_sheet.investments_short_term for y in years]); r += 1
    _row(ws, r, "Advance to Suppliers", [y.balance_sheet.advance_to_suppliers for y in years]); r += 1
    _row(ws, r, "Other Current Assets", [y.balance_sheet.other_current_assets for y in years]); r += 1
    _row(ws, r, "TOTAL CURRENT ASSETS", [m["current_assets"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "CURRENT LIABILITIES", span)
    _row(ws, r, "Short-term Borrowing from Banks", [y.balance_sheet.short_term_borrowing_banks for y in years]); r += 1
    _row(ws, r, "Sundry Creditors", [y.balance_sheet.sundry_creditors for y in years]); r += 1
    _row(ws, r, "Advance from Customers", [y.balance_sheet.advance_from_customers for y in years]); r += 1
    _row(ws, r, "Provision for Tax", [y.balance_sheet.provision_for_tax for y in years]); r += 1
    _row(ws, r, "Other Current Liabilities", [y.balance_sheet.other_current_liabilities for y in years]); r += 1
    _row(ws, r, "TOTAL CURRENT LIABILITIES", [m["current_liab"] for m in metrics], total=True); r += 1
    _row(ws, r, "Other Current Liabilities (excl. Bank Borrowing)", [m["other_current_liab"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Net Working Capital (NWC)", [m["nwc"] for m in metrics], total=True); r += 1
    _row(ws, r, "Working Capital Gap (CA – OCL)", [m["wcg"] for m in metrics], total=True); r += 1

    # ============ SHEET 5: Form V — MPBF ============
    ws = wb.create_sheet("Form V - MPBF")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "MPBF CALCULATION (Tandon Committee)", span)
    _row(ws, r, "Total Current Assets (CA)", [m["current_assets"] for m in metrics], bold=True); r += 1
    _row(ws, r, "Other Current Liabilities (OCL, excl. Bank)", [m["other_current_liab"] for m in metrics]); r += 1
    _row(ws, r, "Working Capital Gap (WCG = CA – OCL)", [m["wcg"] for m in metrics], total=True); r += 1
    _row(ws, r, "Net Working Capital (NWC)", [m["nwc"] for m in metrics]); r += 1
    r = _section(ws, r, "METHOD I  — MPBF = 0.75 × WCG", span)
    _row(ws, r, "75% of WCG", [0.75 * m["wcg"] for m in metrics]); r += 1
    _row(ws, r, "Less: Actual NWC", [m["nwc"] for m in metrics]); r += 1
    _row(ws, r, "MPBF (Method I)", [max(0, 0.75 * m["wcg"] - m["nwc"]) for m in metrics], total=True); r += 1
    _row(ws, r, "Min of (0.75×WCG) [if NWC provided fully]", [m["mpbf_method_1"] for m in metrics]); r += 1
    r = _section(ws, r, "METHOD II — MPBF = 0.75 × CA – OCL", span)
    _row(ws, r, "75% of Current Assets", [0.75 * m["current_assets"] for m in metrics]); r += 1
    _row(ws, r, "Less: Other Current Liabilities", [m["other_current_liab"] for m in metrics]); r += 1
    _row(ws, r, "Less: Actual NWC (25% of CA)", [0.25 * m["current_assets"] for m in metrics]); r += 1
    _row(ws, r, "MPBF (Method II)", [max(0, 0.75 * m["current_assets"] - m["other_current_liab"] - 0.25 * m["current_assets"]) for m in metrics], total=True); r += 1
    _row(ws, r, "Method II (gross): 0.75×CA – OCL", [m["mpbf_method_2"] for m in metrics]); r += 1

    # ============ SHEET 6: Form VI — Fund Flow ============
    ws = wb.create_sheet("Form VI - Fund Flow")
    ws.column_dimensions["A"].width = 50
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 22
    r = 1
    ws.cell(row=r, column=1, value="FUND FLOW STATEMENT (between consecutive years)")
    _h(ws.cell(row=r, column=1))
    r += 2
    # Build fund flows
    ffs = []
    for i in range(1, len(years)):
        ff = compute_fund_flow(metrics[i - 1], metrics[i], years[i - 1], years[i])
        ffs.append((years[i - 1].year_label, years[i].year_label, ff))
    if not ffs:
        ws.cell(row=r, column=1, value="Need at least 2 years to compute fund flow.")
        _label(ws.cell(row=r, column=1))
    else:
        # header
        ws.cell(row=r, column=1, value="Particulars")
        _h(ws.cell(row=r, column=1))
        for i, (a, b, _) in enumerate(ffs):
            c = ws.cell(row=r, column=2 + i, value=f"{a} → {b}")
            _h(c)
        ws.row_dimensions[r].height = 32
        r += 1
        # Collect all source/use labels
        all_src_labels: List[str] = []
        for _, _, ff in ffs:
            for k in ff["sources"].keys():
                if k not in all_src_labels:
                    all_src_labels.append(k)
        all_use_labels: List[str] = []
        for _, _, ff in ffs:
            for k in ff["uses"].keys():
                if k not in all_use_labels:
                    all_use_labels.append(k)
        # SOURCES
        c = ws.cell(row=r, column=1, value="SOURCES OF FUNDS"); _sh(c)
        for i in range(len(ffs)):
            _sh(ws.cell(row=r, column=2 + i, value=""))
        r += 1
        for label in all_src_labels:
            ws.cell(row=r, column=1, value=label); _label(ws.cell(row=r, column=1))
            for i, (_, _, ff) in enumerate(ffs):
                v = ff["sources"].get(label, 0.0)
                ws.cell(row=r, column=2 + i, value=round(v, 2)); _b(ws.cell(row=r, column=2 + i))
            r += 1
        ws.cell(row=r, column=1, value="TOTAL SOURCES"); _total(ws.cell(row=r, column=1), value=False)
        for i, (_, _, ff) in enumerate(ffs):
            ws.cell(row=r, column=2 + i, value=round(ff["total_sources"], 2)); _total(ws.cell(row=r, column=2 + i))
        r += 2
        # USES
        c = ws.cell(row=r, column=1, value="USES OF FUNDS"); _sh(c)
        for i in range(len(ffs)):
            _sh(ws.cell(row=r, column=2 + i, value=""))
        r += 1
        for label in all_use_labels:
            ws.cell(row=r, column=1, value=label); _label(ws.cell(row=r, column=1))
            for i, (_, _, ff) in enumerate(ffs):
                v = ff["uses"].get(label, 0.0)
                ws.cell(row=r, column=2 + i, value=round(v, 2)); _b(ws.cell(row=r, column=2 + i))
            r += 1
        ws.cell(row=r, column=1, value="TOTAL USES"); _total(ws.cell(row=r, column=1), value=False)
        for i, (_, _, ff) in enumerate(ffs):
            ws.cell(row=r, column=2 + i, value=round(ff["total_uses"], 2)); _total(ws.cell(row=r, column=2 + i))
        r += 1
        ws.cell(row=r, column=1, value="SURPLUS / (DEFICIT)"); _total(ws.cell(row=r, column=1), value=False)
        for i, (_, _, ff) in enumerate(ffs):
            ws.cell(row=r, column=2 + i, value=round(ff["surplus_deficit"], 2)); _total(ws.cell(row=r, column=2 + i))

    # ============ SHEET 7: Form VII — Summary of Financial Statements for Ratio Analysis ============
    ws = wb.create_sheet("Form VII - Summary")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "OPERATING STATEMENT — INCOME", span)
    _row(ws, r, "(1) Domestic Sales", [y.op_statement.domestic_sales for y in years]); r += 1
    _row(ws, r, "(2) Export Sales", [y.op_statement.export_sales for y in years]); r += 1
    _row(ws, r, "(3) Gross Sales (1+2)", [m["gross_sales"] for m in metrics], bold=True); r += 1
    _row(ws, r, "(4) Less: Excise / GST", [y.op_statement.less_excise for y in years]); r += 1
    _row(ws, r, "(5) Net Sales (3-4)", [m["net_sales"] for m in metrics], total=True); r += 1
    _row(ws, r, "(6) Other Income", [y.op_statement.other_income for y in years]); r += 1
    _row(ws, r, "(7) Gross Income (5+6)", [m["total_revenue"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "OPERATING STATEMENT — EXPENSES", span)
    _row(ws, r, "(1) Raw Material Consumed", [m["rm_consumed"] for m in metrics]); r += 1
    _row(ws, r, "(2) Power & Fuel", [y.op_statement.power_fuel for y in years]); r += 1
    _row(ws, r, "(3) Direct Labour", [y.op_statement.direct_labour for y in years]); r += 1
    _row(ws, r, "(4) Other Mfg. Expenses", [y.op_statement.other_mfg_expenses for y in years]); r += 1
    _row(ws, r, "(5) Depreciation", [y.op_statement.depreciation for y in years]); r += 1
    _row(ws, r, "(6) Total Cost of Production", [m["cost_of_production"] for m in metrics], bold=True); r += 1
    _row(ws, r, "(7) Total Cost of Sales", [m["cost_of_sales"] for m in metrics], total=True); r += 1
    _row(ws, r, "(8) Gross Profit", [m["gross_profit"] for m in metrics], total=True); r += 1
    _row(ws, r, "(9) Selling + Admin Expenses", [m["operating_expenses"] for m in metrics]); r += 1
    _row(ws, r, "(10) Operating Profit before Finance Charges (PBIT)", [m["operating_profit"] for m in metrics], total=True); r += 1
    _row(ws, r, "(11) Finance Charges (Interest)", [m["interest_total"] for m in metrics]); r += 1
    _row(ws, r, "(12) Operating Profit after Finance Charges", [m["operating_profit"] - m["interest_total"] for m in metrics], bold=True); r += 1
    _row(ws, r, "(13) Profit Before Tax (PBT)", [m["pbt"] for m in metrics], total=True); r += 1
    _row(ws, r, "(14) Provision for Tax", [m["tax"] for m in metrics]); r += 1
    _row(ws, r, "(15) Profit After Tax (PAT)", [m["pat"] for m in metrics], total=True); r += 1
    _row(ws, r, "(16) Cash Profit (PBDIT)", [m["pbdit"] for m in metrics]); r += 1
    _row(ws, r, "(17) Profit before Interest & Tax (PBIT)", [m["pbit"] for m in metrics]); r += 1
    r = _section(ws, r, "BALANCE SHEET — LIABILITIES", span)
    _row(ws, r, "(1) Short-term Borrowings from Banks", [y.balance_sheet.short_term_borrowing_banks for y in years]); r += 1
    _row(ws, r, "(2) Sundry Creditors", [y.balance_sheet.sundry_creditors for y in years]); r += 1
    _row(ws, r, "(3) Other Current Liabilities", [m["other_current_liab"] - y.balance_sheet.sundry_creditors for y, m in zip(years, metrics)]); r += 1
    _row(ws, r, "(4) Total Current Liabilities", [m["current_liab"] for m in metrics], total=True); r += 1
    _row(ws, r, "(5) Term Liabilities", [m["term_liab"] for m in metrics], total=True); r += 1
    _row(ws, r, "(6) Net Worth", [m["net_worth"] for m in metrics], total=True); r += 1
    _row(ws, r, "(7) Tangible Net Worth", [m["tangible_net_worth"] for m in metrics], total=True); r += 1
    r = _section(ws, r, "BALANCE SHEET — ASSETS", span)
    _row(ws, r, "(1) Cash & Bank + Investments", [y.balance_sheet.cash_bank_balances + y.balance_sheet.investments_short_term for y in years]); r += 1
    _row(ws, r, "(2) Receivables (Domestic + Export)", [m["receivables_total"] for m in metrics]); r += 1
    _row(ws, r, "(3) Inventory (RM+WIP+FG)", [m["inventory_total"] for m in metrics]); r += 1
    _row(ws, r, "(4) Other Current Assets", [m["current_assets"] - m["receivables_total"] - m["inventory_total"] - y.balance_sheet.cash_bank_balances - y.balance_sheet.investments_short_term for y, m in zip(years, metrics)]); r += 1
    _row(ws, r, "(5) Total Current Assets", [m["current_assets"] for m in metrics], total=True); r += 1
    _row(ws, r, "(6) Gross Block", [y.balance_sheet.gross_block for y in years]); r += 1
    _row(ws, r, "(7) Less: Accumulated Depreciation", [y.balance_sheet.accumulated_depreciation for y in years]); r += 1
    _row(ws, r, "(8) Net Block + Capital WIP", [m["fixed_assets_total"] for m in metrics], bold=True); r += 1
    _row(ws, r, "(9) Other Non-Current Assets", [m["non_current_assets"] - m["fixed_assets_total"] for m in metrics]); r += 1
    _row(ws, r, "(10) Intangible Assets", [y.balance_sheet.intangible_assets for y in years]); r += 1
    _row(ws, r, "(11) Total Assets", [m["total_assets"] for m in metrics], total=True); r += 1
    _row(ws, r, "(12) Capital Employed (NW + TL)", [m["capital_employed"] for m in metrics], total=True); r += 1
    _row(ws, r, "(13) Net Working Capital (NWC)", [m["nwc"] for m in metrics], total=True); r += 1

    # ============ SHEET 8: Form VIII — Ratio Analysis (categorized) ============
    ws = wb.create_sheet("Form VIII - Ratios")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "(A) LONG-TERM SOLVENCY RATIOS", span)
    _row(ws, r, "(1) Debt Equity Ratio (TL / TNW)  [x]", [m["debt_equity"] for m in metrics]); r += 1
    _row(ws, r, "(2) Net Worth to Total Assets  [%]", [m["net_worth_to_total_assets"] for m in metrics]); r += 1
    _row(ws, r, "(3) Debt to Net Worth  [x]", [m["debt_to_net_worth"] for m in metrics]); r += 1
    _row(ws, r, "(4) Capital Gearing Ratio  [x]", [m["capital_gearing"] for m in metrics]); r += 1
    _row(ws, r, "(5) Fixed Assets to Long-Term Funds  [x]", [m["fixed_assets_to_long_term"] for m in metrics]); r += 1
    _row(ws, r, "(6) Proprietary Ratio (NW / TA)  [%]", [m["proprietary_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(7) Interest Coverage (PBDIT / Int.)  [x]", [m["interest_coverage"] for m in metrics]); r += 1
    _row(ws, r, "(8) Debt Service Coverage (DSCR)  [x]", [m["dscr"] for m in metrics]); r += 1
    _row(ws, r, "(9) TOL / TNW  [x]", [m["tol_tnw"] for m in metrics]); r += 1
    r = _section(ws, r, "(B) SHORT-TERM SOLVENCY RATIOS", span)
    _row(ws, r, "(1) Current Ratio (CA / CL)  [x]", [m["current_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(2) Quick / Acid Test Ratio  [x]", [m["quick_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(3) Absolute Liquid Ratio  [x]", [m["absolute_liquid_ratio"] for m in metrics]); r += 1
    r = _section(ws, r, "(C) PROFITABILITY RATIOS", span)
    _row(ws, r, "(1) Return on Capital Employed (ROCE)  [%]", [m["roce"] for m in metrics]); r += 1
    _row(ws, r, "(2) Gross Profit Margin  [%]", [m["gp_margin"] for m in metrics]); r += 1
    _row(ws, r, "(3) Net Profit Margin (PAT / Net Sales)  [%]", [m["np_margin"] for m in metrics]); r += 1
    _row(ws, r, "(4) Cash Profit Ratio  [%]", [m["cash_profit_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(5) Return on Net Worth (PAT / NW)  [%]", [m["return_on_net_worth"] for m in metrics]); r += 1
    _row(ws, r, "(6) Operating Profit Margin (before int.)  [%]", [m["op_margin"] for m in metrics]); r += 1
    _row(ws, r, "(7) Operating Profit Margin (after int.)  [%]", [m["op_margin_after_int"] for m in metrics]); r += 1
    r = _section(ws, r, "(D) ACTIVITY RATIOS", span)
    _row(ws, r, "(1) Inventory Turnover (Sales / Inv.)  [x]", [m["stock_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(2) Debtors Turnover (Sales / Recv.)  [x]", [m["debtors_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(3) Creditors Turnover (Purchases / Cred.)  [x]", [m["creditors_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(4) Debtors Turnover Period  [days]", [m["debtors_velocity_days"] for m in metrics]); r += 1
    _row(ws, r, "(5) Creditors Turnover Period  [days]", [m["creditors_velocity_days"] for m in metrics]); r += 1
    _row(ws, r, "(6) Fixed Assets Turnover  [x]", [m["fixed_assets_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(7) Total Assets Turnover  [x]", [m["assets_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(8) Working Capital Turnover  [x]", [m["working_capital_turnover"] for m in metrics]); r += 1
    _row(ws, r, "(9) Sales to Capital Employed  [x]", [m["sales_to_capital_employed"] for m in metrics]); r += 1
    r = _section(ws, r, "(E) OPERATING RATIOS (% of Net Sales)", span)
    _row(ws, r, "(1) Domestic Sales Proportion  [%]", [m["domestic_sales_pct"] for m in metrics]); r += 1
    _row(ws, r, "(2) Export Sales Proportion  [%]", [m["export_sales_pct"] for m in metrics]); r += 1
    _row(ws, r, "(3) Material Cost Ratio  [%]", [m["material_cost_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(4) Direct Labour Cost Ratio  [%]", [m["direct_labour_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(5) Other Overheads Ratio  [%]", [m["other_overheads_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(6) Indirect Cost Ratio  [%]", [m["indirect_cost_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(7) Interest Cost Ratio  [%]", [m["interest_cost_ratio"] for m in metrics]); r += 1
    _row(ws, r, "(8) Operating Cost Ratio  [%]", [m["operating_cost_ratio"] for m in metrics]); r += 1

    # ============ SHEET 9: Form IX — Statement of Changes in Working Capital ============
    ws = wb.create_sheet("Form IX - Changes in WC")
    r = _write_year_header(ws, 1, years)
    _set_widths(ws)
    r = _section(ws, r, "CURRENT ASSETS", span)
    _row(ws, r, "(1) Cash and Bank Balances", [y.balance_sheet.cash_bank_balances for y in years]); r += 1
    _row(ws, r, "(2) Investments — Short-term", [y.balance_sheet.investments_short_term for y in years]); r += 1
    _row(ws, r, "(3) Receivables — Domestic", [y.balance_sheet.receivables_domestic for y in years]); r += 1
    _row(ws, r, "(4) Receivables — Export", [y.balance_sheet.receivables_export for y in years]); r += 1
    _row(ws, r, "(5) Inventory — Raw Material", [y.balance_sheet.inventory_rm for y in years]); r += 1
    _row(ws, r, "(6) Inventory — WIP", [y.balance_sheet.inventory_wip for y in years]); r += 1
    _row(ws, r, "(7) Inventory — Finished Goods", [y.balance_sheet.inventory_fg for y in years]); r += 1
    _row(ws, r, "(8) Advances to Suppliers", [y.balance_sheet.advance_to_suppliers for y in years]); r += 1
    _row(ws, r, "(9) Other Current Assets", [y.balance_sheet.other_current_assets for y in years]); r += 1
    _row(ws, r, "(10) Total Current Assets (TCA)", [m["current_assets"] for m in metrics], total=True); r += 1
    # Change in TCA year-on-year
    tca_change = [0.0] + [metrics[i]["current_assets"] - metrics[i - 1]["current_assets"] for i in range(1, len(metrics))]
    _row(ws, r, "(11) Change in Current Assets (YoY)", tca_change, bold=True); r += 1
    r = _section(ws, r, "CURRENT LIABILITIES", span)
    _row(ws, r, "(1) Short-term Borrowings from Banks", [y.balance_sheet.short_term_borrowing_banks for y in years]); r += 1
    _row(ws, r, "(2) Sundry Creditors", [y.balance_sheet.sundry_creditors for y in years]); r += 1
    _row(ws, r, "(3) Advance from Customers", [y.balance_sheet.advance_from_customers for y in years]); r += 1
    _row(ws, r, "(4) Provision for Tax", [y.balance_sheet.provision_for_tax for y in years]); r += 1
    _row(ws, r, "(5) Other Current Liabilities", [y.balance_sheet.other_current_liabilities for y in years]); r += 1
    _row(ws, r, "(6) Total Current Liabilities (TCL)", [m["current_liab"] for m in metrics], total=True); r += 1
    tcl_change = [0.0] + [metrics[i]["current_liab"] - metrics[i - 1]["current_liab"] for i in range(1, len(metrics))]
    _row(ws, r, "(7) Change in Current Liabilities (YoY)", tcl_change, bold=True); r += 1
    r = _section(ws, r, "NET WORKING CAPITAL", span)
    _row(ws, r, "(8) Net Working Capital (TCA – TCL)", [m["nwc"] for m in metrics], total=True); r += 1
    nwc_change = [0.0] + [metrics[i]["nwc"] - metrics[i - 1]["nwc"] for i in range(1, len(metrics))]
    _row(ws, r, "(9) Increase / (Decrease) in NWC (YoY)", nwc_change, total=True); r += 1
    _row(ws, r, "(10) Working Capital Gap (CA – OCL)", [m["wcg"] for m in metrics], bold=True); r += 1

    # ============ SHEET 10: Sensitivity Analysis (optional — preserves existing auto-gen) ============
    if stmt.sensitivity.enabled and stmt.years:
        ws = wb.create_sheet("Sensitivity Analysis")
        sens = compute_sensitivity(stmt)
        target_label = sens.get("target_year_label", "—")
        ws.cell(row=1, column=1, value=f"SENSITIVITY ANALYSIS — Target Year: {target_label}")
        _h(ws.cell(row=1, column=1))
        ws.row_dimensions[1].height = 28
        ws.column_dimensions["A"].width = 46
        n_sc = len(sens["scenarios"])
        for i in range(n_sc):
            ws.column_dimensions[get_column_letter(2 + i)].width = 22
        r = 3
        # header
        ws.cell(row=r, column=1, value="Particulars")
        _h(ws.cell(row=r, column=1))
        for i, sc in enumerate(sens["scenarios"]):
            ws.cell(row=r, column=2 + i, value=sc["name"])
            _h(ws.cell(row=r, column=2 + i))
        ws.row_dimensions[r].height = 32
        r += 1
        # deltas
        for delta_label, key in [
            ("Sales Δ (%)", "sales_delta_pct"),
            ("Raw Material Cost Δ (%)", "rm_cost_delta_pct"),
            ("Interest Δ (%)", "interest_delta_pct"),
            ("Other Expenses Δ (%)", "other_expenses_delta_pct"),
        ]:
            ws.cell(row=r, column=1, value=delta_label)
            _label(ws.cell(row=r, column=1))
            for i, sc in enumerate(sens["scenarios"]):
                ws.cell(row=r, column=2 + i, value=round(float(sc["deltas"].get(key, 0)), 2))
                _b(ws.cell(row=r, column=2 + i))
            r += 1
        # blank row
        r += 1
        # metrics
        ws.cell(row=r, column=1, value="RESULTING METRICS (₹ in Lakhs)")
        _sh(ws.cell(row=r, column=1))
        for i in range(n_sc):
            _sh(ws.cell(row=r, column=2 + i))
        r += 1
        for metric_label, mkey, is_total in [
            ("Net Sales", "net_sales", False),
            ("Gross Profit", "gross_profit", False),
            ("Operating Profit", "operating_profit", True),
            ("Interest", "interest_total", False),
            ("PBT", "pbt", False),
            ("Tax", "tax", False),
            ("PAT", "pat", True),
            ("Cash Accrual", "cash_accrual", True),
            ("Net Profit Margin (%)", "np_margin", False),
            ("Interest Coverage (x)", "interest_coverage", False),
            ("DSCR (x)", "dscr", True),
        ]:
            ws.cell(row=r, column=1, value=metric_label)
            if is_total:
                _total(ws.cell(row=r, column=1), value=False)
            else:
                _label(ws.cell(row=r, column=1))
            for i, sc in enumerate(sens["scenarios"]):
                v = sc["metrics"].get(mkey, 0)
                ws.cell(row=r, column=2 + i, value=round(float(v), 2))
                if is_total:
                    _total(ws.cell(row=r, column=2 + i))
                else:
                    _b(ws.cell(row=r, column=2 + i))
            r += 1

    # =========== LOCK ALL SHEETS (read-only for downstream users) ===========
    _EXCEL_LOCK_PASSWORD = "dprforge-cma"
    for sh in wb.worksheets:
        sh.protection = SheetProtection(
            sheet=True, password=_EXCEL_LOCK_PASSWORD,
            objects=True, scenarios=True,
            formatCells=True, formatColumns=True, formatRows=True,
            insertColumns=True, insertRows=True, insertHyperlinks=True,
            deleteColumns=True, deleteRows=True,
            sort=True, autoFilter=True, pivotTables=True,
            selectLockedCells=False, selectUnlockedCells=False,
        )
        sh.protection.enable()
    # Workbook structure also locked so user can't add/remove sheets
    wb.security = WorkbookProtection(workbookPassword=_EXCEL_LOCK_PASSWORD, lockStructure=True, lockWindows=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_blank_template(audited: int = 2, provisional: int = 1, projected: int = 3) -> bytes:
    """Blank Excel template that user can fill and re-upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "CMA DATA — BLANK ENTRY TEMPLATE"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A2"] = "Fill the numbers in the OpStatement and BalanceSheet sheets. Then upload back to your CMA project."
    ws["A2"].font = Font(italic=True, color="475569")
    ws["A4"] = "Notes:"
    ws["A4"].font = Font(bold=True)
    notes = [
        "• Enter all figures in ₹ Lakhs (recommended).",
        "• Year columns: audited / provisional / projected — rename labels in row 1 if needed.",
        "• Sheets: 'OpStatement' (P&L) and 'BalanceSheet'. The first column lists items.",
        "• Do not change item labels in column A — they are matched on upload.",
        "• You can also fill 'ExistingLimits' for Form I.",
    ]
    for i, t in enumerate(notes):
        ws.cell(row=5 + i, column=1, value=t)
    ws.column_dimensions["A"].width = 100

    # Build column headers
    year_types: List[str] = (["audited"] * audited) + (["provisional"] * provisional) + (["projected"] * projected)
    year_labels: List[str] = []
    current = "FY 2022-23"
    for t in year_types:
        year_labels.append(current)
        current = _next_fy_label(current)

    op_fields = [
        ("domestic_sales", "Domestic Sales"),
        ("export_sales", "Export Sales"),
        ("less_excise", "Less: Excise / GST"),
        ("other_income", "Other Income"),
        ("opening_stock_rm", "Opening Stock — Raw Material"),
        ("raw_materials_purchase", "Purchases — Raw Material"),
        ("closing_stock_rm", "Closing Stock — Raw Material"),
        ("power_fuel", "Power & Fuel"),
        ("direct_labour", "Direct Labour / Wages"),
        ("other_mfg_expenses", "Other Manufacturing Expenses"),
        ("depreciation", "Depreciation"),
        ("opening_stock_wip", "Opening WIP"),
        ("closing_stock_wip", "Closing WIP"),
        ("opening_stock_fg", "Opening Finished Goods"),
        ("closing_stock_fg", "Closing Finished Goods"),
        ("selling_expenses", "Selling Expenses"),
        ("admin_expenses", "Administrative Expenses"),
        ("interest_on_bank_borrowing", "Interest on Bank Borrowings"),
        ("interest_on_others", "Interest on Others"),
        ("tax_rate", "Tax Rate (%)"),
    ]
    bs_fields = [
        ("short_term_borrowing_banks", "Short-term Borrowing from Banks"),
        ("sundry_creditors", "Sundry Creditors"),
        ("advance_from_customers", "Advance from Customers"),
        ("provision_for_tax", "Provision for Tax"),
        ("other_current_liabilities", "Other Current Liabilities"),
        ("debentures", "Debentures"),
        ("term_loan_banks", "Term Loan — Banks"),
        ("term_loan_others", "Term Loan — Others"),
        ("deferred_payment_credits", "Deferred Payment Credits"),
        ("other_term_liabilities", "Other Term Liabilities"),
        ("paid_up_capital", "Paid-up Capital"),
        ("reserves_surplus", "Reserves & Surplus"),
        ("revaluation_reserve", "Revaluation Reserve"),
        ("other_reserves", "Other Reserves"),
        ("cash_bank_balances", "Cash & Bank Balances"),
        ("investments_short_term", "Short-term Investments"),
        ("receivables_domestic", "Receivables — Domestic"),
        ("receivables_export", "Receivables — Export"),
        ("inventory_rm", "Inventory — Raw Material"),
        ("inventory_wip", "Inventory — WIP"),
        ("inventory_fg", "Inventory — Finished Goods"),
        ("advance_to_suppliers", "Advance to Suppliers"),
        ("other_current_assets", "Other Current Assets"),
        ("gross_block", "Gross Block"),
        ("accumulated_depreciation", "Accumulated Depreciation"),
        ("capital_wip", "Capital WIP"),
        ("investments_long_term", "Long-term Investments"),
        ("intangible_assets", "Intangible Assets"),
        ("deferred_revenue_expenditure", "Deferred Revenue Expenditure"),
        ("other_non_current_assets", "Other Non-Current Assets"),
    ]

    def _build_sheet(name: str, fields: List, defaults: Optional[Dict] = None):
        s = wb.create_sheet(name)
        s.cell(row=1, column=1, value="key")
        s.cell(row=1, column=2, value="Particulars")
        _h(s.cell(row=1, column=1)); _h(s.cell(row=1, column=2))
        for i, lbl in enumerate(year_labels):
            c = s.cell(row=1, column=3 + i, value=f"{lbl} ({year_types[i]})")
            _h(c)
        for r_i, (k, label) in enumerate(fields):
            s.cell(row=2 + r_i, column=1, value=k)
            s.cell(row=2 + r_i, column=2, value=label)
            _label(s.cell(row=2 + r_i, column=1))
            _label(s.cell(row=2 + r_i, column=2))
            for j in range(len(year_labels)):
                s.cell(row=2 + r_i, column=3 + j, value=0)
                _b(s.cell(row=2 + r_i, column=3 + j))
        s.column_dimensions["A"].width = 32
        s.column_dimensions["B"].width = 36
        for j in range(len(year_labels)):
            s.column_dimensions[get_column_letter(3 + j)].width = 18

    _build_sheet("OpStatement", op_fields)
    _build_sheet("BalanceSheet", bs_fields)

    # ExistingLimits sheet
    s = wb.create_sheet("ExistingLimits")
    headers = ["facility", "bank", "nature", "existing_limit", "outstanding", "proposed_limit", "security", "rate_of_interest", "margin_pct"]
    for i, h in enumerate(headers):
        c = s.cell(row=1, column=i + 1, value=h)
        _h(c)
    # 3 empty rows
    for r_i in range(2, 6):
        for j in range(len(headers)):
            _label(s.cell(row=r_i, column=j + 1))
    for i in range(len(headers)):
        s.column_dimensions[get_column_letter(i + 1)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_uploaded_excel(content: bytes) -> Dict[str, Any]:
    """Read an uploaded blank-template and reconstruct years + existing_limits."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if "OpStatement" not in wb.sheetnames or "BalanceSheet" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Invalid template: 'OpStatement' or 'BalanceSheet' sheet missing.")

    def _read_sheet(ws) -> (List[str], List[str], List[Dict[str, float]]):
        # row 1: key | Particulars | year_label1 | year_label2 ...
        header = [c.value for c in ws[1]]
        n_year_cols = len(header) - 2
        year_headers = [str(header[2 + i]) if header[2 + i] else f"Year {i+1}" for i in range(n_year_cols)]
        # Detect year types from header text "(audited)" suffix
        year_types = []
        year_labels = []
        for h in year_headers:
            h_low = h.lower()
            if "(audited)" in h_low or "audited" in h_low:
                year_types.append("audited")
            elif "(provisional)" in h_low or "provisional" in h_low:
                year_types.append("provisional")
            elif "(projected)" in h_low or "projected" in h_low:
                year_types.append("projected")
            else:
                year_types.append("audited")
            # strip type suffix
            lbl = h
            for suf in ["(audited)", "(provisional)", "(projected)"]:
                lbl = lbl.replace(suf, "").replace(suf.title(), "")
            year_labels.append(lbl.strip())
        # rows -> dict per year
        per_year: List[Dict[str, float]] = [dict() for _ in range(n_year_cols)]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            key = row[0]
            if not key:
                continue
            for i in range(n_year_cols):
                v = row[2 + i] if 2 + i < len(row) else 0
                try:
                    per_year[i][str(key)] = float(v or 0)
                except (TypeError, ValueError):
                    per_year[i][str(key)] = 0.0
        return year_labels, year_types, per_year

    op_labels, op_types, op_per_year = _read_sheet(wb["OpStatement"])
    bs_labels, bs_types, bs_per_year = _read_sheet(wb["BalanceSheet"])

    n = min(len(op_per_year), len(bs_per_year))
    years: List[CMAYearData] = []
    for i in range(n):
        op = CMAOpStatement(**{k: v for k, v in op_per_year[i].items() if k in CMAOpStatement.model_fields})
        bs = CMABalanceSheet(**{k: v for k, v in bs_per_year[i].items() if k in CMABalanceSheet.model_fields})
        years.append(CMAYearData(
            year_label=op_labels[i] or f"Year {i+1}",
            year_type=op_types[i] if i < len(op_types) else "audited",
            op_statement=op,
            balance_sheet=bs,
        ))

    existing_limits: List[CMAExistingLimits] = []
    if "ExistingLimits" in wb.sheetnames:
        ws = wb["ExistingLimits"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all((c is None or c == "") for c in row):
                continue
            facility = row[0] or ""
            if not facility:
                continue
            try:
                existing_limits.append(CMAExistingLimits(
                    facility=str(facility),
                    bank=str(row[1] or ""),
                    nature=str(row[2] or ""),
                    existing_limit=float(row[3] or 0),
                    outstanding=float(row[4] or 0),
                    proposed_limit=float(row[5] or 0),
                    security=str(row[6] or ""),
                    rate_of_interest=float(row[7] or 0),
                    margin_pct=float(row[8] or 0),
                ))
            except (TypeError, ValueError):
                continue

    return {"years": [y.model_dump() for y in years], "existing_limits": [e.model_dump() for e in existing_limits]}


# =============================== PDF EXPORT ===============================

def generate_cma_pdf(stmt: CMAStatement, *, seller: Optional[Dict[str, Any]] = None) -> bytes:
    """Generate a bank-ready CMA PDF (landscape A4) with branded header.

    `seller` optionally provides company branding (name, gstin, contact) — falls back
    to "Mother Bless Digital Solutions" defaults.
    """
    seller = seller or {
        "name": "Mother Bless Digital Solutions",
        "gstin": "08KQRPS8229A1Z6",
        "primary_phone": "7300213623",
        "email": "motherblessopc@gmail.com",
        "city": "Bagidora",
        "state": "Rajasthan",
    }

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"CMA — {stmt.company_name or 'Statement'}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CMATitle", parent=styles["Title"], fontSize=18,
        textColor=colors.HexColor("#0F172A"), spaceAfter=2, alignment=1,
    )
    sub_title = ParagraphStyle(
        "CMASub", parent=styles["BodyText"], fontSize=9,
        textColor=colors.HexColor("#475569"), alignment=1, spaceAfter=2,
    )
    brand = ParagraphStyle(
        "Brand", parent=styles["BodyText"], fontSize=8,
        textColor=colors.HexColor("#64748B"), alignment=1, spaceAfter=4,
    )
    h2 = ParagraphStyle(
        "CMAH2", parent=styles["Heading2"], fontSize=12,
        textColor=colors.HexColor("#1D4ED8"), spaceBefore=8, spaceAfter=4,
    )
    body = ParagraphStyle(
        "Body", parent=styles["BodyText"], fontSize=9, leading=12,
        textColor=colors.HexColor("#0F172A"),
    )
    small = ParagraphStyle(
        "Small", parent=styles["BodyText"], fontSize=8, leading=10,
        textColor=colors.HexColor("#334155"),
    )
    cell = ParagraphStyle(
        "Cell", parent=styles["BodyText"], fontSize=8, leading=10, alignment=0,
    )
    cell_hdr = ParagraphStyle(
        "CellHdr", parent=styles["BodyText"], fontSize=8.5, leading=10,
        textColor=colors.white, alignment=1, fontName="Helvetica-Bold",
    )

    def _fmt(v):
        try:
            n = float(v)
        except (TypeError, ValueError):
            return str(v) if v is not None else ""
        if n == 0:
            return "—"
        return f"{n:,.2f}"

    def _fmt_pct(v):
        try:
            return f"{float(v):,.2f}%"
        except (TypeError, ValueError):
            return "—"

    def _fmt_x(v):
        try:
            return f"{float(v):,.2f}x"
        except (TypeError, ValueError):
            return "—"

    def _fmt_days(v):
        try:
            return f"{float(v):,.0f} d"
        except (TypeError, ValueError):
            return "—"

    story = []

    # Brand band
    story.append(Paragraph(f"<b>{seller.get('name')}</b>", brand))
    story.append(Paragraph(
        f"GSTIN: {seller.get('gstin', '')} &nbsp;|&nbsp; "
        f"{seller.get('city', '')}, {seller.get('state', '')} &nbsp;|&nbsp; "
        f"{seller.get('primary_phone', '')} &nbsp;|&nbsp; {seller.get('email', '')}",
        brand,
    ))
    story.append(Paragraph("CMA DATA — Credit Monitoring Arrangement", title_style))
    story.append(Paragraph(
        "Prepared as per RBI / Tandon Committee framework — all figures ₹ in Lakhs unless stated",
        sub_title,
    ))
    story.append(Spacer(1, 4 * mm))

    # Company info card (2-column table)
    company_info = [
        [Paragraph("<b>Company / Borrower</b>", small), Paragraph(stmt.company_name or "—", body),
         Paragraph("<b>Constitution</b>", small), Paragraph(stmt.constitution or "—", body)],
        [Paragraph("<b>Banker</b>", small), Paragraph(stmt.banker_name or "—", body),
         Paragraph("<b>Facility Type</b>", small), Paragraph(stmt.facility_type or "—", body)],
        [Paragraph("<b>Registered Address</b>", small),
         Paragraph(stmt.registered_address or "—", body),
         Paragraph("<b>Prepared On</b>", small),
         Paragraph(datetime.now(timezone.utc).strftime("%d %b %Y"), body)],
    ]
    info_tbl = Table(company_info, colWidths=[34 * mm, 95 * mm, 30 * mm, 84 * mm])
    info_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 5 * mm))

    metrics = [compute_year_metrics(y) for y in stmt.years]
    n_years = len(stmt.years)
    # Wrapped year headers using Paragraph (so they word-wrap correctly)
    year_paragraphs = [Paragraph(
        f"<b>{y.year_label}</b><br/><font size=7 color='#CBD5E1'>({y.year_type.title()})</font>",
        cell_hdr,
    ) for y in stmt.years]

    # Compute column widths to fill landscape A4 (~273mm usable)
    PARTICULARS_W = 72 * mm
    USABLE_W = 273 * mm
    YEAR_W = max(22 * mm, (USABLE_W - PARTICULARS_W) / max(1, n_years))

    def _table(rows, col_widths=None, header_label="Particulars"):
        if col_widths is None:
            col_widths = [PARTICULARS_W] + [YEAR_W] * n_years
        t = Table(rows, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # ---------------- Form I — Limits ----------------
    story.append(Paragraph("Form I — Existing & Proposed Credit Limits (₹ in Lakhs)", h2))
    if stmt.existing_limits:
        rows = [[
            Paragraph("<b>S.No.</b>", cell_hdr), Paragraph("<b>Facility</b>", cell_hdr),
            Paragraph("<b>Bank</b>", cell_hdr), Paragraph("<b>Nature</b>", cell_hdr),
            Paragraph("<b>Existing</b>", cell_hdr), Paragraph("<b>Outstanding</b>", cell_hdr),
            Paragraph("<b>Proposed</b>", cell_hdr), Paragraph("<b>Security</b>", cell_hdr),
            Paragraph("<b>ROI %</b>", cell_hdr), Paragraph("<b>Margin %</b>", cell_hdr),
        ]]
        for i, lim in enumerate(stmt.existing_limits):
            rows.append([
                str(i + 1),
                Paragraph(lim.facility or "—", cell), Paragraph(lim.bank or "—", cell),
                Paragraph(lim.nature or "—", cell),
                _fmt(lim.existing_limit), _fmt(lim.outstanding), _fmt(lim.proposed_limit),
                Paragraph(lim.security or "—", cell),
                _fmt(lim.rate_of_interest), _fmt(lim.margin_pct),
            ])
        col_w = [12, 38, 30, 28, 26, 28, 26, 40, 22, 22]  # mm — sums to ~272mm
        story.append(Table(
            rows, colWidths=[x * mm for x in col_w], repeatRows=1,
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("ALIGN", (4, 1), (6, -1), "RIGHT"),
                ("ALIGN", (8, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]),
        ))
    else:
        story.append(Paragraph("<i>No existing / proposed credit limits provided.</i>", small))

    # ---------------- Form II — Operating Statement ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form II — Operating Statement (₹ in Lakhs)", h2))
    op_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_op(label, vals, bold=False, indent=False):
        lbl = ("&nbsp;&nbsp;&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
        op_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_op("Gross Sales", [m["gross_sales"] for m in metrics], bold=True)
    add_op("Less: Excise / GST", [y.op_statement.less_excise for y in stmt.years], indent=True)
    add_op("Net Sales", [m["net_sales"] for m in metrics], bold=True)
    add_op("Add: Other Income", [y.op_statement.other_income for y in stmt.years], indent=True)
    add_op("Raw Material Consumed", [m["rm_consumed"] for m in metrics])
    add_op("Cost of Production", [m["cost_of_production"] for m in metrics])
    add_op("Cost of Sales", [m["cost_of_sales"] for m in metrics], bold=True)
    add_op("Gross Profit", [m["gross_profit"] for m in metrics], bold=True)
    add_op("Operating Expenses (S&D + Admin)", [m["operating_expenses"] for m in metrics])
    add_op("Operating Profit", [m["operating_profit"] for m in metrics], bold=True)
    add_op("Interest", [m["interest_total"] for m in metrics])
    add_op("PBDIT", [m["pbdit"] for m in metrics])
    add_op("PBT", [m["pbt"] for m in metrics])
    add_op("Tax", [m["tax"] for m in metrics])
    add_op("PAT", [m["pat"] for m in metrics], bold=True)
    add_op("Add: Depreciation", [y.op_statement.depreciation for y in stmt.years], indent=True)
    add_op("Cash Accrual", [m["cash_accrual"] for m in metrics], bold=True)
    story.append(_table(op_rows))

    # ---------------- Form III — Balance Sheet ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form III — Balance Sheet Analysis (₹ in Lakhs)", h2))
    bs_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_bs(label, vals, bold=False, indent=False):
        lbl = ("&nbsp;&nbsp;&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
        bs_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_bs("LIABILITIES", [""] * n_years, bold=True)
    add_bs("Current Liabilities", [m["current_liab"] for m in metrics], bold=True, indent=True)
    add_bs("Term Liabilities", [m["term_liab"] for m in metrics], bold=True, indent=True)
    add_bs("Net Worth", [m["net_worth"] for m in metrics], bold=True, indent=True)
    add_bs("Tangible Net Worth", [m["tangible_net_worth"] for m in metrics], indent=True)
    add_bs("Total Liabilities", [m["total_liab"] for m in metrics], bold=True)
    add_bs("ASSETS", [""] * n_years, bold=True)
    add_bs("Current Assets", [m["current_assets"] for m in metrics], bold=True, indent=True)
    add_bs("Inventory", [m["inventory_total"] for m in metrics], indent=True)
    add_bs("Receivables", [m["receivables_total"] for m in metrics], indent=True)
    add_bs("Quick Assets", [m["quick_assets"] for m in metrics], indent=True)
    add_bs("Net Block (Gross − Depn.)", [m["net_block"] for m in metrics], indent=True)
    add_bs("Fixed Assets (Net Block + CWIP)", [m["fixed_assets_total"] for m in metrics], indent=True)
    add_bs("Non-Current Assets (Total)", [m["non_current_assets"] for m in metrics], indent=True)
    add_bs("Total Assets", [m["total_assets"] for m in metrics], bold=True)
    story.append(_table(bs_rows))

    # ---------------- Form IV — Comparative Statement of CA & CL ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form IV — Comparative Statement of Current Assets & Current Liabilities (₹ in Lakhs)", h2))
    ca_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_ca(label, vals, bold=False, indent=False):
        lbl = ("&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
        ca_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_ca("CURRENT ASSETS", [""] * n_years, bold=True)
    add_ca("Cash & Bank Balances", [y.balance_sheet.cash_bank_balances for y in stmt.years], indent=True)
    add_ca("Short-term Investments", [y.balance_sheet.investments_short_term for y in stmt.years], indent=True)
    add_ca("Receivables — Domestic", [y.balance_sheet.receivables_domestic for y in stmt.years], indent=True)
    add_ca("Receivables — Export", [y.balance_sheet.receivables_export for y in stmt.years], indent=True)
    add_ca("Inventory (RM + WIP + FG)", [m["inventory_total"] for m in metrics], indent=True)
    add_ca("Advances to Suppliers", [y.balance_sheet.advance_to_suppliers for y in stmt.years], indent=True)
    add_ca("Other Current Assets", [y.balance_sheet.other_current_assets for y in stmt.years], indent=True)
    add_ca("TOTAL CURRENT ASSETS (T1)", [m["current_assets"] for m in metrics], bold=True, indent=True)
    add_ca("CURRENT LIABILITIES", [""] * n_years, bold=True)
    add_ca("Short-term Borrowings from Banks", [y.balance_sheet.short_term_borrowing_banks for y in stmt.years], indent=True)
    add_ca("Sundry Creditors", [y.balance_sheet.sundry_creditors for y in stmt.years], indent=True)
    add_ca("Advance from Customers", [y.balance_sheet.advance_from_customers for y in stmt.years], indent=True)
    add_ca("Provision for Tax", [y.balance_sheet.provision_for_tax for y in stmt.years], indent=True)
    add_ca("Other Current Liabilities", [y.balance_sheet.other_current_liabilities for y in stmt.years], indent=True)
    add_ca("TOTAL CURRENT LIABILITIES (T2)", [m["current_liab"] for m in metrics], bold=True, indent=True)
    add_ca("OCL (Excl. Bank Borrowings)", [m["other_current_liab"] for m in metrics], bold=True, indent=True)
    add_ca("Net Working Capital (T1 − T2)", [m["nwc"] for m in metrics], bold=True, indent=True)
    story.append(_table(ca_rows))

    # ---------------- Form V — MPBF ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form V — MPBF (Tandon Committee Method)", h2))
    mpbf_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_mp(label, vals, bold=False):
        lbl = f"<b>{label}</b>" if bold else label
        mpbf_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_mp("Current Assets (CA)", [m["current_assets"] for m in metrics])
    add_mp("Other Current Liabilities (OCL)", [m["other_current_liab"] for m in metrics])
    add_mp("Working Capital Gap (CA − OCL)", [m["wcg"] for m in metrics], bold=True)
    add_mp("Net Working Capital (NWC)", [m["nwc"] for m in metrics])
    add_mp("MPBF — Method I  [0.75 × WCG]", [m["mpbf_method_1"] for m in metrics], bold=True)
    add_mp("MPBF — Method II [0.75 × CA − OCL]", [m["mpbf_method_2"] for m in metrics], bold=True)
    story.append(_table(mpbf_rows))

    # ---------------- Form VI — Funds Flow Statement ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form VI — Funds Flow Statement (₹ in Lakhs)", h2))
    if len(stmt.years) < 2:
        story.append(Paragraph("<i>Need at least 2 years to compute fund flow.</i>", small))
    else:
        ff_list = []
        for i in range(1, len(stmt.years)):
            ff_list.append(compute_fund_flow(metrics[i - 1], metrics[i], stmt.years[i - 1], stmt.years[i]))
        # Headers: Particulars + each transition
        ff_headers = [Paragraph("<b>Particulars</b>", cell_hdr)]
        for i in range(1, len(stmt.years)):
            ff_headers.append(Paragraph(
                f"<b>{stmt.years[i-1].year_label} → {stmt.years[i].year_label}</b>",
                cell_hdr,
            ))
        ff_rows = [ff_headers]
        # Collect labels
        all_src = []
        all_use = []
        for ff in ff_list:
            for k in ff["sources"].keys():
                if k not in all_src:
                    all_src.append(k)
            for k in ff["uses"].keys():
                if k not in all_use:
                    all_use.append(k)

        def add_ff(label, vals, bold=False, indent=False):
            lbl = ("&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
            ff_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

        add_ff("SOURCES OF FUNDS", [""] * len(ff_list), bold=True)
        for label in all_src:
            add_ff(label, [ff["sources"].get(label, 0) for ff in ff_list], indent=True)
        add_ff("TOTAL SOURCES", [ff["total_sources"] for ff in ff_list], bold=True, indent=True)
        add_ff("APPLICATION OF FUNDS", [""] * len(ff_list), bold=True)
        for label in all_use:
            add_ff(label, [ff["uses"].get(label, 0) for ff in ff_list], indent=True)
        add_ff("TOTAL USES", [ff["total_uses"] for ff in ff_list], bold=True, indent=True)
        add_ff("SURPLUS / (DEFICIT)", [ff["surplus_deficit"] for ff in ff_list], bold=True)
        ff_cols = [PARTICULARS_W] + [(USABLE_W - PARTICULARS_W) / max(1, len(ff_list))] * len(ff_list)
        story.append(_table(ff_rows, col_widths=ff_cols))

    # ---------------- Form VII — Summary of Financial Statements for Ratio Analysis ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form VII — Summary of Financial Statements for Ratio Analysis (₹ in Lakhs)", h2))
    summary_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_sm(label, vals, bold=False, indent=False):
        lbl = ("&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
        summary_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_sm("OPERATING STATEMENT — INCOME", [""] * n_years, bold=True)
    add_sm("Domestic Sales", [y.op_statement.domestic_sales for y in stmt.years], indent=True)
    add_sm("Export Sales", [y.op_statement.export_sales for y in stmt.years], indent=True)
    add_sm("Gross Sales", [m["gross_sales"] for m in metrics], bold=True, indent=True)
    add_sm("Net Sales", [m["net_sales"] for m in metrics], bold=True, indent=True)
    add_sm("Other Income", [y.op_statement.other_income for y in stmt.years], indent=True)
    add_sm("Gross Income", [m["total_revenue"] for m in metrics], bold=True, indent=True)
    add_sm("OPERATING STATEMENT — EXPENSES", [""] * n_years, bold=True)
    add_sm("Raw Material Consumed", [m["rm_consumed"] for m in metrics], indent=True)
    add_sm("Direct Labour", [y.op_statement.direct_labour for y in stmt.years], indent=True)
    add_sm("Depreciation", [y.op_statement.depreciation for y in stmt.years], indent=True)
    add_sm("Total Cost of Sales", [m["cost_of_sales"] for m in metrics], bold=True, indent=True)
    add_sm("Gross Profit", [m["gross_profit"] for m in metrics], bold=True, indent=True)
    add_sm("Operating Profit (PBIT)", [m["operating_profit"] for m in metrics], bold=True, indent=True)
    add_sm("Interest", [m["interest_total"] for m in metrics], indent=True)
    add_sm("Profit Before Tax (PBT)", [m["pbt"] for m in metrics], bold=True, indent=True)
    add_sm("Tax Provision", [m["tax"] for m in metrics], indent=True)
    add_sm("Profit After Tax (PAT)", [m["pat"] for m in metrics], bold=True, indent=True)
    add_sm("Cash Profit (PBDIT)", [m["pbdit"] for m in metrics], indent=True)
    add_sm("BALANCE SHEET — LIABILITIES", [""] * n_years, bold=True)
    add_sm("Short-term Bank Borrowings", [y.balance_sheet.short_term_borrowing_banks for y in stmt.years], indent=True)
    add_sm("Sundry Creditors", [y.balance_sheet.sundry_creditors for y in stmt.years], indent=True)
    add_sm("Total Current Liabilities", [m["current_liab"] for m in metrics], bold=True, indent=True)
    add_sm("Term Liabilities", [m["term_liab"] for m in metrics], bold=True, indent=True)
    add_sm("Net Worth", [m["net_worth"] for m in metrics], bold=True, indent=True)
    add_sm("Tangible Net Worth", [m["tangible_net_worth"] for m in metrics], bold=True, indent=True)
    add_sm("BALANCE SHEET — ASSETS", [""] * n_years, bold=True)
    add_sm("Cash & Bank + Investments", [y.balance_sheet.cash_bank_balances + y.balance_sheet.investments_short_term for y in stmt.years], indent=True)
    add_sm("Receivables (Domestic + Export)", [m["receivables_total"] for m in metrics], indent=True)
    add_sm("Inventory (RM + WIP + FG)", [m["inventory_total"] for m in metrics], indent=True)
    add_sm("Total Current Assets", [m["current_assets"] for m in metrics], bold=True, indent=True)
    add_sm("Net Fixed Assets (Net Block + CWIP)", [m["fixed_assets_total"] for m in metrics], bold=True, indent=True)
    add_sm("Intangible Assets", [y.balance_sheet.intangible_assets for y in stmt.years], indent=True)
    add_sm("TOTAL ASSETS", [m["total_assets"] for m in metrics], bold=True, indent=True)
    add_sm("Capital Employed (NW + TL)", [m["capital_employed"] for m in metrics], bold=True, indent=True)
    add_sm("Net Working Capital", [m["nwc"] for m in metrics], bold=True, indent=True)
    story.append(_table(summary_rows))

    # ---------------- Form VIII — Ratio Analysis (categorized) ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form VIII — Ratio Analysis", h2))
    ratio_rows = [[Paragraph("<b>Ratio</b>", cell_hdr)] + year_paragraphs]

    def add_section(title):
        ratio_rows.append([Paragraph(f"<b>{title}</b>", cell)] + [Paragraph("", cell)] * n_years)

    def add_ratio(label, key, fmt):
        ratio_rows.append([Paragraph("&nbsp;&nbsp;" + label, cell)] + [fmt(m.get(key, 0)) for m in metrics])

    add_section("(A) Long-term Solvency Ratios")
    add_ratio("Debt / Equity Ratio", "debt_equity", _fmt_x)
    add_ratio("Net Worth to Total Assets", "net_worth_to_total_assets", _fmt_pct)
    add_ratio("Debt to Net Worth", "debt_to_net_worth", _fmt_x)
    add_ratio("Capital Gearing Ratio", "capital_gearing", _fmt_x)
    add_ratio("Fixed Assets / Long-Term Funds", "fixed_assets_to_long_term", _fmt_x)
    add_ratio("Proprietary Ratio", "proprietary_ratio", _fmt_pct)
    add_ratio("Interest Coverage", "interest_coverage", _fmt_x)
    add_ratio("Debt Service Coverage (DSCR)", "dscr", _fmt_x)
    add_ratio("TOL / TNW", "tol_tnw", _fmt_x)
    add_section("(B) Short-term Solvency Ratios")
    add_ratio("Current Ratio", "current_ratio", _fmt_x)
    add_ratio("Quick / Acid Test Ratio", "quick_ratio", _fmt_x)
    add_ratio("Absolute Liquid Ratio", "absolute_liquid_ratio", _fmt_x)
    add_section("(C) Profitability Ratios")
    add_ratio("Return on Capital Employed (ROCE)", "roce", _fmt_pct)
    add_ratio("Gross Profit Margin", "gp_margin", _fmt_pct)
    add_ratio("Net Profit Margin", "np_margin", _fmt_pct)
    add_ratio("Cash Profit Ratio", "cash_profit_ratio", _fmt_pct)
    add_ratio("Return on Net Worth", "return_on_net_worth", _fmt_pct)
    add_ratio("Op. Profit Margin (before int.)", "op_margin", _fmt_pct)
    add_ratio("Op. Profit Margin (after int.)", "op_margin_after_int", _fmt_pct)
    add_section("(D) Activity Ratios")
    add_ratio("Inventory Turnover", "stock_turnover", _fmt_x)
    add_ratio("Debtors Turnover", "debtors_turnover", _fmt_x)
    add_ratio("Creditors Turnover", "creditors_turnover", _fmt_x)
    add_ratio("Debtors Turnover Period", "debtors_velocity_days", _fmt_days)
    add_ratio("Creditors Turnover Period", "creditors_velocity_days", _fmt_days)
    add_ratio("Fixed Assets Turnover", "fixed_assets_turnover", _fmt_x)
    add_ratio("Total Assets Turnover", "assets_turnover", _fmt_x)
    add_ratio("Working Capital Turnover", "working_capital_turnover", _fmt_x)
    add_ratio("Sales / Capital Employed", "sales_to_capital_employed", _fmt_x)
    add_section("(E) Operating Ratios")
    add_ratio("Domestic Sales Proportion", "domestic_sales_pct", _fmt_pct)
    add_ratio("Export Sales Proportion", "export_sales_pct", _fmt_pct)
    add_ratio("Material Cost Ratio", "material_cost_ratio", _fmt_pct)
    add_ratio("Direct Labour Cost Ratio", "direct_labour_ratio", _fmt_pct)
    add_ratio("Other Overheads Ratio", "other_overheads_ratio", _fmt_pct)
    add_ratio("Indirect Cost Ratio", "indirect_cost_ratio", _fmt_pct)
    add_ratio("Interest Cost Ratio", "interest_cost_ratio", _fmt_pct)
    add_ratio("Operating Cost Ratio", "operating_cost_ratio", _fmt_pct)
    story.append(_table(ratio_rows))

    # ---------------- Form IX — Statement of Changes in Working Capital ----------------
    story.append(PageBreak())
    story.append(Paragraph("Form IX — Statement of Changes in Working Capital (₹ in Lakhs)", h2))
    wc_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] + year_paragraphs]

    def add_wc(label, vals, bold=False, indent=False):
        lbl = ("&nbsp;&nbsp;" if indent else "") + (f"<b>{label}</b>" if bold else label)
        wc_rows.append([Paragraph(lbl, cell)] + [_fmt(v) for v in vals])

    add_wc("CURRENT ASSETS", [""] * n_years, bold=True)
    add_wc("Cash and Bank Balances", [y.balance_sheet.cash_bank_balances for y in stmt.years], indent=True)
    add_wc("Investments — Short-term", [y.balance_sheet.investments_short_term for y in stmt.years], indent=True)
    add_wc("Receivables — Domestic", [y.balance_sheet.receivables_domestic for y in stmt.years], indent=True)
    add_wc("Receivables — Export", [y.balance_sheet.receivables_export for y in stmt.years], indent=True)
    add_wc("Inventory — Raw Material", [y.balance_sheet.inventory_rm for y in stmt.years], indent=True)
    add_wc("Inventory — WIP", [y.balance_sheet.inventory_wip for y in stmt.years], indent=True)
    add_wc("Inventory — Finished Goods", [y.balance_sheet.inventory_fg for y in stmt.years], indent=True)
    add_wc("Advances to Suppliers", [y.balance_sheet.advance_to_suppliers for y in stmt.years], indent=True)
    add_wc("Other Current Assets", [y.balance_sheet.other_current_assets for y in stmt.years], indent=True)
    add_wc("Total Current Assets (TCA)", [m["current_assets"] for m in metrics], bold=True, indent=True)
    tca_change = [0.0] + [metrics[i]["current_assets"] - metrics[i - 1]["current_assets"] for i in range(1, len(metrics))]
    add_wc("Change in Current Assets (YoY)", tca_change, bold=True, indent=True)
    add_wc("CURRENT LIABILITIES", [""] * n_years, bold=True)
    add_wc("Short-term Bank Borrowings", [y.balance_sheet.short_term_borrowing_banks for y in stmt.years], indent=True)
    add_wc("Sundry Creditors", [y.balance_sheet.sundry_creditors for y in stmt.years], indent=True)
    add_wc("Advance from Customers", [y.balance_sheet.advance_from_customers for y in stmt.years], indent=True)
    add_wc("Provision for Tax", [y.balance_sheet.provision_for_tax for y in stmt.years], indent=True)
    add_wc("Other Current Liabilities", [y.balance_sheet.other_current_liabilities for y in stmt.years], indent=True)
    add_wc("Total Current Liabilities (TCL)", [m["current_liab"] for m in metrics], bold=True, indent=True)
    tcl_change = [0.0] + [metrics[i]["current_liab"] - metrics[i - 1]["current_liab"] for i in range(1, len(metrics))]
    add_wc("Change in Current Liabilities (YoY)", tcl_change, bold=True, indent=True)
    add_wc("NET WORKING CAPITAL", [""] * n_years, bold=True)
    add_wc("Net Working Capital (TCA − TCL)", [m["nwc"] for m in metrics], bold=True, indent=True)
    nwc_change = [0.0] + [metrics[i]["nwc"] - metrics[i - 1]["nwc"] for i in range(1, len(metrics))]
    add_wc("Increase / (Decrease) in NWC", nwc_change, bold=True, indent=True)
    add_wc("Working Capital Gap (CA − OCL)", [m["wcg"] for m in metrics], bold=True, indent=True)
    story.append(_table(wc_rows))

    # ---------------- Sensitivity Analysis (preserved — optional) ----------------
    if stmt.sensitivity.enabled and stmt.years:
        sens = compute_sensitivity(stmt)
        story.append(PageBreak())
        story.append(Paragraph(
            f"Sensitivity Analysis (Target Year: {sens.get('target_year_label', '—')})",
            h2,
        ))
        sc_names = [sc["name"] for sc in sens["scenarios"]]
        sens_rows = [[Paragraph("<b>Particulars</b>", cell_hdr)] +
                     [Paragraph(f"<b>{n}</b>", cell_hdr) for n in sc_names]]
        for delta_label, dkey in [
            ("Sales Δ (%)", "sales_delta_pct"),
            ("RM Cost Δ (%)", "rm_cost_delta_pct"),
            ("Interest Δ (%)", "interest_delta_pct"),
            ("Other Exp. Δ (%)", "other_expenses_delta_pct"),
        ]:
            sens_rows.append([Paragraph(delta_label, cell)] +
                             [_fmt(sc["deltas"].get(dkey, 0)) for sc in sens["scenarios"]])
        for metric_label, mkey in [
            ("Net Sales", "net_sales"),
            ("Operating Profit", "operating_profit"),
            ("PBT", "pbt"),
            ("PAT", "pat"),
            ("Cash Accrual", "cash_accrual"),
            ("NP Margin (%)", "np_margin"),
            ("Interest Coverage", "interest_coverage"),
            ("DSCR", "dscr"),
        ]:
            sens_rows.append([Paragraph(metric_label, cell)] +
                             [_fmt(sc["metrics"].get(mkey, 0)) for sc in sens["scenarios"]])
        sens_cols = [PARTICULARS_W] + [(USABLE_W - PARTICULARS_W) / max(1, len(sc_names))] * len(sc_names)
        story.append(_table(sens_rows, col_widths=sens_cols))

    # ---------------- Footer ----------------
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"<i>Generated by DPRForge — CMA Data Preparation Module on "
        f"{datetime.now(timezone.utc).strftime('%d %b %Y')}. "
        f"Figures shown in lakhs unless stated otherwise. "
        f"For full sheets and ageing schedules refer to the Excel download.</i>",
        small,
    ))
    story.append(Spacer(1, 1 * mm))
    story.append(Paragraph(
        f"<font color='#94A3B8'>{seller.get('name')} • GSTIN {seller.get('gstin', '')}</font>",
        small,
    ))

    # Add page numbers
    def _on_page(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#94A3B8"))
        canvas.drawRightString(
            landscape(A4)[0] - 12 * mm, 8 * mm,
            f"Page {_doc.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return buf.getvalue()


# =============================== ROUTES ===============================

# =============================== AUTO-PROJECTION HELPERS ===============================

# Op statement fields that should grow with sales (revenue + variable costs)
_OP_GROWTH_FIELDS = [
    "domestic_sales", "export_sales", "less_excise", "other_income",
    "opening_stock_rm", "raw_materials_purchase", "closing_stock_rm",
    "power_fuel", "direct_labour", "other_mfg_expenses",
    "opening_stock_wip", "closing_stock_wip",
    "opening_stock_fg", "closing_stock_fg",
    "selling_expenses", "admin_expenses",
    "interest_on_bank_borrowing", "interest_on_others",
    "depreciation",
]
# Balance-sheet fields that should also scale with business size
_BS_GROWTH_FIELDS = [
    "short_term_borrowing_banks", "sundry_creditors", "advance_from_customers",
    "provision_for_tax", "other_current_liabilities",
    "term_loan_banks", "term_loan_others", "debentures",
    "deferred_payment_credits", "other_term_liabilities",
    "paid_up_capital", "reserves_surplus", "other_reserves",
    "cash_bank_balances", "investments_short_term",
    "receivables_domestic", "receivables_export",
    "inventory_rm", "inventory_wip", "inventory_fg",
    "advance_to_suppliers", "other_current_assets",
    "gross_block", "accumulated_depreciation", "capital_wip",
    "investments_long_term", "intangible_assets",
    "deferred_revenue_expenditure", "other_non_current_assets",
]


def _net_sales_of(y: CMAYearData) -> float:
    op = y.op_statement
    return (op.domestic_sales + op.export_sales) - op.less_excise


def _detect_growth(stmt: CMAStatement) -> Dict[str, Any]:
    """Compute YoY net-sales growth % from audited years. Default 15% if not derivable."""
    audited = [y for y in stmt.years if y.year_type == "audited"]
    growths: List[float] = []
    if len(audited) >= 2:
        for i in range(1, len(audited)):
            prev_ns = _net_sales_of(audited[i - 1])
            curr_ns = _net_sales_of(audited[i])
            if prev_ns > 0 and curr_ns > 0:
                growths.append((curr_ns / prev_ns - 1.0) * 100.0)
    detected = round(sum(growths) / len(growths), 2) if growths else None
    return {
        "detected_growth_pct": detected,
        "default_growth_pct": detected if detected is not None else 15.0,
        "audited_years_count": len(audited),
        "audited_years_with_sales": sum(1 for y in audited if _net_sales_of(y) > 0),
        "can_auto_project": bool(growths),
    }


def _apply_auto_projections(stmt: CMAStatement, payload: Dict[str, Any]) -> (CMAStatement, Dict[str, Any]):
    """Fill provisional + projected years from the LAST audited year using a YoY growth multiplier."""
    fill_provisional = bool(payload.get("fill_provisional", True))
    fill_projected = bool(payload.get("fill_projected", True))
    fields = payload.get("fields") or ["op_statement", "balance_sheet"]

    hint = _detect_growth(stmt)
    if payload.get("growth_rate_pct") is not None:
        try:
            growth_pct = float(payload["growth_rate_pct"])
        except (TypeError, ValueError):
            growth_pct = hint["default_growth_pct"]
    else:
        growth_pct = hint["default_growth_pct"]
    g = 1.0 + (growth_pct / 100.0)

    # Find the "base" year = last audited (or last year with any sales > 0)
    base_idx = -1
    for i, y in enumerate(stmt.years):
        if y.year_type == "audited":
            base_idx = i
    if base_idx == -1:
        # Fallback: use year 0 (won't have data, but at least no crash)
        base_idx = 0
    base_year = stmt.years[base_idx]

    years = [y.model_copy(deep=True) for y in stmt.years]

    for j in range(base_idx + 1, len(years)):
        y = years[j]
        if y.year_type == "provisional" and not fill_provisional:
            continue
        if y.year_type == "projected" and not fill_projected:
            continue
        # Years after base — multiplier compounds
        n = j - base_idx
        mult = g ** n
        if "op_statement" in fields:
            for k in _OP_GROWTH_FIELDS:
                base_v = getattr(base_year.op_statement, k, 0) or 0
                setattr(y.op_statement, k, round(float(base_v) * mult, 2))
            # Tax rate doesn't grow — copy as-is
            y.op_statement.tax_rate = base_year.op_statement.tax_rate
        if "balance_sheet" in fields:
            for k in _BS_GROWTH_FIELDS:
                base_v = getattr(base_year.balance_sheet, k, 0) or 0
                setattr(y.balance_sheet, k, round(float(base_v) * mult, 2))
            # Revaluation reserve — copy as-is (not growth driven)
            y.balance_sheet.revaluation_reserve = base_year.balance_sheet.revaluation_reserve

    new_stmt = stmt.model_copy(update={"years": years})
    return new_stmt, {"applied_growth_pct": growth_pct, "base_year_index": base_idx,
                      "base_year_label": base_year.year_label, "hint": hint}


# =============================== INDUSTRY BENCHMARKS (curated) ===============================

# Indian SME averages — sourced from public CMA banking norms (approximate, used as a reference).
INDUSTRY_BENCHMARKS: Dict[str, Dict[str, float]] = {
    "manufacturing":          {"current_ratio": 1.33, "quick_ratio": 0.85, "debt_equity": 2.0, "tol_tnw": 4.0, "np_margin": 5.5,  "op_margin": 9.0,  "gp_margin": 22.0, "roce": 14.0, "stock_turnover": 5.0,  "debtors_velocity_days": 60,  "creditors_velocity_days": 45,  "interest_coverage": 2.5, "dscr": 1.5},
    "trading":                {"current_ratio": 1.25, "quick_ratio": 0.70, "debt_equity": 3.0, "tol_tnw": 5.0, "np_margin": 2.5,  "op_margin": 4.5,  "gp_margin": 12.0, "roce": 12.0, "stock_turnover": 8.0,  "debtors_velocity_days": 45,  "creditors_velocity_days": 30,  "interest_coverage": 2.0, "dscr": 1.35},
    "service":                {"current_ratio": 1.50, "quick_ratio": 1.30, "debt_equity": 1.5, "tol_tnw": 3.0, "np_margin": 12.0, "op_margin": 18.0, "gp_margin": 45.0, "roce": 20.0, "stock_turnover": 0.0,  "debtors_velocity_days": 30,  "creditors_velocity_days": 30,  "interest_coverage": 3.0, "dscr": 1.6},
    "textile":                {"current_ratio": 1.30, "quick_ratio": 0.80, "debt_equity": 2.5, "tol_tnw": 4.5, "np_margin": 4.0,  "op_margin": 8.0,  "gp_margin": 18.0, "roce": 11.0, "stock_turnover": 4.5,  "debtors_velocity_days": 75,  "creditors_velocity_days": 50,  "interest_coverage": 2.0, "dscr": 1.3},
    "food_processing":        {"current_ratio": 1.30, "quick_ratio": 0.75, "debt_equity": 2.0, "tol_tnw": 4.0, "np_margin": 5.0,  "op_margin": 9.5,  "gp_margin": 20.0, "roce": 13.0, "stock_turnover": 6.0,  "debtors_velocity_days": 45,  "creditors_velocity_days": 40,  "interest_coverage": 2.4, "dscr": 1.45},
    "auto_parts":             {"current_ratio": 1.35, "quick_ratio": 0.90, "debt_equity": 2.0, "tol_tnw": 4.0, "np_margin": 6.0,  "op_margin": 10.0, "gp_margin": 22.0, "roce": 15.0, "stock_turnover": 5.5,  "debtors_velocity_days": 75,  "creditors_velocity_days": 60,  "interest_coverage": 2.5, "dscr": 1.5},
    "construction":           {"current_ratio": 1.20, "quick_ratio": 0.60, "debt_equity": 2.5, "tol_tnw": 5.0, "np_margin": 4.5,  "op_margin": 8.0,  "gp_margin": 14.0, "roce": 10.0, "stock_turnover": 3.0,  "debtors_velocity_days": 90,  "creditors_velocity_days": 60,  "interest_coverage": 1.8, "dscr": 1.3},
    "pharma":                 {"current_ratio": 1.50, "quick_ratio": 1.00, "debt_equity": 1.2, "tol_tnw": 2.5, "np_margin": 12.0, "op_margin": 20.0, "gp_margin": 50.0, "roce": 18.0, "stock_turnover": 4.0,  "debtors_velocity_days": 75,  "creditors_velocity_days": 60,  "interest_coverage": 4.0, "dscr": 1.8},
    "chemicals":              {"current_ratio": 1.35, "quick_ratio": 0.85, "debt_equity": 2.0, "tol_tnw": 4.0, "np_margin": 7.0,  "op_margin": 12.0, "gp_margin": 25.0, "roce": 15.0, "stock_turnover": 4.5,  "debtors_velocity_days": 60,  "creditors_velocity_days": 45,  "interest_coverage": 2.8, "dscr": 1.55},
    "it_software":            {"current_ratio": 2.00, "quick_ratio": 1.90, "debt_equity": 0.5, "tol_tnw": 1.5, "np_margin": 18.0, "op_margin": 25.0, "gp_margin": 55.0, "roce": 25.0, "stock_turnover": 0.0,  "debtors_velocity_days": 60,  "creditors_velocity_days": 30,  "interest_coverage": 6.0, "dscr": 2.5},
    "retail":                 {"current_ratio": 1.20, "quick_ratio": 0.50, "debt_equity": 2.5, "tol_tnw": 4.0, "np_margin": 3.0,  "op_margin": 5.5,  "gp_margin": 18.0, "roce": 12.0, "stock_turnover": 7.0,  "debtors_velocity_days": 15,  "creditors_velocity_days": 30,  "interest_coverage": 2.2, "dscr": 1.4},
    "engineering":            {"current_ratio": 1.30, "quick_ratio": 0.85, "debt_equity": 2.0, "tol_tnw": 4.0, "np_margin": 6.5,  "op_margin": 11.0, "gp_margin": 24.0, "roce": 14.0, "stock_turnover": 4.5,  "debtors_velocity_days": 75,  "creditors_velocity_days": 50,  "interest_coverage": 2.5, "dscr": 1.5},
    "plastics":               {"current_ratio": 1.30, "quick_ratio": 0.80, "debt_equity": 2.2, "tol_tnw": 4.0, "np_margin": 5.0,  "op_margin": 9.0,  "gp_margin": 20.0, "roce": 13.0, "stock_turnover": 5.0,  "debtors_velocity_days": 60,  "creditors_velocity_days": 45,  "interest_coverage": 2.3, "dscr": 1.4},
    "agriculture":            {"current_ratio": 1.30, "quick_ratio": 0.65, "debt_equity": 2.0, "tol_tnw": 3.5, "np_margin": 6.0,  "op_margin": 10.0, "gp_margin": 22.0, "roce": 12.0, "stock_turnover": 3.5,  "debtors_velocity_days": 30,  "creditors_velocity_days": 30,  "interest_coverage": 2.2, "dscr": 1.4},
    "hospitality":            {"current_ratio": 1.20, "quick_ratio": 1.00, "debt_equity": 2.5, "tol_tnw": 4.0, "np_margin": 8.0,  "op_margin": 15.0, "gp_margin": 40.0, "roce": 12.0, "stock_turnover": 0.0,  "debtors_velocity_days": 30,  "creditors_velocity_days": 45,  "interest_coverage": 2.0, "dscr": 1.4},
}


def _lookup_benchmark(industry: str) -> Dict[str, Any]:
    """Find best-matching benchmark by lowercased substring."""
    key = (industry or "").lower().strip().replace(" ", "_").replace("&", "and").replace("/", "_")
    # exact
    if key in INDUSTRY_BENCHMARKS:
        return {"industry": key, "source": "curated", "ratios": INDUSTRY_BENCHMARKS[key]}
    # substring
    for k in INDUSTRY_BENCHMARKS:
        if k in key or key in k:
            return {"industry": k, "matched_from": industry, "source": "curated", "ratios": INDUSTRY_BENCHMARKS[k]}
    # default = manufacturing
    return {"industry": "manufacturing", "matched_from": industry, "source": "curated_default", "ratios": INDUSTRY_BENCHMARKS["manufacturing"]}


# =============================== AI HELPERS (gpt-5.2) ===============================

import os
import json as _json
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore
    _LLM_AVAILABLE = True
except ImportError:
    _LLM_AVAILABLE = False
    LlmChat = None  # type: ignore
    UserMessage = None  # type: ignore


def _extract_json(text: str) -> Dict[str, Any]:
    """Best-effort JSON extraction from a chat response."""
    if not text:
        return {}
    # Try fenced ```json blocks
    import re
    m = re.search(r"```json\s*(\{.*?\})\s*```", text, re.S)
    if m:
        try:
            return _json.loads(m.group(1))
        except Exception:
            pass
    # Try first {...} block
    m = re.search(r"\{.*\}", text, re.S)
    if m:
        try:
            return _json.loads(m.group(0))
        except Exception:
            pass
    return {}


async def _llm_growth_suggestion(stmt: CMAStatement, audited: List[CMAYearData]) -> Dict[str, Any]:
    if not _LLM_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI library not available on server")
    api_key = os.environ.get("EMERGENT_LLM_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")

    # Summarise audited line items as a compact prompt
    summary_lines = []
    for i, y in enumerate(audited):
        op = y.op_statement
        bs = y.balance_sheet
        summary_lines.append(
            f"Year {i+1} [{y.year_label}]: "
            f"domestic_sales={op.domestic_sales}, export_sales={op.export_sales}, "
            f"raw_materials_purchase={op.raw_materials_purchase}, direct_labour={op.direct_labour}, "
            f"power_fuel={op.power_fuel}, other_mfg_expenses={op.other_mfg_expenses}, "
            f"selling_expenses={op.selling_expenses}, admin_expenses={op.admin_expenses}, "
            f"depreciation={op.depreciation}, interest_on_bank_borrowing={op.interest_on_bank_borrowing}, "
            f"sundry_creditors={bs.sundry_creditors}, receivables_domestic={bs.receivables_domestic}, "
            f"inventory_rm={bs.inventory_rm}, gross_block={bs.gross_block}, term_loan_banks={bs.term_loan_banks}"
        )
    n_proj = sum(1 for y in stmt.years if y.year_type in ("provisional", "projected"))
    prompt = (
        f"You are a financial analyst helping prepare a bank CMA for an Indian SME in industry: '{stmt.industry or 'Manufacturing'}'.\n"
        f"Audited line items (₹ Lakhs):\n" + "\n".join(summary_lines) + "\n\n"
        f"Suggest realistic YoY growth percentages (decimal, e.g. 12.5 = 12.5%) for each line item for the next {n_proj} years. "
        f"Differentiate revenue, variable cost, fixed cost, depreciation, interest, working-capital balance-sheet items. "
        f"Return ONLY a JSON object like: "
        f'{{"field_growth_pct": {{"domestic_sales": 15, "export_sales": 12, "raw_materials_purchase": 14, "direct_labour": 10, "power_fuel": 8, "other_mfg_expenses": 8, "selling_expenses": 10, "admin_expenses": 6, "depreciation": 5, "interest_on_bank_borrowing": 8, "interest_on_others": 0, "sundry_creditors": 12, "receivables_domestic": 15, "inventory_rm": 14, "gross_block": 8, "term_loan_banks": -10, "paid_up_capital": 0, "reserves_surplus": 18}}, "reasoning": "Brief one-line justification"}}. '
        f"No commentary, no markdown. Just the JSON."
    )

    chat = LlmChat(api_key=api_key, session_id=f"cma_ai_growth_{stmt.cma_id}",
                   system_message="You are a precise Indian banking CMA analyst. Return only valid JSON.")
    chat.with_model("openai", "gpt-5.2")
    msg = UserMessage(text=prompt)
    response = await chat.send_message(msg)
    parsed = _extract_json(str(response))
    if "field_growth_pct" not in parsed:
        # Fallback: return uniform default
        return {
            "field_growth_pct": {"domestic_sales": 15, "raw_materials_purchase": 13, "direct_labour": 10,
                                 "power_fuel": 8, "other_mfg_expenses": 8, "selling_expenses": 10,
                                 "admin_expenses": 6, "depreciation": 5, "interest_on_bank_borrowing": 8,
                                 "sundry_creditors": 12, "receivables_domestic": 15, "inventory_rm": 14,
                                 "gross_block": 10, "term_loan_banks": -5, "reserves_surplus": 18},
            "reasoning": "Fallback defaults (AI response could not be parsed)",
            "source": "fallback",
        }
    parsed["source"] = "gpt-5.2"
    return parsed


async def _llm_benchmark_refine(industry: str) -> Dict[str, Any]:
    if not _LLM_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI library not available on server")
    api_key = os.environ.get("EMERGENT_LLM_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")
    # Pre-seed with curated as a baseline for the LLM
    base = _lookup_benchmark(industry)
    prompt = (
        f"You are an Indian banking CMA analyst. Provide realistic average financial ratios for an SME in industry: '{industry}'.\n"
        f"Curated baseline (for reference, may be off): {base['ratios']}.\n"
        f"Return ONLY a JSON object with this exact shape: "
        f'{{"ratios": {{"current_ratio":1.3, "quick_ratio":0.8, "debt_equity":2.0, "tol_tnw":4.0, "np_margin":5.0, "op_margin":9.0, "gp_margin":22.0, "roce":14.0, "stock_turnover":5.0, "debtors_velocity_days":60, "creditors_velocity_days":45, "interest_coverage":2.5, "dscr":1.5}}, "reasoning":"one-line note"}}. '
        f"All numeric (not strings). Indian SME context, fiscal year basis. No markdown."
    )
    chat = LlmChat(api_key=api_key, session_id=f"cma_ai_bench_{industry[:30]}",
                   system_message="You are a precise Indian SME credit analyst. Return only valid JSON.")
    chat.with_model("openai", "gpt-5.2")
    msg = UserMessage(text=prompt)
    response = await chat.send_message(msg)
    parsed = _extract_json(str(response))
    ratios = parsed.get("ratios") or {}
    if not ratios:
        return {"industry": industry, "source": "curated_fallback", "ratios": base["ratios"], "reasoning": "AI response not parseable"}
    return {"industry": industry, "source": "gpt-5.2", "ratios": ratios, "reasoning": parsed.get("reasoning", "")}


def _stmt_compact_context(stmt: CMAStatement) -> str:
    """Compact textual representation of the CMA used as LLM context."""
    lines = [f"Company: {stmt.company_name} | Industry: {stmt.industry or 'N/A'} | Banker: {stmt.banker_name or 'N/A'} | Facility: {stmt.facility_type}"]
    for y in stmt.years:
        m = compute_year_metrics(y)
        lines.append(
            f"  {y.year_label} ({y.year_type}): NetSales={m['net_sales']:.0f}, GP={m['gross_profit']:.0f}, "
            f"PAT={m['pat']:.0f}, CurrentRatio={m['current_ratio']:.2f}, DE={m['debt_equity']:.2f}, "
            f"DSCR={m['dscr']:.2f}, NPMargin={m['np_margin']:.1f}%, ROCE={m['roce']:.1f}%, "
            f"MPBF_I={m['mpbf_method_1']:.0f}, MPBF_II={m['mpbf_method_2']:.0f}"
        )
    if stmt.existing_limits:
        lines.append("Existing/Proposed Limits:")
        for l in stmt.existing_limits:
            lines.append(f"  - {l.facility} @ {l.bank}: Existing ₹{l.existing_limit}, Proposed ₹{l.proposed_limit}, ROI {l.rate_of_interest}%")
    return "\n".join(lines)


async def _llm_cma_summary(stmt: CMAStatement) -> str:
    if not _LLM_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI library not available")
    api_key = os.environ.get("EMERGENT_LLM_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")
    ctx = _stmt_compact_context(stmt)
    prompt = (
        "You are an Indian SME credit analyst writing a one-page bank-facing CMA narrative. "
        "Be specific, concise (no fluff), and use rupee/percentage figures. Structure with 4 sections: "
        "1) Business snapshot (2 lines), 2) Financial performance trend (3-4 lines), "
        "3) Key ratios & risk observations (3-4 lines, mention if any ratio is weak vs typical SME averages), "
        "4) Working-capital / MPBF recommendation (2-3 lines). "
        "Total length 200-280 words. Plain text, no markdown formatting.\n\n"
        f"CMA DATA:\n{ctx}"
    )
    chat = LlmChat(api_key=api_key, session_id=f"cma_summary_{stmt.cma_id}",
                   system_message="You are an experienced Indian SME credit analyst. Write professional, factual CMA narratives. No emojis.")
    chat.with_model("openai", "gpt-5.2")
    msg = UserMessage(text=prompt)
    response = await chat.send_message(msg)
    return str(response).strip()


async def _llm_cma_chat(stmt: CMAStatement, question: str) -> str:
    if not _LLM_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI library not available")
    api_key = os.environ.get("EMERGENT_LLM_KEY", "").strip()
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")
    ctx = _stmt_compact_context(stmt)
    prompt = (
        f"You are helping a CA / SME owner understand a CMA. Here is the CMA snapshot:\n\n{ctx}\n\n"
        f"User's question: {question}\n\n"
        f"Answer concisely (3-6 sentences) using the data above. If the question is outside CMA scope, say so politely. "
        f"Use ₹ for currency. Plain text, no markdown."
    )
    chat = LlmChat(api_key=api_key, session_id=f"cma_chat_{stmt.cma_id}",
                   system_message="You are a precise Indian banking CMA expert. Answer in plain text using user's data. No emojis, no markdown.")
    chat.with_model("openai", "gpt-5.2")
    msg = UserMessage(text=prompt)
    response = await chat.send_message(msg)
    return str(response).strip()


def _apply_per_field_growth(stmt: CMAStatement, rates: Dict[str, float]) -> CMAStatement:
    """Apply per-field growth % from last audited year, compounding (1+r)^n across subsequent years."""
    # Find last audited year
    base_idx = -1
    for i, y in enumerate(stmt.years):
        if y.year_type == "audited":
            base_idx = i
    if base_idx == -1:
        base_idx = 0
    base = stmt.years[base_idx]
    years = [y.model_copy(deep=True) for y in stmt.years]
    for j in range(base_idx + 1, len(years)):
        n = j - base_idx
        y = years[j]
        for k in _OP_GROWTH_FIELDS:
            r = float(rates.get(k, 0)) / 100.0
            base_v = getattr(base.op_statement, k, 0) or 0
            setattr(y.op_statement, k, round(float(base_v) * ((1 + r) ** n), 2))
        y.op_statement.tax_rate = base.op_statement.tax_rate
        for k in _BS_GROWTH_FIELDS:
            r = float(rates.get(k, 0)) / 100.0
            base_v = getattr(base.balance_sheet, k, 0) or 0
            setattr(y.balance_sheet, k, round(float(base_v) * ((1 + r) ** n), 2))
        y.balance_sheet.revaluation_reserve = base.balance_sheet.revaluation_reserve
    return stmt.model_copy(update={"years": years})


# =============================== CARRY-FORWARD RECONCILIATION ===============================

def _reconcile_carry_forward(stmt: CMAStatement) -> CMAStatement:
    """Apply automatic CMA carry-forward / consistency rules between years.

    Rules (Year N → Year N+1):
      1. Op-statement closing stocks (RM/WIP/FG) → Op-statement opening stocks of FY N+1
      2. Op-statement closing stocks of FY N → Balance-sheet inventory_rm / wip / fg of SAME FY N
      3. Gross block of FY N+1 ≥ Gross block of FY N (no shrinking — user may override / add capex)
      4. Accumulated depreciation of FY N+1 = AccDep of FY N + depreciation expense of FY N
      5. Reserves & Surplus of FY N+1 = R&S of FY N + PAT of FY N (assumes zero dividend / drawings)
      6. Capital WIP and intangibles of FY N+1 ≥ Year N's value (carry-forward; can grow with capex)

    First year (Year 1) opening values are ALWAYS user-entered — never overwritten.
    Audited years keep their actual balance-sheet numbers (rules 3-5 skip them).
    """
    if not stmt.years or len(stmt.years) < 1:
        return stmt
    years = [y.model_copy(deep=True) for y in stmt.years]

    # Rule 2 — closing stock → balance sheet inventory (for EVERY year, including audited)
    for y in years:
        cs_rm = float(y.op_statement.closing_stock_rm or 0)
        cs_wip = float(y.op_statement.closing_stock_wip or 0)
        cs_fg = float(y.op_statement.closing_stock_fg or 0)
        if cs_rm > 0:
            y.balance_sheet.inventory_rm = cs_rm
        if cs_wip > 0:
            y.balance_sheet.inventory_wip = cs_wip
        if cs_fg > 0:
            y.balance_sheet.inventory_fg = cs_fg

    # Rules 1, 3, 4, 5, 6 — between consecutive years (Year 1 is NEVER touched here)
    for i in range(1, len(years)):
        prev = years[i - 1]
        curr = years[i]

        # Rule 1: opening stocks of current year = closing stocks of previous year
        curr.op_statement.opening_stock_rm = round(float(prev.op_statement.closing_stock_rm or 0), 2)
        curr.op_statement.opening_stock_wip = round(float(prev.op_statement.closing_stock_wip or 0), 2)
        curr.op_statement.opening_stock_fg = round(float(prev.op_statement.closing_stock_fg or 0), 2)

        # Rules 3-6 only auto-apply when the year is NOT audited (audited = factual data)
        if curr.year_type != "audited":
            # Rule 3: Gross block carries forward (at least previous year's value)
            prev_gb = float(prev.balance_sheet.gross_block or 0)
            curr_gb = float(curr.balance_sheet.gross_block or 0)
            if curr_gb < prev_gb:
                curr.balance_sheet.gross_block = round(prev_gb, 2)

            # Rule 4: Accumulated dep = prev acc dep + prev year's depreciation expense
            prev_accdep = float(prev.balance_sheet.accumulated_depreciation or 0)
            prev_dep_exp = float(prev.op_statement.depreciation or 0)
            curr.balance_sheet.accumulated_depreciation = round(prev_accdep + prev_dep_exp, 2)

            # Rule 5: Reserves & surplus = prev R&S + prev PAT (PAT computed dynamically)
            prev_metrics = compute_year_metrics(prev)
            prev_pat = float(prev_metrics.get("pat", 0))
            prev_rs = float(prev.balance_sheet.reserves_surplus or 0)
            curr.balance_sheet.reserves_surplus = round(prev_rs + prev_pat, 2)

            # Rule 6: Capital WIP and intangibles never shrink unless explicitly written off
            prev_cwip = float(prev.balance_sheet.capital_wip or 0)
            if float(curr.balance_sheet.capital_wip or 0) < prev_cwip:
                curr.balance_sheet.capital_wip = round(prev_cwip, 2)
            prev_intg = float(prev.balance_sheet.intangible_assets or 0)
            if float(curr.balance_sheet.intangible_assets or 0) < prev_intg:
                curr.balance_sheet.intangible_assets = round(prev_intg, 2)

    return stmt.model_copy(update={"years": years})


# =============================== DPR → CMA MAPPER ===============================

def _build_cma_from_project(project: Dict[str, Any], user_id: str) -> CMAStatement:
    """Convert a DPR project dict into a pre-filled CMAStatement.

    Mapping (best-effort, user can refine in wizard):
      - business_name → company_name
      - constitution / industry → as is
      - historical_actuals (audited years) + projections (projected years) → CMAYearData
      - project_cost (sum) → gross_block (Year 1)
      - means_of_finance.term_loan → term_loan_banks
      - means_of_finance.promoter_contribution → paid_up_capital
      - working_capital.cash_required → cash_bank_balances seed
    """
    business_name = project.get("business_name", "")
    industry = project.get("industry") or project.get("business_type", "")
    constitution_map = {
        "Proprietorship": "Proprietary",
        "Partnership": "Partnership",
        "LLP": "LLP",
        "Pvt Ltd": "Private Limited",
        "Public Ltd": "Public Limited",
    }
    constitution = constitution_map.get(project.get("constitution", ""), project.get("constitution") or "Private Limited")

    # ---- Build year list ----
    years: List[CMAYearData] = []
    hist = project.get("historical_actuals") or []
    # audited years from historical
    for i, h in enumerate(hist):
        op = CMAOpStatement(
            domestic_sales=float(h.get("revenue") or 0),
            raw_materials_purchase=float(h.get("raw_material") or 0),
            direct_labour=float(h.get("salaries") or 0),
            power_fuel=float(h.get("power_fuel") or 0),
            other_mfg_expenses=float(h.get("rent") or 0),
            admin_expenses=float(h.get("other_expenses") or 0),
            depreciation=float(h.get("depreciation") or 0),
            interest_on_bank_borrowing=float(h.get("interest") or 0),
            tax_rate=25.0,
        )
        years.append(CMAYearData(
            year_label=h.get("year_label") or f"Audited Y{i+1}",
            year_type="audited",
            op_statement=op,
            balance_sheet=CMABalanceSheet(),
        ))

    # projected years from projections
    prj = project.get("projections") or []
    current_label = "FY 2024-25"
    if years:
        current_label = _next_fy_label(years[-1].year_label)
    for i, p in enumerate(prj):
        op = CMAOpStatement(
            domestic_sales=float(p.get("revenue") or 0),
            raw_materials_purchase=float(p.get("raw_material") or 0),
            direct_labour=float(p.get("salaries") or 0),
            power_fuel=float(p.get("power_fuel") or 0),
            other_mfg_expenses=float(p.get("rent") or 0),
            admin_expenses=float(p.get("other_expenses") or 0),
            depreciation=float(p.get("depreciation") or 0),
            interest_on_bank_borrowing=float(p.get("interest") or 0),
            tax_rate=float(p.get("tax_rate") or 25.0),
        )
        years.append(CMAYearData(
            year_label=current_label,
            year_type="projected",
            op_statement=op,
            balance_sheet=CMABalanceSheet(),
        ))
        current_label = _next_fy_label(current_label)

    # If we have at least one year, seed the FIRST year balance sheet with cost-of-project + finance
    if years:
        project_cost_total = sum(float(c.get("amount") or 0) for c in (project.get("project_cost") or []))
        mof = project.get("means_of_finance") or {}
        wc = project.get("working_capital") or {}
        first_bs = years[0].balance_sheet
        first_bs.gross_block = project_cost_total
        first_bs.term_loan_banks = float(mof.get("term_loan") or 0)
        first_bs.short_term_borrowing_banks = float(mof.get("working_capital_loan") or 0)
        first_bs.paid_up_capital = float(mof.get("promoter_contribution") or 0)
        first_bs.cash_bank_balances = float(wc.get("cash_required") or 0)
        years[0].balance_sheet = first_bs

    # ---- Existing limits seed from means_of_finance ----
    existing_limits: List[CMAExistingLimits] = []
    mof = project.get("means_of_finance") or {}
    if mof.get("term_loan"):
        existing_limits.append(CMAExistingLimits(
            facility="Term Loan",
            bank=project.get("loan_scheme", ""),
            nature="Fund based",
            proposed_limit=float(mof.get("term_loan") or 0),
            rate_of_interest=float(mof.get("interest_rate") or 0),
        ))
    if mof.get("working_capital_loan"):
        existing_limits.append(CMAExistingLimits(
            facility="Cash Credit",
            bank=project.get("loan_scheme", ""),
            nature="Fund based",
            proposed_limit=float(mof.get("working_capital_loan") or 0),
            rate_of_interest=float(mof.get("interest_rate") or 0),
        ))

    # ---- Applicant address ----
    applicant = project.get("applicant") or {}
    addr_parts = [applicant.get("address_line1", ""), applicant.get("address_line2", ""),
                  applicant.get("city", ""), applicant.get("state", ""), applicant.get("pincode", "")]
    address = ", ".join([p for p in addr_parts if p]).strip(", ")
    promoters_list = project.get("promoters") or []
    promoters_text = "; ".join([p.get("name", "") for p in promoters_list if p.get("name")])

    return CMAStatement(
        user_id=user_id,
        company_name=business_name,
        constitution=constitution,
        industry=industry,
        business_activity=industry,
        registered_address=address,
        pan=applicant.get("pan", ""),
        promoters=promoters_text,
        banker_name="",
        facility_type="Composite" if (mof.get("term_loan") and mof.get("working_capital_loan")) else ("TL" if mof.get("term_loan") else "CC"),
        purpose=f"Generated from DPR — {business_name}",
        years=years,
        existing_limits=existing_limits,
        source_dpr_project_id=project.get("project_id", ""),
    )


def register_cma_routes(api_router: APIRouter, db, get_current_user, User):
    """Mount all CMA endpoints onto the provided api_router."""

    # ---------- Pricing config (read from `db.settings` if present) ----------
    SETTINGS_ID = "company_settings"
    DEFAULT_CMA_PRICING = {
        "cma_pricing_mode": "free",       # "free" | "paid"
        "cma_price_inr": 499,             # one-time price for paid mode
        "cma_promo_banner": "🎉 LIMITED-TIME: CMA module is FREE — download unlimited bank-ready CMA Excel + PDF. Tell your CA friends!",
    }

    async def _get_cma_pricing() -> Dict[str, Any]:
        doc = await db.settings.find_one({"_id": SETTINGS_ID}, {"_id": 0}) or {}
        return {
            "mode": doc.get("cma_pricing_mode") or DEFAULT_CMA_PRICING["cma_pricing_mode"],
            "price_inr": int(doc.get("cma_price_inr") or DEFAULT_CMA_PRICING["cma_price_inr"]),
            "promo_banner": doc.get("cma_promo_banner") or DEFAULT_CMA_PRICING["cma_promo_banner"],
        }

    async def _get_seller_info(_db) -> Dict[str, Any]:
        """Build the seller dict expected by invoice_module from company settings."""
        # Static defaults (mirror server.py COMPANY_INFO)
        seller = {
            "name": "Mother Bless Digital Solutions",
            "address_line1": "Shop No. 32, Above State Bank ATM",
            "address_line2": "Sagwara Road",
            "city": "Bagidora",
            "state": "Rajasthan",
            "pincode": "314035",
            "country": "India",
            "primary_phone": "7300213623",
            "gstin": "08KQRPS8229A1Z6",
            "email": "motherblessopc@gmail.com",
        }
        doc = await _db.settings.find_one({"_id": SETTINGS_ID}, {"_id": 0}) or {}
        for k in ("name", "address_line1", "address_line2", "city", "state", "pincode", "country", "primary_phone", "gstin", "email"):
            if doc.get(k):
                seller[k] = doc[k]
        return seller

    async def _ensure_can_download(stmt: CMAStatement, user) -> None:
        pricing = await _get_cma_pricing()
        # Free mode → everyone can download
        if pricing["mode"] == "free":
            return
        # Admins always
        if getattr(user, "is_admin", False):
            return
        # Paid mode → CMA must be marked paid
        if stmt.payment_status != "paid":
            raise HTTPException(
                status_code=402,
                detail=f"CMA is in PAID mode (₹{pricing['price_inr']}). Please complete payment to download.",
            )

    async def _load(cma_id: str, user) -> CMAStatement:
        doc = await db.cma_statements.find_one({"cma_id": cma_id, "user_id": user.user_id}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="CMA not found")
        # Coerce datetime fields
        for k in ("created_at", "updated_at", "paid_at"):
            if isinstance(doc.get(k), str):
                try:
                    doc[k] = datetime.fromisoformat(doc[k])
                except (ValueError, TypeError):
                    doc[k] = None
        return CMAStatement(**doc)

    @api_router.get("/cma/statements", response_model=List[CMAStatement])
    async def list_statements(user=Depends(get_current_user)):
        cursor = db.cma_statements.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1)
        items = []
        async for doc in cursor:
            for k in ("created_at", "updated_at"):
                if isinstance(doc.get(k), str):
                    try:
                        doc[k] = datetime.fromisoformat(doc[k])
                    except (ValueError, TypeError):
                        doc[k] = datetime.now(timezone.utc)
            items.append(CMAStatement(**doc))
        return items

    @api_router.post("/cma/statements", response_model=CMAStatement)
    async def create_statement(payload: CMACreate, user=Depends(get_current_user)):
        years = _scaffold_years(payload)
        stmt = CMAStatement(
            user_id=user.user_id,
            company_name=payload.company_name,
            constitution=payload.constitution,
            industry=payload.industry,
            facility_type=payload.facility_type,
            banker_name=payload.banker_name,
            years=years,
        )
        await db.cma_statements.insert_one(stmt.model_dump())
        return stmt

    @api_router.get("/cma/statements/{cma_id}", response_model=CMAStatement)
    async def get_statement(cma_id: str, user=Depends(get_current_user)):
        return await _load(cma_id, user)

    @api_router.put("/cma/statements/{cma_id}", response_model=CMAStatement)
    async def update_statement(cma_id: str, payload: CMAUpdate, user=Depends(get_current_user)):
        existing = await _load(cma_id, user)
        update = payload.model_dump(exclude_unset=True)
        merged = existing.model_dump()
        merged.update(update)
        merged["updated_at"] = datetime.now(timezone.utc)
        new_stmt = CMAStatement(**merged)
        # Auto-apply carry-forward rules on every save:
        #   • Year N closing stock (RM/WIP/FG) → Year N+1 opening stock
        #   • Year N gross block → Year N+1 (at least; user can add capex)
        #   • Year N accumulated dep + Year N depreciation → Year N+1 acc dep
        #   • Year N R&S + Year N PAT → Year N+1 R&S
        # Year 1 values are user-entered and never overwritten.
        new_stmt = _reconcile_carry_forward(new_stmt)
        new_stmt.updated_at = datetime.now(timezone.utc)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": new_stmt.model_dump()},
        )
        return new_stmt

    @api_router.delete("/cma/statements/{cma_id}")
    async def delete_statement(cma_id: str, user=Depends(get_current_user)):
        result = await db.cma_statements.delete_one({"cma_id": cma_id, "user_id": user.user_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="CMA not found")
        return {"ok": True}

    @api_router.get("/cma/statements/{cma_id}/analysis")
    async def analysis(cma_id: str, user=Depends(get_current_user)):
        stmt = await _load(cma_id, user)
        return build_analysis(stmt)

    @api_router.get("/cma/statements/{cma_id}/download/excel")
    async def download_excel(cma_id: str, user=Depends(get_current_user)):
        stmt = await _load(cma_id, user)
        await _ensure_can_download(stmt, user)
        content = generate_cma_excel(stmt)
        safe_name = (stmt.company_name or "CMA").replace(" ", "_").replace("/", "_")[:40]
        filename = f"CMA_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @api_router.get("/cma/statements/{cma_id}/download/pdf")
    async def download_pdf(cma_id: str, user=Depends(get_current_user)):
        stmt = await _load(cma_id, user)
        await _ensure_can_download(stmt, user)
        seller = await _get_seller_info(db)
        content = generate_cma_pdf(stmt, seller=seller)
        safe_name = (stmt.company_name or "CMA").replace(" ", "_").replace("/", "_")[:40]
        filename = f"CMA_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @api_router.get("/cma/template/blank.xlsx")
    async def blank_template(audited: int = 2, provisional: int = 1, projected: int = 3,
                             user=Depends(get_current_user)):
        content = generate_blank_template(audited, provisional, projected)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="CMA_Blank_Template.xlsx"'},
        )

    @api_router.post("/cma/statements/{cma_id}/import-excel", response_model=CMAStatement)
    async def import_excel(cma_id: str, file: UploadFile = File(...),
                           user=Depends(get_current_user)):
        existing = await _load(cma_id, user)
        content = await file.read()
        try:
            parsed = parse_uploaded_excel(content)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

        merged = existing.model_dump()
        if parsed.get("years"):
            merged["years"] = parsed["years"]
        if parsed.get("existing_limits"):
            merged["existing_limits"] = parsed["existing_limits"]
        merged["updated_at"] = datetime.now(timezone.utc)
        new_stmt = CMAStatement(**merged)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": new_stmt.model_dump()},
        )
        return new_stmt

    # =============================== AUTO-PROJECT ===============================

    @api_router.post("/cma/statements/{cma_id}/auto-project", response_model=CMAStatement)
    async def auto_project(cma_id: str, payload: Dict[str, Any] = None,
                           user=Depends(get_current_user)):
        """Auto-fill provisional + projected years using a YoY growth rate derived from audited years."""
        payload = payload or {}
        stmt = await _load(cma_id, user)
        new_stmt, info = _apply_auto_projections(stmt, payload)
        # Always reconcile carry-forward after auto-project (closing→opening, gross block, R&S etc.)
        new_stmt = _reconcile_carry_forward(new_stmt)
        new_stmt.updated_at = datetime.now(timezone.utc)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": new_stmt.model_dump()},
        )
        return new_stmt

    @api_router.post("/cma/statements/{cma_id}/reconcile", response_model=CMAStatement)
    async def reconcile(cma_id: str, user=Depends(get_current_user)):
        """Apply CMA carry-forward rules: closing→opening stock, gross block carry-forward, accumulated
        depreciation, reserves & surplus reconciliation. Audited-year balance-sheet items are preserved."""
        stmt = await _load(cma_id, user)
        new_stmt = _reconcile_carry_forward(stmt)
        new_stmt.updated_at = datetime.now(timezone.utc)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": new_stmt.model_dump()},
        )
        return new_stmt

    @api_router.get("/cma/statements/{cma_id}/growth-hint")
    async def growth_hint(cma_id: str, user=Depends(get_current_user)):
        """Return the auto-detected YoY growth % from audited years (and other hints)."""
        stmt = await _load(cma_id, user)
        return _detect_growth(stmt)

    # =============================== PRICING ===============================

    @api_router.get("/cma/pricing")
    async def cma_pricing(user=Depends(get_current_user)):
        """Public (auth-required) endpoint for the frontend to read current pricing mode."""
        return await _get_cma_pricing()

    @api_router.get("/admin/cma/pricing")
    async def admin_get_cma_pricing(user=Depends(get_current_user)):
        if not getattr(user, "is_admin", False):
            raise HTTPException(status_code=403, detail="Admin only")
        return await _get_cma_pricing()

    @api_router.post("/admin/cma/pricing")
    async def admin_update_cma_pricing(payload: Dict[str, Any], user=Depends(get_current_user)):
        if not getattr(user, "is_admin", False):
            raise HTTPException(status_code=403, detail="Admin only")
        update = {}
        if "mode" in payload:
            mode = str(payload["mode"]).lower()
            if mode not in ("free", "paid"):
                raise HTTPException(status_code=400, detail="mode must be 'free' or 'paid'")
            update["cma_pricing_mode"] = mode
        if "price_inr" in payload:
            try:
                p = int(payload["price_inr"])
                if p <= 0:
                    raise ValueError
                update["cma_price_inr"] = p
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="price_inr must be a positive integer")
        if "promo_banner" in payload:
            update["cma_promo_banner"] = str(payload["promo_banner"])
        if not update:
            raise HTTPException(status_code=400, detail="Nothing to update")
        update["updated_at"] = datetime.now(timezone.utc).isoformat()
        update["updated_by"] = user.email
        await db.settings.update_one({"_id": SETTINGS_ID}, {"$set": update}, upsert=True)
        return await _get_cma_pricing()

    @api_router.post("/admin/cma/{cma_id}/mark-paid")
    async def admin_mark_cma_paid(cma_id: str, user=Depends(get_current_user)):
        """Admin: manually mark a CMA as paid (UPI / Bank transfer flow)."""
        if not getattr(user, "is_admin", False):
            raise HTTPException(status_code=403, detail="Admin only")
        doc = await db.cma_statements.find_one({"cma_id": cma_id}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="CMA not found")
        await db.cma_statements.update_one(
            {"cma_id": cma_id},
            {"$set": {
                "payment_status": "paid",
                "paid_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        return {"ok": True, "cma_id": cma_id, "payment_status": "paid"}

    # =============================== SELF-SERVE PAYMENT ===============================

    async def _compute_cma_price(user, pricing: Dict[str, Any]) -> Dict[str, Any]:
        """Return {base_price, final_price, founding_discount_pct, is_free, is_founding_ca}."""
        base = int(pricing.get("price_inr", 499))
        is_free = pricing.get("mode") == "free"
        # Re-fetch user to get latest is_founding_ca flag
        u = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "is_founding_ca": 1, "wallet_balance": 1})
        is_founding_ca = bool((u or {}).get("is_founding_ca", False))
        discount_pct = 50 if is_founding_ca else 0
        final = 0 if is_free else max(0, int(round(base * (1 - discount_pct / 100.0))))
        return {
            "is_free": is_free,
            "base_price_inr": base,
            "final_price_inr": final,
            "founding_discount_pct": discount_pct,
            "is_founding_ca": is_founding_ca,
            "wallet_balance": float((u or {}).get("wallet_balance", 0)),
            "promo_banner": pricing.get("promo_banner", ""),
            "mode": pricing.get("mode"),
        }

    @api_router.get("/cma/statements/{cma_id}/pricing")
    async def cma_statement_pricing(cma_id: str, user=Depends(get_current_user)):
        """Per-CMA pricing for the current user (applies founding-CA discount)."""
        stmt = await _load(cma_id, user)
        pricing = await _get_cma_pricing()
        price = await _compute_cma_price(user, pricing)
        price["cma_id"] = cma_id
        price["payment_status"] = stmt.payment_status
        return price

    @api_router.post("/cma/statements/{cma_id}/submit-payment", response_model=CMAStatement)
    async def cma_submit_payment(cma_id: str, payload: CMAPaymentSubmit, user=Depends(get_current_user)):
        """User submits UPI/Bank txn ID — payment_status moves to 'submitted', admin verifies later."""
        stmt = await _load(cma_id, user)
        if stmt.payment_status == "paid":
            raise HTTPException(status_code=400, detail="Already paid")
        pricing = await _get_cma_pricing()
        price = await _compute_cma_price(user, pricing)
        if price["is_free"]:
            raise HTTPException(status_code=400, detail="CMA module is currently FREE — no payment needed.")
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": {
                "payment_status": "submitted",
                "payment_txn_id": payload.txn_id,
                "payment_amount": float(payload.amount or price["final_price_inr"]),
                "payment_method": payload.method,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        # Also log to admin payments queue (same collection used by DPR)
        await db.payments.insert_one({
            "log_id": str(uuid.uuid4()),
            "kind": "cma",
            "cma_id": cma_id,
            "user_id": user.user_id,
            "user_email": user.email,
            "amount": float(payload.amount or price["final_price_inr"]),
            "method": payload.method,
            "txn_id": payload.txn_id,
            "verification_status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        # Auto-generate GST tax invoice (idempotent)
        try:
            from invoice_module import create_invoice_for_payment
            seller = await _get_seller_info(db)
            user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0}) or {
                "user_id": user.user_id, "email": user.email, "name": getattr(user, "name", ""),
            }
            await create_invoice_for_payment(
                db, kind="cma", ref_id=cma_id, user=user_doc,
                amount_paid=float(payload.amount or price["final_price_inr"]),
                payment_method=payload.method,
                payment_txn_id=payload.txn_id,
                seller=seller,
                buyer_name=stmt.company_name or user_doc.get("name"),
                buyer_address=stmt.registered_address or "",
            )
        except Exception as _inv_err:
            import logging as _lg
            _lg.getLogger(__name__).warning(f"[invoice] CMA submit-payment invoice failed: {_inv_err}")
        return await _load(cma_id, user)

    @api_router.post("/cma/statements/{cma_id}/pay-from-wallet", response_model=CMAStatement)
    async def cma_pay_from_wallet(cma_id: str, user=Depends(get_current_user)):
        """Deduct CMA price from wallet & mark paid instantly."""
        stmt = await _load(cma_id, user)
        if stmt.payment_status == "paid":
            raise HTTPException(status_code=400, detail="Already paid")
        pricing = await _get_cma_pricing()
        price = await _compute_cma_price(user, pricing)
        if price["is_free"]:
            raise HTTPException(status_code=400, detail="CMA module is currently FREE — no payment needed.")
        if price["wallet_balance"] < price["final_price_inr"]:
            raise HTTPException(
                status_code=402,
                detail=f"Insufficient wallet balance. Need ₹{price['final_price_inr']}, you have ₹{price['wallet_balance']:.0f}",
            )
        now = datetime.now(timezone.utc)
        await db.users.update_one(
            {"user_id": user.user_id},
            {"$inc": {"wallet_balance": -price["final_price_inr"]}},
        )
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": {
                "payment_status": "paid",
                "payment_txn_id": f"CMA-WALLET-{uuid.uuid4().hex[:8].upper()}",
                "payment_amount": float(price["final_price_inr"]),
                "payment_method": "Wallet",
                "paid_at": now.isoformat(),
                "updated_at": now.isoformat(),
            }},
        )
        await db.wallet_txns.insert_one({
            "txn_uid": str(uuid.uuid4()),
            "user_id": user.user_id,
            "user_email": user.email,
            "type": "debit",
            "amount": float(price["final_price_inr"]),
            "cma_id": cma_id,
            "status": "verified",
            "created_at": now.isoformat(),
        })
        # Auto-generate GST tax invoice (idempotent)
        try:
            from invoice_module import create_invoice_for_payment
            seller = await _get_seller_info(db)
            user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0}) or {
                "user_id": user.user_id, "email": user.email, "name": getattr(user, "name", ""),
            }
            await create_invoice_for_payment(
                db, kind="cma", ref_id=cma_id, user=user_doc,
                amount_paid=float(price["final_price_inr"]),
                payment_method="Wallet",
                payment_txn_id=f"CMA-WALLET-{cma_id[:8]}",
                seller=seller,
                buyer_name=stmt.company_name or user_doc.get("name"),
                buyer_address=stmt.registered_address or "",
            )
        except Exception as _inv_err:
            import logging as _lg
            _lg.getLogger(__name__).warning(f"[invoice] CMA wallet-pay invoice failed: {_inv_err}")
        return await _load(cma_id, user)

    @api_router.post("/admin/cma/payments/{log_id}/verify")
    async def admin_verify_cma_payment(log_id: str, user=Depends(get_current_user)):
        if not getattr(user, "is_admin", False):
            raise HTTPException(status_code=403, detail="Admin only")
        log = await db.payments.find_one({"log_id": log_id}, {"_id": 0})
        if not log or log.get("kind") != "cma":
            raise HTTPException(status_code=404, detail="CMA payment log not found")
        await db.payments.update_one({"log_id": log_id}, {"$set": {"verification_status": "verified", "verified_at": datetime.now(timezone.utc).isoformat()}})
        await db.cma_statements.update_one(
            {"cma_id": log["cma_id"]},
            {"$set": {"payment_status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}},
        )
        return {"ok": True, "cma_id": log["cma_id"]}

    @api_router.post("/cma/statements/{cma_id}/razorpay-order")
    async def cma_razorpay_order(cma_id: str, user=Depends(get_current_user)):
        """Create a Razorpay order for CMA (uses the same admin-configured keys as DPR)."""
        stmt = await _load(cma_id, user)
        if stmt.payment_status == "paid":
            raise HTTPException(status_code=400, detail="Already paid")
        # Reuse company settings (DPR's razorpay_*)
        settings_doc = await db.settings.find_one({"_id": SETTINGS_ID}, {"_id": 0}) or {}
        key_id = settings_doc.get("razorpay_key_id", "")
        key_secret = settings_doc.get("razorpay_key_secret", "")
        enabled = bool(settings_doc.get("razorpay_enabled", False))
        if not enabled or not key_id or not key_secret:
            raise HTTPException(status_code=400, detail="Razorpay is not configured. Please use UPI / Wallet flow, or ask admin to enable Razorpay.")
        pricing = await _get_cma_pricing()
        price = await _compute_cma_price(user, pricing)
        if price["is_free"]:
            raise HTTPException(status_code=400, detail="CMA is currently FREE — no payment needed.")
        try:
            import razorpay  # type: ignore
            client = razorpay.Client(auth=(key_id, key_secret))
            order = client.order.create({
                "amount": int(price["final_price_inr"] * 100),  # paise
                "currency": "INR",
                "receipt": f"cma_{cma_id[:12]}",
                "payment_capture": 1,
            })
            return {
                "order_id": order["id"],
                "key_id": key_id,
                "amount": order["amount"],
                "currency": order["currency"],
                "company_name": settings_doc.get("name", "DPRForge"),
                "cma_id": cma_id,
            }
        except ImportError:
            raise HTTPException(status_code=500, detail="razorpay SDK not installed on server")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Razorpay order creation failed: {e}")

    @api_router.post("/cma/statements/{cma_id}/razorpay-verify", response_model=CMAStatement)
    async def cma_razorpay_verify(cma_id: str, payload: CMARazorpayVerify, user=Depends(get_current_user)):
        stmt = await _load(cma_id, user)
        settings_doc = await db.settings.find_one({"_id": SETTINGS_ID}, {"_id": 0}) or {}
        key_id = settings_doc.get("razorpay_key_id", "")
        key_secret = settings_doc.get("razorpay_key_secret", "")
        if not key_id or not key_secret:
            raise HTTPException(status_code=400, detail="Razorpay not configured")
        try:
            import razorpay  # type: ignore
            client = razorpay.Client(auth=(key_id, key_secret))
            client.utility.verify_payment_signature({
                "razorpay_order_id": payload.razorpay_order_id,
                "razorpay_payment_id": payload.razorpay_payment_id,
                "razorpay_signature": payload.razorpay_signature,
            })
        except ImportError:
            raise HTTPException(status_code=500, detail="razorpay SDK not installed")
        except Exception:
            raise HTTPException(status_code=400, detail="Signature verification failed")
        now = datetime.now(timezone.utc)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": {
                "payment_status": "paid",
                "payment_txn_id": payload.razorpay_payment_id,
                "payment_method": "Razorpay",
                "paid_at": now.isoformat(),
                "updated_at": now.isoformat(),
            }},
        )
        return await _load(cma_id, user)

    # =============================== AI GROWTH SUGGEST ===============================

    @api_router.post("/cma/statements/{cma_id}/ai-suggest-growth")
    async def ai_suggest_growth(cma_id: str, user=Depends(get_current_user)):
        """Ask GPT-5.2 for per-line-item growth % based on the user's audited years."""
        stmt = await _load(cma_id, user)
        audited = [y for y in stmt.years if y.year_type == "audited"]
        if len(audited) < 2:
            raise HTTPException(status_code=400, detail="Need at least 2 audited years for AI suggestions")
        try:
            return await _llm_growth_suggestion(stmt, audited)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"AI suggestion failed: {e}")

    @api_router.post("/cma/statements/{cma_id}/apply-ai-growth", response_model=CMAStatement)
    async def apply_ai_growth(cma_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
        """Apply user-confirmed per-field growth rates returned by /ai-suggest-growth (user can edit before applying)."""
        stmt = await _load(cma_id, user)
        rates: Dict[str, float] = payload.get("field_growth_pct") or {}
        if not rates:
            raise HTTPException(status_code=400, detail="field_growth_pct dict required")
        new_stmt = _apply_per_field_growth(stmt, rates)
        new_stmt = _reconcile_carry_forward(new_stmt)
        new_stmt.updated_at = datetime.now(timezone.utc)
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": new_stmt.model_dump()},
        )
        return new_stmt

    # =============================== AI ASSISTANT (summary + chat) ===============================

    @api_router.post("/cma/statements/{cma_id}/ai-summary")
    async def cma_ai_summary(cma_id: str, user=Depends(get_current_user)):
        """Generate a 1-page bank-facing narrative summary using GPT-5.2."""
        stmt = await _load(cma_id, user)
        try:
            text = await _llm_cma_summary(stmt)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"AI summary failed: {e}")
        await db.cma_statements.update_one(
            {"cma_id": cma_id, "user_id": user.user_id},
            {"$set": {"ai_summary": text, "updated_at": datetime.now(timezone.utc).isoformat()}},
        )
        return {"cma_id": cma_id, "summary": text, "source": "gpt-5.2"}

    @api_router.post("/cma/statements/{cma_id}/ai-chat")
    async def cma_ai_chat(cma_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
        """Conversational assistant — user asks a question about THIS CMA, GPT-5.2 answers with data context."""
        question = (payload or {}).get("question", "").strip()
        if not question:
            raise HTTPException(status_code=400, detail="question required")
        stmt = await _load(cma_id, user)
        try:
            answer = await _llm_cma_chat(stmt, question)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"AI chat failed: {e}")
        return {"cma_id": cma_id, "question": question, "answer": answer, "source": "gpt-5.2"}

    # =============================== BENCHMARKS ===============================

    @api_router.get("/cma/benchmarks/{industry}")
    async def get_benchmarks(industry: str, user=Depends(get_current_user)):
        return _lookup_benchmark(industry)

    @api_router.post("/cma/benchmarks/ai-refine")
    async def ai_refine_benchmarks(payload: Dict[str, Any], user=Depends(get_current_user)):
        industry = (payload or {}).get("industry", "")
        if not industry:
            raise HTTPException(status_code=400, detail="industry is required")
        try:
            return await _llm_benchmark_refine(industry)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"AI refine failed: {e}")

    # =============================== GENERATE FROM DPR ===============================

    @api_router.post("/cma/from-dpr/{project_id}", response_model=CMAStatement)
    async def generate_from_dpr(project_id: str, user=Depends(get_current_user)):
        """Create a new CMA pre-filled from an existing DPR project owned by the user."""
        project = await db.projects.find_one({"project_id": project_id, "user_id": user.user_id}, {"_id": 0})
        if not project:
            raise HTTPException(status_code=404, detail="DPR project not found")

        stmt = _build_cma_from_project(project, user_id=user.user_id)
        await db.cma_statements.insert_one(stmt.model_dump())
        return stmt
